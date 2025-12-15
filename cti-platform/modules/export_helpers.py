"""
Odysafe CTI Platform
Copyright (C) 2025 Bastien GUIDONE

This program is free software: you can redistribute it and/or modify
it under the terms of the GNU Affero General Public License as published by
the Free Software Foundation, either version 3 of the License, or
(at your option) any later version.

This program is distributed in the hope that it will be useful,
but WITHOUT ANY WARRANTY; without even the implied warranty of
MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
GNU Affero General Public License for more details.

You should have received a copy of the GNU Affero General Public License
along with this program.  If not, see <https://www.gnu.org/licenses/>.

Helper functions for IOC exports to eliminate code duplication
"""
import csv
import json
import logging
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Optional, Iterator, Tuple
from flask import send_file

logger = logging.getLogger(__name__)


def get_iocs_for_export(db, data: Dict) -> Tuple[List[Dict], Optional[str]]:
    """
    Unified function to retrieve IOCs for export based on filters.
    Returns (iocs_list, source_context)
    """
    source_ids = data.get('source_ids', [])
    group_ids = data.get('group_ids', [])
    ioc_ids = data.get('ioc_ids', [])
    ioc_types = data.get('ioc_types', [])  # NEW: Filter by IOC types
    
    source_context = ""
    iocs = []
    
    # Build filters dict for database queries
    filters = {}
    if ioc_types:
        filters['ioc_types'] = ioc_types
    
    # If group_ids is provided, retrieve corresponding source_ids
    if group_ids:
        all_source_ids = []
        for group_id in group_ids:
            group_source_ids = db.get_sources_by_group(group_id)
            all_source_ids.extend(group_source_ids)
        source_ids = list(set(all_source_ids))  # Remove duplicates
    
    # Retrieve IOCs
    if source_ids:
        contexts = []
        for source_id in source_ids:
            source = db.get_source(source_id)
            if source and source.get('context'):
                contexts.append(source['context'])
            source_iocs = db.get_iocs_by_source(source_id)
            # Apply IOC type filter if specified
            if ioc_types:
                source_iocs = [ioc for ioc in source_iocs if ioc.get('ioc_type') in ioc_types]
            iocs.extend(source_iocs)
        
        # Combine contexts if multiple sources
        if contexts:
            source_context = "\n\n---\n\n".join(contexts)
    elif ioc_ids:
        for ioc_id in ioc_ids:
            ioc = db.get_ioc(ioc_id)
            if ioc:
                # Apply IOC type filter if specified
                if not ioc_types or ioc.get('ioc_type') in ioc_types:
                    iocs.append(ioc)
                # Get source context from first IOC
                if not source_context:
                    source_info = db.get_source(ioc.get('source_id'))
                    if source_info:
                        source_context = source_info.get('context', '')
    else:
        # Get all IOCs with streaming for large datasets
        # Collect all IOCs from streaming iterator with filters
        iocs = []
        for batch in db.get_all_iocs_streaming(filters=filters if filters else None, limit=None):
            iocs.extend(batch)
    
    # Deduplicate IOCs by type and value (keep first occurrence)
    # This ensures only unique IOCs are exported, regardless of source
    # This is especially important for STIX export to avoid duplicate objects
    seen_iocs = {}
    unique_iocs = []
    for ioc in iocs:
        ioc_type = ioc.get('ioc_type', '').lower().strip()
        ioc_value = ioc.get('ioc_value', '').strip()
        if ioc_value:
            # Create unique key: type + normalized value (case-insensitive)
            key = (ioc_type, ioc_value.lower())
            if key not in seen_iocs:
                seen_iocs[key] = True
                unique_iocs.append(ioc)
    
    if len(unique_iocs) < len(iocs):
        logger.info(f"Export deduplication: {len(iocs)} -> {len(unique_iocs)} unique IOCs")
    
    return unique_iocs, source_context


def export_txt(db, data: Dict, output_folder: Path) -> Path:
    """Export IOCs to TXT format with types"""
    iocs, _ = get_iocs_for_export(db, data)
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_file = output_folder / "iocs" / f"export_{timestamp}.txt"
    output_file.parent.mkdir(parents=True, exist_ok=True)
    
    with open(output_file, 'w', encoding='utf-8') as f:
        for ioc in iocs:
            f.write(f"{ioc['ioc_type']}\t{ioc['ioc_value']}\n")
    
    return output_file


def export_txt_simple(db, data: Dict, output_folder: Path) -> Path:
    """Export IOCs to TXT format - values only (one per line) for firewall/EDR compatibility"""
    iocs, _ = get_iocs_for_export(db, data)
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_file = output_folder / "iocs" / f"export_simple_{timestamp}.txt"
    output_file.parent.mkdir(parents=True, exist_ok=True)
    
    with open(output_file, 'w', encoding='utf-8') as f:
        for ioc in iocs:
            f.write(f"{ioc['ioc_value']}\n")
    
    return output_file


def export_json(db, data: Dict, output_folder: Path) -> Path:
    """Export IOCs to JSON format with metadata"""
    iocs, _ = get_iocs_for_export(db, data)
    
    # Enrich IOCs with source and tag information
    source_ids_unique = list(set(ioc['source_id'] for ioc in iocs))
    # Build sources map efficiently without creating intermediate list
    sources_map = {}
    for sid in source_ids_unique:
        source = db.get_source(sid)
        if source:
            sources_map[source['id']] = source
    
    for ioc in iocs:
        source = sources_map.get(ioc['source_id'])
        if source:
            ioc['source'] = source
        # Get tags for this IOC
        full_ioc = db.get_ioc(ioc['id'])
        if full_ioc and full_ioc.get('tags'):
            ioc['tags'] = full_ioc['tags']
        else:
            ioc['tags'] = []
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_file = output_folder / "iocs" / f"export_{timestamp}.json"
    output_file.parent.mkdir(parents=True, exist_ok=True)
    
    export_data = {
        'export_date': datetime.now().isoformat(),
        'total_iocs': len(iocs),
        'iocs': iocs
    }
    
    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(export_data, f, indent=2, ensure_ascii=False, default=str)
    
    return output_file


def export_csv(db, data: Dict, output_folder: Path) -> Path:
    """Export IOCs to CSV format"""
    iocs, _ = get_iocs_for_export(db, data)
    
    # Enrich IOCs with tags if needed
    for ioc in iocs:
        if 'tags' not in ioc:
            full_ioc = db.get_ioc(ioc['id'])
            if full_ioc and full_ioc.get('tags'):
                ioc['tags'] = full_ioc['tags']
            else:
                ioc['tags'] = []
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_file = output_folder / "iocs" / f"export_{timestamp}.csv"
    output_file.parent.mkdir(parents=True, exist_ok=True)
    
    with open(output_file, 'w', encoding='utf-8', newline='') as f:
        writer = csv.writer(f)
        # Write header
        writer.writerow(['Type', 'Value', 'Source', 'Created At', 'Tags'])
        
        # Write IOCs
        for ioc in iocs:
            source = db.get_source(ioc.get('source_id'))
            source_name = source.get('name', '') if source else ioc.get('source_name', '')
            
            # Handle tags - can be list of strings or list of dicts
            tags_str = ''
            tags = ioc.get('tags', [])
            if tags:
                if isinstance(tags, list):
                    if tags and isinstance(tags[0], dict):
                        tags_str = ', '.join([tag.get('name', '') for tag in tags if tag.get('name')])
                    else:
                        tags_str = ', '.join([str(tag) for tag in tags if tag])
            
            writer.writerow([
                ioc.get('ioc_type', ''),
                ioc.get('ioc_value', ''),
                source_name,
                ioc.get('created_at', ''),
                tags_str
            ])
    
    return output_file


def export_csv_firewall(db, data: Dict, output_folder: Path) -> Path:
    """Export IOCs to CSV format - simplified for firewall/EDR compatibility"""
    iocs, _ = get_iocs_for_export(db, data)
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_file = output_folder / "iocs" / f"export_firewall_{timestamp}.csv"
    output_file.parent.mkdir(parents=True, exist_ok=True)
    
    with open(output_file, 'w', encoding='utf-8', newline='') as f:
        writer = csv.writer(f)
        # Write header - simple format with just value
        writer.writerow(['value'])
        
        # Write IOCs - just the values
        for ioc in iocs:
            writer.writerow([ioc.get('ioc_value', '')])
    
    return output_file


def export_json_simple(db, data: Dict, output_folder: Path) -> Path:
    """Export IOCs to JSON format - simplified structure grouped by type"""
    iocs, _ = get_iocs_for_export(db, data)
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_file = output_folder / "iocs" / f"export_simple_{timestamp}.json"
    output_file.parent.mkdir(parents=True, exist_ok=True)
    
    # Group IOCs by type
    iocs_by_type = {}
    all_values = []
    
    for ioc in iocs:
        ioc_type = ioc.get('ioc_type', 'unknown')
        ioc_value = ioc.get('ioc_value', '')
        
        if ioc_type not in iocs_by_type:
            iocs_by_type[ioc_type] = []
        iocs_by_type[ioc_type].append(ioc_value)
        all_values.append(ioc_value)
    
    # Create export structure
    export_data = {
        'export_date': datetime.now().isoformat(),
        'total_iocs': len(iocs),
        'iocs': all_values,  # Simple list of all values
        'iocs_by_type': iocs_by_type  # Grouped by type
    }
    
    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(export_data, f, indent=2, ensure_ascii=False, default=str)
    
    return output_file


def export_xlsx(db, data: Dict, output_folder: Path) -> Path:
    """Export IOCs to XLSX format with elegant formatting"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise RuntimeError("openpyxl is not available. Please install it: pip install openpyxl")
    
    iocs, _ = get_iocs_for_export(db, data)
    
    # Enrich IOCs with tags if needed
    for ioc in iocs:
        if 'tags' not in ioc:
            full_ioc = db.get_ioc(ioc['id'])
            if full_ioc and full_ioc.get('tags'):
                ioc['tags'] = full_ioc['tags']
            else:
                ioc['tags'] = []
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_file = output_folder / "iocs" / f"export_{timestamp}.xlsx"
    output_file.parent.mkdir(parents=True, exist_ok=True)
    
    # Create workbook
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    # Color palette (elegant and modern)
    colors = {
        'header_bg': 'F5F5F7',  # Light gray
        'header_text': '1D1D1F',  # Dark gray
        'row_bg_light': 'FFFFFF',  # White
        'row_bg_alt': 'FAFAFA',  # Very light gray
        'border': 'E5E5E7',  # Light border
        'accent': '8B5CF6',  # Purple accent
        'accent_light': 'E9D5FF',  # Light purple
        'network_accent': 'D1FAE5',  # Light green for network IOCs
        'file_accent': 'FEE2E2',  # Light red for file hashes
        'summary_bg': 'F9FAFB',  # Very light gray for summary
    }
    
    # Style definitions
    header_font = Font(name='Calibri', size=11, bold=True, color=colors['header_text'])
    cell_font = Font(name='Calibri', size=10, color='1D1D1F')
    header_fill = PatternFill(start_color=colors['header_bg'], end_color=colors['header_bg'], fill_type='solid')
    network_fill = PatternFill(start_color=colors['network_accent'], end_color=colors['network_accent'], fill_type='solid')
    file_fill = PatternFill(start_color=colors['file_accent'], end_color=colors['file_accent'], fill_type='solid')
    accent_fill = PatternFill(start_color=colors['accent_light'], end_color=colors['accent_light'], fill_type='solid')
    border_style = Border(
        left=Side(style='thin', color=colors['border']),
        right=Side(style='thin', color=colors['border']),
        top=Side(style='thin', color=colors['border']),
        bottom=Side(style='thin', color=colors['border'])
    )
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    # Sheet 1: All IOCs
    ws_all = wb.create_sheet("All IOCs", 0)
    headers = ['Type', 'Value', 'Source', 'First Seen', 'Last Seen', 'Tags', 'Notes']
    
    # Write headers
    for col_idx, header in enumerate(headers, 1):
        cell = ws_all.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border_style
    
    # Write IOCs
    for row_idx, ioc in enumerate(iocs, 2):
        # Alternate row colors
        row_fill = PatternFill(
            start_color=colors['row_bg_alt'] if row_idx % 2 == 0 else colors['row_bg_light'],
            end_color=colors['row_bg_alt'] if row_idx % 2 == 0 else colors['row_bg_light'],
            fill_type='solid'
        )
        
        source = db.get_source(ioc.get('source_id'))
        source_name = source.get('name', '') if source else ioc.get('source_name', '')
        
        # Handle tags
        tags_str = ''
        tags = ioc.get('tags', [])
        if tags:
            if isinstance(tags, list):
                if tags and isinstance(tags[0], dict):
                    tags_str = ', '.join([tag.get('name', '') for tag in tags if tag.get('name')])
                else:
                    tags_str = ', '.join([str(tag) for tag in tags if tag])
        
        # Type cell with accent color for certain types
        ioc_type = ioc.get('ioc_type', '')
        type_cell = ws_all.cell(row=row_idx, column=1, value=ioc_type)
        if ioc_type in ['ip4', 'ip6', 'fqdn', 'url']:
            type_cell.fill = network_fill
        elif ioc_type in ['md5', 'sha1', 'sha256', 'sha512']:
            type_cell.fill = file_fill
        
        # Value cell with monospace font
        value_cell = ws_all.cell(row=row_idx, column=2, value=ioc.get('ioc_value', ''))
        value_cell.font = Font(name='Consolas', size=10, color='1D1D1F')  # Monospace for values
        
        # Other cells
        ws_all.cell(row=row_idx, column=3, value=source_name)
        ws_all.cell(row=row_idx, column=4, value=ioc.get('first_seen', ''))
        ws_all.cell(row=row_idx, column=5, value=ioc.get('last_seen', ''))
        ws_all.cell(row=row_idx, column=6, value=tags_str)
        ws_all.cell(row=row_idx, column=7, value=ioc.get('notes', ''))
        
        # Apply styles to all cells in row
        for col_idx in range(1, len(headers) + 1):
            cell = ws_all.cell(row=row_idx, column=col_idx)
            if col_idx == 2:  # Value column - monospace
                cell.font = Font(name='Consolas', size=10, color='1D1D1F')
            else:
                cell.font = cell_font
            # Don't override type cell fill if it has accent color
            if col_idx != 1 or ioc_type not in ['ip4', 'ip6', 'fqdn', 'url', 'md5', 'sha1', 'sha256', 'sha512']:
                cell.fill = row_fill
            cell.alignment = left_align if col_idx in [2, 6, 7] else center_align
            cell.border = border_style
    
    # Auto-adjust column widths
    column_widths = {
        'A': 12,  # Type
        'B': 40,  # Value
        'C': 25,  # Source
        'D': 18,  # First Seen
        'E': 18,  # Last Seen
        'F': 30,  # Tags
        'G': 40,  # Notes
    }
    for col_letter, width in column_widths.items():
        ws_all.column_dimensions[col_letter].width = width
    
    # Freeze header row
    ws_all.freeze_panes = 'A2'
    
    # Group IOCs by type for additional sheets
    iocs_by_type = {}
    for ioc in iocs:
        ioc_type = ioc.get('ioc_type', 'unknown')
        if ioc_type not in iocs_by_type:
            iocs_by_type[ioc_type] = []
        iocs_by_type[ioc_type].append(ioc)
    
    # Create sheets for common IOC types (if more than 5 IOCs of that type)
    common_types = ['ip4', 'ip6', 'fqdn', 'url', 'md5', 'sha1', 'sha256', 'email']
    for ioc_type in common_types:
        if ioc_type in iocs_by_type and len(iocs_by_type[ioc_type]) > 5:
            ws_type = wb.create_sheet(ioc_type.upper(), len(wb.worksheets))
            
            # Write headers
            type_headers = ['Value', 'Source', 'First Seen', 'Last Seen', 'Tags']
            for col_idx, header in enumerate(type_headers, 1):
                cell = ws_type.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align
                cell.border = border_style
            
            # Write IOCs of this type
            for row_idx, ioc in enumerate(iocs_by_type[ioc_type], 2):
                row_fill = PatternFill(
                    start_color=colors['row_bg_alt'] if row_idx % 2 == 0 else colors['row_bg_light'],
                    end_color=colors['row_bg_alt'] if row_idx % 2 == 0 else colors['row_bg_light'],
                    fill_type='solid'
                )
                
                source = db.get_source(ioc.get('source_id'))
                source_name = source.get('name', '') if source else ioc.get('source_name', '')
                
                tags_str = ''
                tags = ioc.get('tags', [])
                if tags:
                    if isinstance(tags, list):
                        if tags and isinstance(tags[0], dict):
                            tags_str = ', '.join([tag.get('name', '') for tag in tags if tag.get('name')])
                        else:
                            tags_str = ', '.join([str(tag) for tag in tags if tag])
                
                value_cell = ws_type.cell(row=row_idx, column=1, value=ioc.get('ioc_value', ''))
                value_cell.font = Font(name='Consolas', size=10, color='1D1D1F')
                ws_type.cell(row=row_idx, column=2, value=source_name)
                ws_type.cell(row=row_idx, column=3, value=ioc.get('first_seen', ''))
                ws_type.cell(row=row_idx, column=4, value=ioc.get('last_seen', ''))
                ws_type.cell(row=row_idx, column=5, value=tags_str)
                
                for col_idx in range(1, len(type_headers) + 1):
                    cell = ws_type.cell(row=row_idx, column=col_idx)
                    cell.font = cell_font if col_idx != 1 else Font(name='Consolas', size=10, color='1D1D1F')
                    cell.fill = row_fill
                    cell.alignment = left_align if col_idx in [1, 5] else center_align
                    cell.border = border_style
            
            # Auto-adjust column widths
            type_widths = {'A': 50, 'B': 25, 'C': 18, 'D': 18, 'E': 30}
            for col_letter, width in type_widths.items():
                ws_type.column_dimensions[col_letter].width = width
            
            ws_type.freeze_panes = 'A2'
    
    # Summary sheet
    ws_summary = wb.create_sheet("Summary", 0)
    
    # Summary data
    summary_data = [
        ['Export Information', ''],
        ['Export Date', datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
        ['Total IOCs', len(iocs)],
        ['', ''],
        ['IOC Types Breakdown', 'Count'],
    ]
    
    # Add type counts
    type_counts = {}
    for ioc in iocs:
        ioc_type = ioc.get('ioc_type', 'unknown')
        type_counts[ioc_type] = type_counts.get(ioc_type, 0) + 1
    
    for ioc_type, count in sorted(type_counts.items(), key=lambda x: x[1], reverse=True):
        summary_data.append([ioc_type, count])
    
    # Write summary with elegant formatting
    for row_idx, (label, value) in enumerate(summary_data, 1):
        label_cell = ws_summary.cell(row=row_idx, column=1, value=label)
        value_cell = ws_summary.cell(row=row_idx, column=2, value=value)
        
        if row_idx == 1:  # Title row
            label_cell.font = Font(name='Calibri', size=14, bold=True, color=colors['header_text'])
            label_cell.fill = PatternFill(start_color=colors['summary_bg'], end_color=colors['summary_bg'], fill_type='solid')
            value_cell.fill = PatternFill(start_color=colors['summary_bg'], end_color=colors['summary_bg'], fill_type='solid')
        elif row_idx <= 4:  # Info rows
            label_cell.font = Font(name='Calibri', size=10, bold=True, color='1D1D1F')
            value_cell.font = cell_font
            label_cell.fill = PatternFill(start_color=colors['row_bg_light'], end_color=colors['row_bg_light'], fill_type='solid')
            value_cell.fill = PatternFill(start_color=colors['row_bg_light'], end_color=colors['row_bg_light'], fill_type='solid')
        elif row_idx == 5:  # Header for breakdown
            label_cell.font = header_font
            value_cell.font = header_font
            label_cell.fill = header_fill
            value_cell.fill = header_fill
        else:  # Type breakdown rows
            label_cell.font = cell_font
            value_cell.font = Font(name='Calibri', size=10, bold=True, color=colors['accent'])
            row_fill = PatternFill(
                start_color=colors['row_bg_alt'] if row_idx % 2 == 0 else colors['row_bg_light'],
                end_color=colors['row_bg_alt'] if row_idx % 2 == 0 else colors['row_bg_light'],
                fill_type='solid'
            )
            label_cell.fill = row_fill
            value_cell.fill = row_fill
        
        label_cell.alignment = left_align
        value_cell.alignment = center_align
        label_cell.border = border_style
        value_cell.border = border_style
    
    ws_summary.column_dimensions['A'].width = 25
    ws_summary.column_dimensions['B'].width = 15
    
    # Set row height for header
    ws_summary.row_dimensions[1].height = 25
    
    # Save workbook
    wb.save(str(output_file))
    logger.info(f"XLSX export generated: {output_file}")
    
    return output_file


def export_stix(db, data: Dict, output_folder: Path, convert_to_stix_func) -> Optional[Path]:
    """Export IOCs to STIX format"""
    iocs, source_context = get_iocs_for_export(db, data)
    
    report_name = data.get('report_name', 'CTI Export')
    provided_context = data.get('source_context', '')
    
    # Use provided context or combine with source contexts
    if provided_context and source_context:
        final_context = f"{provided_context}\n\n---\n\n{source_context}"
    elif provided_context:
        final_context = provided_context
    else:
        final_context = source_context or 'CTI Export'
    
    # Convert to STIX
    stix_file = convert_to_stix_func(
        iocs_list=iocs,
        source_context=final_context,
        report_name=report_name
    )
    
    return stix_file

