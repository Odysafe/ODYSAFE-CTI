"""Professional Excel renderer for FLINT schema version 2 reports."""

from datetime import datetime, timezone
from io import BytesIO
from urllib.parse import urlparse

from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


NAVY = "243247"
BLUE = "DCE9F7"
PURPLE = "E9E2F5"
ORANGE = "FCE8D2"
RED = "F7D7D7"
GREEN = "DDF0E5"
GRAY = "EEF1F5"
WHITE = "FFFFFF"
TEXT = "172033"
THIN = Side(style="thin", color="CDD5DF")
MAX_CELL = 32000


def safe_cell(value):
    """Keep user content as text and prevent spreadsheet formula execution."""
    if value is None:
        return ""
    if isinstance(value, bool):
        return "Yes" if value else "No"
    text = str(value).replace("\x00", "")[:MAX_CELL]
    if text.startswith(("=", "+", "-", "@")):
        text = "'" + text
    return text


def nonempty(rows):
    ignored = {"id", "internal_id", "imported"}
    return [row for row in (rows or []) if isinstance(row, dict) and any(
        str(value or "").strip() for key, value in row.items() if key not in ignored
    )]


def defang(value, indicator_type):
    text = str(value or "")
    kind = str(indicator_type or "").lower()
    if kind in {"url", "domain", "ip", "email"}:
        if kind == "url":
            text = text.replace("https://", "hxxps://").replace("http://", "hxxp://")
        if kind == "email":
            text = text.replace("@", "[@]")
        text = text.replace(".", "[.]")
    return text


def style_title(ws, title, data, columns=8):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=columns)
    cell = ws.cell(1, 1, safe_cell(title))
    cell.fill = PatternFill("solid", fgColor=NAVY)
    cell.font = Font(size=18, bold=True, color=WHITE)
    cell.alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 34
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=columns)
    meta = ws.cell(2, 1, safe_cell(f"{data.get('reference', '')} | {data.get('status', 'Draft')} | {data.get('tlp', '')}"))
    meta.fill = PatternFill("solid", fgColor=GRAY)
    meta.font = Font(size=10, bold=True, color=TEXT)
    ws.freeze_panes = "A4"
    ws.sheet_view.showGridLines = False
    ws.oddHeader.center.text = safe_cell(data.get("reference", "FLINT"))
    ws.oddFooter.left.text = safe_cell(data.get("tlp", ""))
    ws.oddFooter.right.text = "Exported UTC " + datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M")
    return 4


def section(ws, row, title, columns=8, color=BLUE):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=columns)
    cell = ws.cell(row, 1, safe_cell(title))
    cell.fill = PatternFill("solid", fgColor=color)
    cell.font = Font(size=11, bold=True, color=TEXT)
    cell.border = Border(bottom=THIN)
    return row + 1


def fields(ws, row, values, columns=8):
    for label, value in values:
        ws.cell(row, 1, safe_cell(label)).font = Font(bold=True, color=TEXT)
        ws.cell(row, 1).border = Border(bottom=THIN)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=columns)
        cell = ws.cell(row, 2, safe_cell(value) or "Not provided")
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        cell.border = Border(bottom=THIN)
        ws.row_dimensions[row].height = max(20, min(90, 15 + len(str(value or "")) // 10))
        row += 1
    return row


def table(ws, row, title, rows, columns, data, color=PURPLE, links=None):
    row = section(ws, row, title, len(columns), color)
    rows = nonempty(rows)
    if not rows:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(columns))
        ws.cell(row, 1, "No data provided.").font = Font(italic=True, color="667085")
        return row + 2
    header_row = row
    for col, (_, label) in enumerate(columns, 1):
        cell = ws.cell(row, col, label)
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.font = Font(bold=True, color=WHITE)
        cell.alignment = Alignment(wrap_text=True)
        cell.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    row += 1
    for idx, item in enumerate(rows):
        for col, (key, _) in enumerate(columns, 1):
            raw = item.get(key, "")
            cell = ws.cell(row, col, safe_cell(raw))
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = Border(left=THIN, right=THIN, bottom=THIN)
            if idx % 2:
                cell.fill = PatternFill("solid", fgColor="F8FAFC")
            if links and key in links and valid_url(raw):
                cell.hyperlink = str(raw)
                cell.font = Font(color="0563C1", underline="single")
        ws.row_dimensions[row].height = 34
        row += 1
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(columns))}{row - 1}"
    return row + 2


def valid_url(value):
    try:
        parsed = urlparse(str(value or ""))
        return parsed.scheme in {"http", "https"} and bool(parsed.netloc)
    except ValueError:
        return False


def finish(ws, widths=None):
    widths = widths or {}
    for idx in range(1, ws.max_column + 1):
        letter = get_column_letter(idx)
        estimated = max((len(str(ws.cell(row, idx).value or "")) for row in range(1, min(ws.max_row, 250) + 1)), default=10) + 2
        ws.column_dimensions[letter].width = min(widths.get(letter, max(12, estimated)), 42)
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_title_rows = "1:3"


def _diamond_text(value):
    text = safe_cell(value) or "Not specified"
    return text[:237] + "..." if len(text) > 240 else text


def _diamond_image(data):
    """Create the colored Diamond Model image embedded in the workbook."""
    from PIL import Image, ImageDraw, ImageFont

    image = Image.new("RGB", (1000, 650), "#FAFAFC")
    draw = ImageDraw.Draw(image)
    try:
        title_font = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf", 16)
        body_font = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf", 13)
    except OSError:
        title_font = body_font = ImageFont.load_default()

    center_x, center_y = 500, 325
    points = [(center_x, 35), (965, center_y), (center_x, 615), (35, center_y)]
    draw.polygon(points, fill="#F5F3FF", outline="#7C3AED", width=5)
    draw.line([points[0], points[2]], fill="#A78BFA", width=2)
    draw.line([points[3], points[1]], fill="#A78BFA", width=2)

    def card(box, title, value, header, body):
        x0, y0, x1, y1 = box
        draw.rounded_rectangle(box, radius=12, fill=body, outline="#6D28D9", width=3)
        draw.rectangle((x0, y0, x1, y0 + 34), fill=header)
        draw.text(((x0 + x1) / 2, y0 + 17), title, fill="white", font=title_font, anchor="mm")
        words = _diamond_text(value).split()
        lines, current = [], ""
        for word in words:
            candidate = f"{current} {word}".strip()
            if draw.textlength(candidate, font=body_font) <= x1 - x0 - 22:
                current = candidate
            else:
                if current: lines.append(current)
                current = word
            if len(lines) == 4: break
        if current and len(lines) < 5: lines.append(current)
        for index, line in enumerate(lines[:5]):
            draw.text((x0 + 12, y0 + 49 + index * 17), line, fill="#1E1033", font=body_font)

    card((365, 52, 635, 176), "ADVERSARY", data.get("diamond_adversary"), "#5B21B6", "#EDE9FE")
    card((52, 263, 322, 387), "CAPABILITY", data.get("diamond_capability"), "#6D28D9", "#EDE9FE")
    card((678, 263, 948, 387), "INFRASTRUCTURE", data.get("diamond_infrastructure"), "#7C3AED", "#DDD6FE")
    card((365, 474, 635, 598), "VICTIM", data.get("diamond_victim"), "#6D28D9", "#C4B5FD")
    draw.rounded_rectangle((415, 275, 585, 375), radius=16, fill="#4C1D95", outline="#EDE9FE", width=3)
    draw.text((500, 316), "EVENT", fill="white", font=title_font, anchor="mm")
    draw.text((500, 345), "Core incident", fill="#DDD6FE", font=body_font, anchor="mm")
    stream = BytesIO()
    image.save(stream, format="PNG")
    stream.seek(0)
    return stream


def diamond_model(ws, row, data, columns):
    row = section(ws, row, "Diamond Model of Intrusion Analysis", columns, PURPLE)
    try:
        image = XLImage(_diamond_image(data))
        image.width, image.height = 650, 422
        ws.add_image(image, f"A{row}")
        for current_row in range(row, row + 23):
            ws.row_dimensions[current_row].height = 19
        return row + 24
    except Exception:
        colors = {"Adversary": "EDE9FE", "Capability": "EDE9FE", "Infrastructure": "DDD6FE", "Victim": "C4B5FD"}
        values = [("Adversary", data.get("diamond_adversary")), ("Capability", data.get("diamond_capability")),
                  ("Infrastructure", data.get("diamond_infrastructure")), ("Victim", data.get("diamond_victim"))]
        for label, value in values:
            ws.cell(row, 1, label).fill = PatternFill("solid", fgColor="6D28D9")
            ws.cell(row, 1).font = Font(bold=True, color=WHITE)
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=columns)
            cell = ws.cell(row, 2, _diamond_text(value))
            cell.fill = PatternFill("solid", fgColor=colors[label])
            cell.alignment = Alignment(wrap_text=True, vertical="center")
            ws.row_dimensions[row].height = 36
            row += 1
        return row + 1


def dashboard(wb, data):
    ws = wb.create_sheet("Dashboard")
    row = style_title(ws, "FLINT EXECUTIVE DASHBOARD", data, 9)
    if data.get("export_incomplete") or data.get("status") == "Draft":
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
        cell = ws.cell(row, 1, "DRAFT / INCOMPLETE")
        cell.fill = PatternFill("solid", fgColor=ORANGE); cell.font = Font(size=14, bold=True, color="8A3D00"); cell.alignment = Alignment(horizontal="center")
        row += 2
    row = section(ws, row, "Report Control", 9)
    row = fields(ws, row, [("Reference",data.get("reference")),("Title",data.get("title")),("Product Type",data.get("product_type")),("Report Status",data.get("status")),("Investigation Status",data.get("investigation_status")),("TLP",data.get("tlp")),("PAP",data.get("pap")),("Priority",data.get("priority")),("Author",data.get("author")),("Created",data.get("created_at")),("Updated",data.get("updated_at")),("Intelligence Requirement / PIR",data.get("pir"))],9)
    row += 1; row = section(ws,row,"Executive Assessment",9)
    row = fields(ws,row,[("Initial Signal",data.get("trigger_details")),("Why It Matters",data.get("why_it_matters")),("Required Action",data.get("required_action")),("Likelihood",data.get("likelihood")),("Analytic Confidence",data.get("analytic_confidence")),("Threat / Activity Type",data.get("threat_type")),("Actor / Campaign / Cluster",data.get("entity_name"))],9)
    row = table(ws,row,"Key Judgements",data.get("keyJudgements"),[("judgement","Judgement"),("likelihood","Likelihood"),("confidence","Confidence")],data)
    counts=[("IOC count",len(nonempty(data.get("iocs")))),("TTP count",len(nonempty(data.get("techniques")))),("Actions count",len(nonempty(data.get("actions")))),("Hunt leads",len(nonempty(data.get("hunts")))),("Open gaps",1 if str(data.get("known_unknowns") or "").strip() else 0)]
    row=section(ws,row,"Operational KPIs",9)
    for col,(label,count) in enumerate(counts,1):
        cell=ws.cell(row,col,f"{label}\n{count}");cell.fill=PatternFill("solid",fgColor=GRAY);cell.font=Font(bold=True,color=TEXT);cell.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True);cell.border=Border(left=THIN,right=THIN,top=THIN,bottom=THIN)
    ws.row_dimensions[row].height=48;row+=2
    row=diamond_model(ws,row,data,9)
    finish(ws)


def executive(wb, data):
    ws=wb.create_sheet("Executive Summary");row=style_title(ws,"EXECUTIVE SUMMARY",data,8)
    row=section(ws,row,"Report Control",8);row=fields(ws,row,[("Reference",data.get("reference")),("Title",data.get("title")),("Author",data.get("author")),("Status",data.get("status")),("TLP / PAP",f"{data.get('tlp','')} / {data.get('pap','')}"),("Priority",data.get("priority")),("Created / Updated",f"{data.get('created_at','')} / {data.get('updated_at','')}"),("PIR",data.get("pir"))],8)
    row=section(ws,row,"Summary",8);row=fields(ws,row,[("Initial Signal / Trigger",data.get("trigger_details")),("Why It Matters",data.get("why_it_matters")),("Required Action",data.get("required_action")),("Who / What Is Affected",data.get("affected_scope")),("Threat / Activity Type",data.get("threat_type")),("Likelihood",data.get("likelihood")),("Analytic Confidence",data.get("analytic_confidence"))],8)
    row=table(ws,row,"Key Judgements",data.get("keyJudgements"),[("judgement","Judgement"),("likelihood","Likelihood"),("confidence","Confidence"),("evidence","Evidence Reference")],data)
    table(ws,row,"Confidence by Area",data.get("confidenceAreas"),[("area","Area"),("confidence","Confidence"),("notes","Notes")],data);finish(ws)


def activity(wb,data):
    ws=wb.create_sheet("Technical Analysis");row=style_title(ws,"TECHNICAL ANALYSIS",data,9)
    row=section(ws,row,"Technical Details",9);row=fields(ws,row,[("Threat History / Context",data.get("relevant_background")),("Attack Vector",data.get("attack_vector")),("Malware / Family",data.get("malware_family")),("Persistence",data.get("persistence")),("C2 Infrastructure",data.get("c2_infrastructure")),("Anchor / Key Discriminator",data.get("anchor"))],9)
    row=table(ws,row,"MITRE ATT&CK",data.get("techniques"),[("id","ID"),("name","Technique"),("tactic","Tactic"),("procedure","Procedure"),("status","Status"),("evidence","Evidence"),("confidence","Confidence"),("source","Source")],data)
    row=diamond_model(ws,row,data,9)
    row=section(ws,row,"Analytic Reasoning",9);row=fields(ws,row,[("Working Hypothesis",data.get("working_hypothesis")),("Evidence For",data.get("evidence_for")),("Evidence Against / Contradictions",data.get("evidence_against")),("Alternative Explanation",data.get("alternative_explanations")),("Next Best Pivot",data.get("next_best_pivot"))],9)
    row=table(ws,row,"Attack Sequence",data.get("attack_sequence"),[("step","Step"),("date","Date / Time"),("technique","Technique"),("procedure","Procedure"),("evidence","Evidence")],data)
    table(ws,row,"Timeline",data.get("timeline"),[("date","Date UTC"),("event","Event"),("status","Status"),("significance","Significance"),("source","Source"),("evidence","Evidence Reference")],data);finish(ws)


def indicators(wb,data):
    ws=wb.create_sheet("IOCs");row=style_title(ws,"INDICATORS OF COMPROMISE",data,8)
    rows=nonempty(data.get("iocs"));render=[]
    for item in rows:
        copy=dict(item);copy["value"]=defang(item.get("value"),item.get("type")) if data.get("defang_export",True) else item.get("value","");render.append(copy)
    table(ws,row,"Indicators",render,[("type","Type"),("value","Value"),("role","Role"),("tlp","TLP"),("confidence","Confidence"),("origin","Origin"),("first_seen","First Seen"),("last_seen","Last Seen"),("operational_status","Status"),("recommended_use","Recommended Use"),("notes","Context / Relationship")],data);finish(ws)


def hunt_detection(wb,data):
    ws=wb.create_sheet("Hunt & Detection");row=style_title(ws,"HUNT AND DETECTION",data,8)
    row=table(ws,row,"Hunt Leads",data.get("hunts"),[("hypothesis","Hypothesis"),("observable","What to Search"),("data_source","Data Source"),("time_window","Time Window"),("expected_evidence","Expected Evidence"),("pivot_hit","Pivot If Hit"),("no_hit","Next Step If No Hit")],data)
    row=table(ws,row,"Hunt Results",data.get("hunt_results"),[("result","Result"),("new_artifacts","New Artifacts"),("interpretation","Interpretation"),("follow_up","Follow-up")],data)
    detections=nonempty(data.get("detections"));summary=[{k:v for k,v in x.items() if k!="content"} for x in detections]
    row=table(ws,row,"Detection Rules",summary,[("type","Type"),("name","Name"),("validation_status","Validation Status"),("telemetry","Required Telemetry"),("false_positives","False Positive Notes")],data)
    for rule in detections:
        row=section(ws,row,f"Rule Content: {rule.get('name') or rule.get('type') or 'Unnamed Rule'}",8,GRAY);ws.merge_cells(start_row=row,start_column=1,end_row=row,end_column=8);cell=ws.cell(row,1,safe_cell(rule.get("content")) or "No content provided.");cell.font=Font(name="Consolas",size=9);cell.alignment=Alignment(wrap_text=True,vertical="top");ws.row_dimensions[row].height=min(300,max(35,15+str(rule.get("content") or "").count("\n")*12));row+=2
    finish(ws)


def actions_gaps(wb,data):
    ws=wb.create_sheet("Actions & Gaps");row=style_title(ws,"ACTIONS AND INTELLIGENCE GAPS",data,8)
    row=table(ws,row,"Actions",data.get("actions"),[("priority","Priority"),("action","Action"),("type","Type"),("owner","Owner"),("due_date","Due Date"),("status","Status"),("trigger","Condition")],data)
    row=section(ws,row,"Intelligence Gaps",8);row=fields(ws,row,[("Summary",data.get("known_unknowns")),("Hunt Questions",data.get("hunt_questions")),("RFIs",data.get("rfis_text"))],8)
    table(ws,row,"Gap Register",data.get("gaps"),[("gap","Gap"),("priority","Priority"),("resolution","What would resolve it?"),("owner","Owner"),("status","Status")],data);finish(ws)


def sources_confidence(wb,data):
    ws=wb.create_sheet("Sources & Assessment");row=style_title(ws,"SOURCES AND ASSESSMENT",data,8)
    row=table(ws,row,"Sources",data.get("sources"),[("name","Source Name"),("date","Date"),("url","URL / Reference"),("type","Type"),("reliability","Reliability"),("primacy","Provenance"),("provenance","Relationship")],data,links={"url"})
    row=section(ws,row,"Assessment",8);fields(ws,row,[("Investigation Status",data.get("investigation_status")),("Overall Likelihood",data.get("overall_likelihood") or data.get("likelihood")),("Analytic Confidence",data.get("overall_analytic_confidence") or data.get("analytic_confidence")),("Confidence Rationale",data.get("confidence_rationale")),("Key Assumptions",data.get("key_assumptions")),("Contradictions / Alternative Explanations",data.get("contradictory_evidence")),("Confidence by Section",data.get("section_confidence"))],8);finish(ws)


def distribution(wb,data):
    ws=wb.create_sheet("Distribution");row=style_title(ws,"DISTRIBUTION",data,6)
    if data.get("tlp")=="TLP:RED":ws.merge_cells(start_row=row,start_column=1,end_row=row,end_column=6);cell=ws.cell(row,1,"TLP:RED | Named recipients only");cell.fill=PatternFill("solid",fgColor=RED);cell.font=Font(bold=True,color="8F1D1D");cell.alignment=Alignment(horizontal="center");row+=2
    row=table(ws,row,"Distribution",data.get("recipients"),[("name","Name"),("role","Role"),("organization","Organisation"),("email","Email")],data)
    row=section(ws,row,"Handling",6);row=fields(ws,row,[("Distribution List",data.get("distribution_list")),("TLP",data.get("tlp")),("PAP",data.get("pap")),("Handling Instructions",data.get("handling_instructions")),("Feedback Contact",data.get("feedback_contact"))],6)
    row=section(ws,row,"Disclaimer",6);fields(ws,row,[("Notice","This product is provided for authorised defensive use. Validate intelligence against organisational policy and current telemetry before taking action.")],6);finish(ws)


def build_workbook(data):
    wb=Workbook();wb.remove(wb.active)
    for builder in (dashboard,executive,activity,indicators,hunt_detection,actions_gaps,sources_confidence,distribution):builder(wb,data)
    stream=BytesIO();wb.save(stream);stream.seek(0)
    # Re-open once to ensure the generated package is structurally valid.
    load_workbook(stream,read_only=True).close();stream.seek(0)
    return stream
