"""
Odysafe CTI Platform
Copyright (C) 2026 Bastien GUIDONE

This program is free software: you can redistribute it and/or modify
it under the terms of the GNU Affero General Public License as published by
the Free Software Foundation, either version 3 of the License, or
(at your option) any later version.

This program is distributed in the hope that it will be useful,
but WITHOUT ANY WARRANTY; without even the implied warranty of
MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
GNU Affero General Public License for more details.

You should have received a copy of the GNU Affero General Public License
along with this program.  If not, see <https://www.gnu.org/licenses/>.

Application Flask principale - CTI Platform
"""
import os
import json
import csv
import logging
import sqlite3
import threading
import time
import zipfile
import tempfile
import socket
import hashlib
from datetime import datetime, timedelta
from pathlib import Path
from flask import Flask, render_template, request, send_file, redirect, url_for, flash, session, g
from typing import List, Dict
from werkzeug.utils import secure_filename
from werkzeug.exceptions import RequestEntityTooLarge
from werkzeug.serving import run_simple

from config import (
    UPLOAD_FOLDER, OUTPUT_FOLDER, ALLOWED_EXTENSIONS, ALLOWED_MIME_TYPES, MAX_FILE_SIZE,
    CLEANUP_DAYS, PORT, HOST, SECRET_KEY, DEBUG,
    USE_SSL, SSL_CERT_FILE, SSL_KEY_FILE
)
from database import Database
from modules.iocsearcher_wrapper import (
    extract_iocs, extract_from_text, extract_from_url,
    IOCSEARCHER_AVAILABLE
)
from modules.storage_monitor import get_storage_info, format_bytes
from modules.progress_tracker import progress_tracker
from modules.api_helpers import api_success, api_error, api_not_found
from modules.github_repo import github_repo_manager
from modules.github_repo_rtm import rtm_repo_manager
from modules.github_repo_data_shield import data_shield_repo_manager
from modules.apt_annotations import apt_annotations_manager
from modules.export_helpers import (
    export_txt, export_json, export_csv, export_stix,
    export_txt_simple, export_csv_firewall, export_json_simple,
    export_xlsx
)
from modules.mitre_stix_parser import (
    MitreStixParser, mitre_json_exists, get_mitre_json_path, get_loaded_parser
)
from modules.auth import (
    require_auth, is_auth_enabled, create_user, verify_user,
    change_password, user_exists, get_current_username
)
from modules import main_bp, upload_bp, iocs_bp, sources_bp, export_bp, settings_bp, stix_graph_bp, cti_resources_bp, mitre_attack_bp, ransomware_matrix_bp
from routes_flash_report import flash_report_bp
from routes_memory import memory_bp
from markdown import markdown as md_markdown

# ANSI color codes for logging
class ColoredFormatter(logging.Formatter):
    """Custom formatter with colors for different log levels"""
    COLORS = {
        'DEBUG': '\033[0;36m',     # Cyan
        'INFO': '\033[0;32m',      # Green
        'WARNING': '\033[1;33m',   # Yellow
        'ERROR': '\033[0;31m',     # Red
        'CRITICAL': '\033[1;31m',  # Bold Red
    }
    RESET = '\033[0m'
    
    def format(self, record):
        # Shorten level name
        level_short = record.levelname[0]  # D, I, W, E, C
        color = self.COLORS.get(record.levelname, self.RESET)
        record.level_short = f"{color}{level_short}{self.RESET}"
        return super().format(record)

# Configure root logger
logging.basicConfig(
    level=logging.INFO,
    format='%(level_short)s %(message)s',
    handlers=[logging.StreamHandler()]
)

# Apply colored formatter to root handler
for handler in logging.root.handlers:
    handler.setFormatter(ColoredFormatter('%(level_short)s %(message)s'))

logger = logging.getLogger(__name__)

# Suppress werkzeug request logs (keep only errors)
werkzeug_logger = logging.getLogger('werkzeug')
werkzeug_logger.setLevel(logging.ERROR)

# Filter out HTTPS connection attempts (TLS handshakes) and SSL errors
class HTTPSFilter(logging.Filter):
    def filter(self, record):
        msg_str = str(record.msg) if hasattr(record, 'msg') else ''
        # Filter out "Bad request version" errors that are TLS handshakes
        if 'Bad request version' in msg_str:
            return False
        # Filter out SSL EOF errors (HTTPS attempts on HTTP server)
        if 'SSLEOFError' in msg_str or 'EOF occurred in violation of protocol' in msg_str:
            return False
        return True

werkzeug_logger.addFilter(HTTPSFilter())

# Suppress verbose urllib3 logs
logging.getLogger('urllib3').setLevel(logging.WARNING)

# Initialisation Flask
app = Flask(__name__)
app.secret_key = SECRET_KEY
app.config['UPLOAD_FOLDER'] = str(UPLOAD_FOLDER)
app.config['MAX_CONTENT_LENGTH'] = MAX_FILE_SIZE
# Always reload changed templates and force browsers to revalidate static assets.
# This prevents an old Flash Report UI from surviving an ODYSAFE update.
app.config['TEMPLATES_AUTO_RELOAD'] = True
app.config['SEND_FILE_MAX_AGE_DEFAULT'] = 0
app.jinja_env.auto_reload = True
# Configuration de la session : pas de timeout automatique, session permanente jusqu'à déconnexion
# Utiliser une valeur très grande (10 ans) car Flask n'accepte pas None pour PERMANENT_SESSION_LIFETIME
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(days=3650)  # 10 ans (session quasi permanente)
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_SECURE'] = USE_SSL  # Secure cookies si SSL activé

# Initialisation base de données
db = Database()

# Register blueprints
app.register_blueprint(flash_report_bp)
app.register_blueprint(memory_bp)

@app.after_request
def prevent_stale_ui_cache(response):
    """Make UI updates visible immediately while retaining conditional static requests."""
    if response.mimetype == 'text/html':
        response.headers['Cache-Control'] = 'no-store, max-age=0'
        response.headers['Pragma'] = 'no-cache'
        response.headers['Expires'] = '0'
    elif request.path.startswith('/static/'):
        response.headers['Cache-Control'] = 'no-cache, max-age=0, must-revalidate'
    return response

@app.before_request
def before_request():
    g.db = db
    if is_auth_enabled(db):
        path = request.path
        public_paths = ['/login', '/api/auth/status', '/api/auth/create-user', '/api/auth/enable']
        if not path.startswith('/static/') and path not in public_paths:
            if 'user_id' not in session:
                if request.is_json or path.startswith('/api/'):
                    return api_error("Authentication required", 401)
                return redirect(url_for('login'))

# Ajouter la fonction markdown au contexte des templates
@app.template_filter('markdown')
def markdown_filter(text):
    """Convert Markdown to HTML"""
    if not text:
        return ''
    return md_markdown(text, extensions=['fenced_code', 'tables', 'nl2br'])

# ========== HELPERS ==========

def allowed_file(filename):
    """Checks if the file has an allowed extension"""
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def validate_file_mime(file_path: Path) -> bool:
    """
    Validates file MIME type using python-magic.
    Returns True if MIME type is allowed, False otherwise.
    """
    try:
        import magic
        mime = magic.Magic(mime=True)
        detected_mime = mime.from_file(str(file_path))
        
        # Normalize MIME type (remove charset, etc.)
        detected_mime = detected_mime.split(';')[0].strip()
        
        # Check if detected MIME is in allowed list
        if detected_mime in ALLOWED_MIME_TYPES:
            return True
        
        # Also check if it's a text file (many log files are detected as text/plain)
        if detected_mime.startswith('text/'):
            # Check extension to be sure
            ext = file_path.suffix.lower().lstrip('.')
            if ext in ['txt', 'log', 'md', 'html', 'htm', 'xml', 'csv']:
                return True
        
        logger.warning(f"File {file_path.name} has disallowed MIME type: {detected_mime}")
        return False
    except ImportError:
        # python-magic not available, skip MIME validation but log warning
        logger.warning("python-magic not available, MIME validation skipped")
        return True  # Allow file if magic is not available
    except Exception as e:
        logger.error(f"MIME validation error for {file_path.name}: {e}")
        # On error, be permissive but log it
        return True

def generate_default_context(source_type: str) -> str:
    """Generates a default context based on source type"""
    now = datetime.now()
    context_map = {
        'file_upload': f"File upload - {now.strftime('%Y-%m-%d %H:%M:%S')}",
        'paste': f"Paste - {now.strftime('%Y-%m-%d %H:%M:%S')}",
        'url': f"URL - {now.strftime('%Y-%m-%d %H:%M:%S')}"
    }
    return context_map.get(source_type, f"Source - {now.strftime('%Y-%m-%d %H:%M:%S')}")

def check_and_apply_auto_rotation(db_instance):
    """Checks and applies auto-rotation of sources if enabled"""
    auto_rotation = db_instance.get_setting('auto_rotation_enabled', 'false').lower() == 'true'
    if auto_rotation:
        max_sources = int(db_instance.get_setting('max_sources', '20'))
        deleted_count = db_instance.rotate_sources_if_needed(max_sources)
        if deleted_count > 0:
            logger.info(f"Auto-rotation: {deleted_count} oldest source(s) deleted")
        return deleted_count
    return 0

_IOC_TYPE_NORM = {
    'ip4': 'ipv4', 'ip': 'ipv4', 'ip4net': 'ipv4',
    'ip6': 'ipv6', 'ip6net': 'ipv6',
    'fqdn': 'domain',
    'md5hash': 'md5', 'sha1hash': 'sha1', 'sha256hash': 'sha256',
    'email': 'email', 'emailaddr': 'email',
}

MEMORY_INDEX_TIMEOUT_SECONDS = max(
    5,
    int(os.getenv('ZETTELFORGE_INDEX_TIMEOUT_SECONDS', '60')),
)
_memory_index_slot = threading.BoundedSemaphore(1)


def _normalize_ioc_type(ioc_type: str) -> str:
    """Normalise raw iocsearcher type names to canonical internal names."""
    return _IOC_TYPE_NORM.get(ioc_type.lower(), ioc_type.lower())


def _remember_source_in_memory(source_id: int, source_info: Dict, iocs: List[Dict],
                               raw_text):
    """Import the bridge inside the bounded operation so import failures stay non-fatal."""
    from modules import zettelforge_bridge as zf

    return zf.remember_source_import(
        source_id,
        source_info,
        iocs,
        raw_text=raw_text,
    )


def _run_memory_index_with_timeout(task_id: str, operation) -> str:
    """Run one Memory operation without allowing it to block source processing forever."""
    deadline = time.monotonic() + MEMORY_INDEX_TIMEOUT_SECONDS
    while not _memory_index_slot.acquire(timeout=0.25):
        if progress_tracker.is_stopped(task_id):
            return 'stopped'
        if time.monotonic() >= deadline:
            logger.warning(
                "CTI Memory indexing queue timed out after %s seconds for task %s",
                MEMORY_INDEX_TIMEOUT_SECONDS,
                task_id,
            )
            return 'timed_out'

    result = {'status': 'failed', 'error': None}
    completed = threading.Event()

    def run_operation():
        try:
            result['status'] = operation()
        except Exception as exc:
            result['error'] = exc
        finally:
            _memory_index_slot.release()
            completed.set()

    memory_thread = threading.Thread(
        target=run_operation,
        name=f"memory-index-{task_id}",
        daemon=True,
    )
    memory_thread.start()

    cancellation_requested = False
    while not completed.wait(timeout=0.25):
        if progress_tracker.is_stopped(task_id):
            cancellation_requested = True
            progress_tracker.update_progress(
                task_id,
                message="Stopping after the current CTI Memory operation...",
            )
        if time.monotonic() >= deadline:
            logger.warning(
                "CTI Memory indexing timed out after %s seconds for task %s",
                MEMORY_INDEX_TIMEOUT_SECONDS,
                task_id,
            )
            return 'stopped' if cancellation_requested else 'timed_out'

    if result['error'] is not None:
        logger.warning('CTI Memory indexing failed for task %s: %s', task_id, result['error'])
        return 'stopped' if cancellation_requested else 'failed'
    if cancellation_requested:
        return 'stopped'
    return result['status'] if result['status'] in {'indexed', 'failed'} else 'failed'


def _run_auto_pipeline(source_id: int, iocs_list: List[Dict], source_info: Dict):
    """Generate and register an IOC report when automatic reports are enabled."""
    auto_generate = db.get_setting('auto_generate_reports', 'false') == 'true'
    if not auto_generate or not iocs_list:
        return

    try:
        reports_folder = OUTPUT_FOLDER / "reports"
        reports_folder.mkdir(parents=True, exist_ok=True)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        report_path = reports_folder / f"auto_ioc_report_{source_id}_{timestamp}.json"
        report_data = {
            "generated_at": datetime.now().isoformat(),
            "source": source_info or {"id": source_id},
            "total_iocs": len(iocs_list),
            "iocs": iocs_list,
        }

        with open(report_path, 'w', encoding='utf-8') as report_file:
            json.dump(report_data, report_file, indent=2, ensure_ascii=False, default=str)

        report_id = db.create_generated_report(
            source_id=source_id,
            report_type="automatic_ioc_report",
            file_path=str(report_path),
        )
        logger.info(
            "Automatic IOC report generated for source %s: report_id=%s, path=%s",
            source_id,
            report_id,
            report_path,
        )
    except Exception:
        logger.exception(
            "Automatic report pipeline failed for source %s; IOC extraction completed but report generation did not",
            source_id,
        )

def _extract_iocs_background(source_id: int, extraction_func, extraction_args: tuple,
                             initial_message: str, initial_percentage: int = 20):
    """Extract and persist IOCs while reporting cancellable background progress."""
    task_id = f"source_{source_id}"

    def stop_processing(message: str) -> None:
        try:
            db.update_source_status(source_id, 'stopped')
        except Exception as status_error:
            logger.error(
                "Failed to mark source %s as stopped: %s",
                source_id,
                status_error,
                exc_info=True,
            )
        progress_tracker.stopped_task(task_id, message)

    try:
        progress_tracker.update_progress(
            task_id,
            percentage=initial_percentage,
            message=initial_message,
        )
        extraction_output = extraction_func(*extraction_args)
        extracted_text = None
        if (
            isinstance(extraction_output, tuple)
            and len(extraction_output) == 2
            and isinstance(extraction_output[0], list)
        ):
            results, extracted_text = extraction_output
        else:
            results = extraction_output

        if progress_tracker.is_stopped(task_id):
            stop_processing("Stopped after IOC extraction")
            return

        progress_tracker.update_progress(
            task_id,
            percentage=60,
            message=f"{len(results)} IOCs extracted, saving...",
        )

        source_info = db.get_source(source_id)
        iocs_list = []
        total_results = len(results)
        progress_interval = max(1, total_results // 20) if total_results else 1

        for index, (ioc_type, ioc_value, raw_value, offset) in enumerate(results, start=1):
            if progress_tracker.is_stopped(task_id):
                stop_processing(
                    f"Stopped after saving {index - 1} of {total_results} extracted IOCs"
                )
                return

            ioc_type = _normalize_ioc_type(ioc_type)
            duplicate_id = db.check_duplicate(ioc_type, ioc_value, source_id)
            if not duplicate_id:
                db.create_ioc(
                    source_id,
                    ioc_type,
                    ioc_value,
                    raw_value,
                    source_info=source_info,
                )
                iocs_list.append({
                    'ioc_type': ioc_type,
                    'ioc_value': ioc_value,
                    'raw_value': raw_value,
                })

            if index == total_results or index % progress_interval == 0:
                save_percentage = 60 + int((index / max(1, total_results)) * 20)
                progress_tracker.update_progress(
                    task_id,
                    percentage=save_percentage,
                    message=f"Saving extracted IOCs: {index}/{total_results}",
                )

        if progress_tracker.is_stopped(task_id):
            stop_processing(f"Stopped after saving {total_results} extracted IOCs")
            return

        progress_tracker.update_progress(
            task_id,
            percentage=82,
            message="IOCs saved, running automatic report pipeline...",
        )
        _run_auto_pipeline(source_id, iocs_list, source_info)

        if progress_tracker.is_stopped(task_id):
            stop_processing("Stopped before CTI Memory indexing")
            return

        progress_tracker.update_progress(
            task_id,
            percentage=85,
            message="IOCs saved, indexing CTI Memory...",
        )
        if extracted_text is None and getattr(extraction_func, '__name__', '') == 'extract_from_text':
            extracted_text = extraction_args[0] if extraction_args else None

        all_iocs = [
            {'ioc_type': _normalize_ioc_type(t), 'ioc_value': v}
            for t, v, _, _ in results
        ]
        memory_index_status = _run_memory_index_with_timeout(
            task_id,
            lambda: _remember_source_in_memory(
                source_id, source_info, all_iocs or iocs_list, extracted_text
            ),
        )

        if memory_index_status == 'stopped':
            stop_processing("Stopped during CTI Memory indexing")
            return
        if memory_index_status == 'failed':
            logger.warning('CTI Memory indexing failed for source %s', source_id)
            try:
                db.upsert_cross_ref(
                    f'odysafe:source:{source_id}',
                    'source',
                    str(source_id),
                    status='failed',
                )
            except Exception as status_error:
                logger.warning(
                    'Failed to persist CTI Memory indexing status for source %s: %s',
                    source_id,
                    status_error,
                )
        elif memory_index_status == 'timed_out':
            logger.warning(
                'CTI Memory indexing timed out for source %s; the bounded background '
                'operation may still finish later',
                source_id,
            )

        db.update_source_status(source_id, 'completed')
        completion_message = f"Processing completed: {len(iocs_list)} new IOCs saved"
        if memory_index_status == 'failed':
            completion_message += "; CTI Memory indexing failed"
        elif memory_index_status == 'timed_out':
            completion_message += "; CTI Memory indexing timed out"
        progress_tracker.complete_task(
            task_id,
            completion_message,
            result_data={
                'memory_indexed': memory_index_status == 'indexed',
                'memory_index_status': memory_index_status,
            },
        )
    except Exception as e:
        logger.error("IOC extraction error for source %s: %s", source_id, e, exc_info=True)
        try:
            db.update_source_status(source_id, 'error')
        except Exception as status_error:
            logger.error(
                "Failed to mark source %s as error after IOC processing failure: %s",
                source_id,
                status_error,
                exc_info=True,
            )
        progress_tracker.error_task(task_id, f"Error: {str(e)}")
    finally:
        cleanup_timer = threading.Timer(
            300.0,
            progress_tracker.remove_task,
            args=(task_id,),
        )
        cleanup_timer.daemon = True
        cleanup_timer.start()


def cleanup_old_files():
    """Cleans up files in uploads/ older than CLEANUP_DAYS days"""
    try:
        cutoff_date = datetime.now() - timedelta(days=CLEANUP_DAYS)
        deleted_count = 0
        
        for file_path in UPLOAD_FOLDER.iterdir():
            if file_path.is_file():
                file_time = datetime.fromtimestamp(file_path.stat().st_mtime)
                if file_time < cutoff_date:
                    try:
                        file_path.unlink()
                        deleted_count += 1
                        logger.info(f"File deleted: {file_path}")
                    except Exception as e:
                        logger.warning(f"Unable to delete {file_path}: {e}")
        
        if deleted_count > 0:
            logger.info(f"Cleanup completed: {deleted_count} file(s) deleted")
    except Exception as e:
        logger.error(f"Cleanup error: {e}")


def checkpoint_database_wal():
    """Checkpoint SQLite's WAL without deleting files managed by SQLite itself."""
    try:
        db_path = Path(db.db_path)
        with sqlite3.connect(str(db_path), timeout=5.0) as conn:
            busy, log_frames, checkpointed_frames = conn.execute(
                "PRAGMA wal_checkpoint(PASSIVE)"
            ).fetchone()

        if busy:
            logger.info(
                "SQLite WAL checkpoint deferred: database is busy "
                "(%s/%s frames checkpointed)",
                checkpointed_frames,
                log_frames,
            )
        elif log_frames:
            logger.info(
                "SQLite WAL checkpoint completed: %s/%s frames checkpointed",
                checkpointed_frames,
                log_frames,
            )
    except sqlite3.OperationalError as exc:
        logger.warning("SQLite WAL checkpoint skipped: %s", exc)
    except Exception as exc:
        logger.error("SQLite WAL checkpoint failed: %s", exc)


# Start automatic cleanup in background
def start_cleanup_scheduler():
    """Starts the cleanup scheduler"""
    def cleanup_loop():
        while True:
            cleanup_old_files()
            checkpoint_database_wal()
            # Wait 24 hours before next cleanup
            time.sleep(86400)  # 24 hours
    
    cleanup_thread = threading.Thread(target=cleanup_loop, daemon=True)
    cleanup_thread.start()
    logger.info("Cleanup scheduler started")

start_cleanup_scheduler()

# ========== AUTHENTICATION ROUTES ==========

@app.route('/login', methods=['GET', 'POST'])
def login():
    """Login page"""
    # Si l'auth n'est pas activée, rediriger vers le dashboard
    if not is_auth_enabled(db):
        return redirect(url_for('dashboard'))
    
    # Si déjà connecté, rediriger vers le dashboard
    if 'user_id' in session:
        return redirect(url_for('dashboard'))
    
    if request.method == 'POST':
        data = request.get_json() if request.is_json else request.form
        username = data.get('username', '').strip()
        password = data.get('password', '')
        
        if not username or not password:
            if request.is_json:
                return api_error("Username and password are required", 400)
            flash("Username and password are required", "error")
            return render_template('login.html')
        
        success, result = verify_user(db, username, password)
        if success:
            session.permanent = True  # Session permanente (pas de timeout)
            session['user_id'] = result
            session['username'] = username
            if request.is_json:
                return api_success({"message": "Login successful"})
            return redirect(url_for('dashboard'))
        else:
            if request.is_json:
                return api_error(result, 401)
            flash(result, "error")
            return render_template('login.html')
    
    return render_template('login.html')

@app.route('/logout', methods=['POST'])
def logout():
    """Logout"""
    session.clear()
    if request.is_json:
        return api_success({"message": "Logged out successfully"})
    return redirect(url_for('login'))

@app.route('/api/auth/status', methods=['GET'])
def api_auth_status():
    """API: Check authentication status"""
    try:
        enabled = is_auth_enabled(db)
        is_logged_in = 'user_id' in session
        username = get_current_username() if is_logged_in else None
        has_users = user_exists(db, 'odysafe')
        
        return api_success({
            "auth_enabled": enabled,
            "is_logged_in": is_logged_in,
            "has_users": has_users,
            "username": username
        })
    except Exception as e:
        logger.error(f"api_auth_status error: {e}")
        return api_error(str(e), 500)

@app.route('/api/auth/enable', methods=['POST'])
def api_auth_enable():
    """API: Enable authentication when odysafe account already exists"""
    try:
        if is_auth_enabled(db):
            return api_success({"auth_enabled": True, "message": "Authentication is already enabled"})
        if not user_exists(db, 'odysafe'):
            return api_error("Set the odysafe password first before enabling authentication", 400)
        db.set_setting('auth_enabled', 'true')
        return api_success({"auth_enabled": True, "message": "Authentication enabled"})
    except Exception as e:
        logger.error(f"api_auth_enable error: {e}")
        return api_error(str(e), 500)


@app.route('/api/auth/create-user', methods=['POST'])
def api_auth_create_user():
    """API: Create odysafe user and enable authentication"""
    try:
        data = request.get_json() or {}
        username = data.get('username', 'odysafe').strip() or 'odysafe'
        password = data.get('password', '')
        password_confirm = data.get('password_confirm', '')
        
        if not password:
            return api_error("Password and confirmation are required", 400)
        
        if password != password_confirm:
            return api_error("Passwords do not match", 400)
        
        if len(password) < 8:
            return api_error("Password must be at least 8 characters long", 400)
        
        # Vérifier que le username est "odysafe"
        if username != 'odysafe':
            return api_error("Username must be 'odysafe'", 400)
        
        # Vérifier si l'utilisateur existe déjà
        if user_exists(db, username):
            return api_error("User already exists", 400)
        
        # Créer l'utilisateur
        success, message = create_user(db, username, password)
        if success:
            # Activer l'authentification après création de l'utilisateur
            db.set_setting('auth_enabled', 'true')
            return api_success({
                "message": message,
                "auth_enabled": True
            })
        else:
            return api_error(message, 400)
    
    except Exception as e:
        logger.error(f"api_auth_create_user error: {e}")
        return api_error(str(e), 500)

@app.route('/api/auth/change-password', methods=['POST'])
@require_auth
def api_auth_change_password():
    """API: Change password (requires authentication)"""
    try:
        if 'username' not in session:
            return api_error("Not authenticated", 401)
        
        data = request.get_json()
        old_password = data.get('old_password', '')
        new_password = data.get('new_password', '')
        new_password_confirm = data.get('new_password_confirm', '')
        
        if not old_password or not new_password:
            return api_error("Old password and new password are required", 400)
        
        if new_password != new_password_confirm:
            return api_error("New passwords do not match", 400)
        
        if len(new_password) < 8:
            return api_error("New password must be at least 8 characters long", 400)
        
        username = session['username']
        success, message = change_password(db, username, old_password, new_password)
        if success:
            return api_success({"message": message})
        else:
            return api_error(message, 400)
    
    except Exception as e:
        logger.error(f"api_auth_change_password error: {e}")
        return api_error(str(e), 500)

@app.route('/api/auth/toggle', methods=['POST'])
@require_auth
def api_auth_toggle():
    """API: Disable authentication (enable via create-user)"""
    try:
        data = request.get_json()
        enabled = data.get('enabled', False)
        
        # On ne peut que désactiver via cette route (l'activation se fait via create-user)
        if enabled:
            return api_error("To enable authentication, please create a user first", 400)
        
        # Vérifier si l'auth est activée avant de désactiver
        current_auth_enabled = is_auth_enabled(db)
        if not current_auth_enabled:
            return api_error("Authentication is already disabled", 400)
        
        # Désactiver l'authentification
        db.set_setting('auth_enabled', 'false')
        
        # Déconnecter tous les utilisateurs
        session.clear()
        
        return api_success({
            "message": "Authentication disabled",
            "auth_enabled": False
        })
    
    except Exception as e:
        logger.error(f"api_auth_toggle error: {e}")
        return api_error(str(e), 500)

# ========== ROUTES GET ==========

@app.route('/')
@require_auth
def dashboard():
    """Home page with statistics"""
    try:
        stats = db.get_statistics()
        recent_limit = int(db.get_setting('recent_sources_limit', '20'))
        recent_sources = db.get_all_sources(limit=recent_limit)
        
        # CTI Resources statistics
        cti_stats = {}
        
        # DeepDarkCTI stats
        try:
            ddc_repo_exists = github_repo_manager.repo_exists()
            ddc_category_info = github_repo_manager.get_category_info()
            if ddc_repo_exists:
                categories = github_repo_manager.fetch_all_categories()
                cti_stats['deepdarkcti'] = {
                    'available': True,
                    'categories_count': len(categories) if categories else 0,
                    'last_update': ddc_category_info.get('last_update') if ddc_category_info else None
                }
            else:
                cti_stats['deepdarkcti'] = {
                    'available': False,
                    'categories_count': 0,
                    'last_update': None
                }
        except Exception as e:
            logger.error(f"Error fetching DeepDarkCTI stats: {e}")
            cti_stats['deepdarkcti'] = {
                'available': False,
                'categories_count': 0,
                'last_update': None
            }
        
        # Ransomware Tool Matrix stats
        try:
            rtm_repo_exists = rtm_repo_manager.repo_exists()
            if rtm_repo_exists:
                rtm_category_info = rtm_repo_manager.get_category_info()
                tools = rtm_repo_manager.fetch_tools()
                threat_intel = rtm_repo_manager.fetch_threat_intel()
                group_profiles = rtm_repo_manager.fetch_group_profiles()
                community_reports = rtm_repo_manager.fetch_community_reports()
                cti_stats['ransomware_matrix'] = {
                    'available': True,
                    'tools_count': sum(len(tool_data.get('tools', [])) for tool_data in tools.values()) if tools else 0,
                    'threat_intel_count': sum(len(intel_list) for intel_list in threat_intel.values()) if threat_intel else 0,
                    'group_profiles_count': len(group_profiles) if group_profiles else 0,
                    'community_reports_count': len(community_reports) if community_reports else 0,
                    'last_update': rtm_category_info.get('last_update') if rtm_category_info else None
                }
            else:
                cti_stats['ransomware_matrix'] = {
                    'available': False,
                    'tools_count': 0,
                    'threat_intel_count': 0,
                    'group_profiles_count': 0,
                    'community_reports_count': 0,
                    'last_update': None
                }
        except Exception as e:
            logger.error(f"Error fetching Ransomware Tool Matrix stats: {e}")
            cti_stats['ransomware_matrix'] = {
                'available': False,
                'tools_count': 0,
                'threat_intel_count': 0,
                'group_profiles_count': 0,
                'community_reports_count': 0,
                'last_update': None
            }
        
        # Data-Shield IPv4 Blocklist stats
        try:
            ds_status = data_shield_repo_manager.get_status()
            
            # Count imported IPs from database
            imported_count = 0
            try:
                with db.connection() as conn:
                    ody = conn.cursor()
                    ody.execute("""
                        SELECT COUNT(*) as count FROM iocs i
                        JOIN sources s ON i.source_id = s.id
                        WHERE s.name = 'data-shield' AND s.is_deleted = 0
                    """)
                    row = ody.fetchone()
                    if row:
                        imported_count = row['count']
            except Exception as e:
                logger.warning(f"Error counting imported IPs: {e}")
            
            # Count IPs added in the last 24 hours from data-shield source
            delta_24h = 0
            try:
                with db.connection() as conn2:
                    row2 = conn2.execute("""
                        SELECT COUNT(*) AS cnt FROM iocs i
                        JOIN sources s ON i.source_id = s.id
                        WHERE s.name = 'data-shield' AND i.is_deleted = 0
                          AND i.created_at >= datetime('now', '-1 day')
                    """).fetchone()
                    if row2:
                        delta_24h = row2['cnt']
            except Exception:
                pass

            cti_stats['data_shield'] = {
                'available': ds_status.get('repo_exists', False),
                'blocklist_exists': ds_status.get('blocklist_exists', False),
                'ip_count': ds_status.get('ip_count', 0),
                'imported_ip_count': imported_count,
                'last_update': ds_status.get('last_update'),
                'delta_24h': delta_24h,
            }
        except Exception as e:
            logger.error(f"Error fetching Data-Shield stats: {e}")
            cti_stats['data_shield'] = {
                'available': False,
                'blocklist_exists': False,
                'ip_count': 0,
                'imported_ip_count': 0,
                'last_update': None,
                'delta_24h': 0,
            }
        
        # MITRE ATT&CK stats
        try:
            # Check if MITRE data is available by checking if the database has MITRE data
            with db.connection() as conn:
                result = conn.execute("SELECT COUNT(*) as count FROM mitre_tactics").fetchone()
                mitre_available = result['count'] > 0 if result else False
                
                if mitre_available:
                    # Get MITRE statistics
                    tactics_count = conn.execute("SELECT COUNT(*) as count FROM mitre_tactics").fetchone()['count']
                    techniques_count = conn.execute("SELECT COUNT(*) as count FROM mitre_techniques").fetchone()['count']
                    groups_count = conn.execute("SELECT COUNT(*) as count FROM mitre_groups").fetchone()['count']
                    
                    cti_stats['mitre_attack'] = {
                        'available': True,
                        'tactics_count': tactics_count,
                        'techniques_count': techniques_count,
                        'groups_count': groups_count,
                        'last_import': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    }
                else:
                    cti_stats['mitre_attack'] = {
                        'available': False,
                        'tactics_count': 0,
                        'techniques_count': 0,
                        'groups_count': 0,
                        'last_import': None
                    }
        except Exception as e:
            logger.error(f"Error fetching MITRE ATT&CK stats: {e}")
            cti_stats['mitre_attack'] = {
                'available': False,
                'tactics_count': 0,
                'techniques_count': 0
            }
        
        return render_template('dashboard.html', stats=stats, recent_sources=recent_sources, cti_stats=cti_stats)
    except Exception as e:
        logger.error(f"Dashboard error: {e}")
        flash("Error loading dashboard", "error")
        return render_template('dashboard.html', stats={}, recent_sources=[], cti_stats={})

@app.route('/upload')
@require_auth
def upload():
    """Upload/import page"""
    return render_template('upload.html')

@app.route('/iocs')
@require_auth
def iocs_list():
    """IOC list with filters"""
    try:
        from modules.ioc_query_urls import get_query_urls
        
        # Retrieve filtering parameters
        ioc_type = request.args.get('type', '').strip()
        search = request.args.get('search', '').strip()
        ioc_value = request.args.get('value', '').strip()
        source_name = request.args.get('source', '').strip()
        date_range = request.args.get('date_range', '')
        date_from = request.args.get('date_from', '')
        date_to = request.args.get('date_to', '')
        page = int(request.args.get('page', 1))
        per_page = 30
        
        filters = {}
        if ioc_type:
            filters['ioc_type'] = ioc_type
        if ioc_value:
            filters['ioc_value'] = ioc_value
        elif search:
            filters['search'] = search
        if source_name:
            filters['source_name'] = source_name
        group_id = request.args.get('group', '').strip()
        if group_id:
            try:
                filters['group_id'] = int(group_id)
            except ValueError:
                pass
        # Duplicate filter
        show_duplicates = request.args.get('duplicates', '').strip()
        if show_duplicates == 'true':
            filters['show_duplicates'] = True
        if date_range:
            filters['date_range'] = date_range
            # Calculate date_from and date_to based on date_range
            from datetime import datetime, timedelta
            now = datetime.now()
            if date_range == '24h':
                filters['date_from'] = (now - timedelta(hours=24)).strftime('%Y-%m-%d')
            elif date_range == '7d':
                filters['date_from'] = (now - timedelta(days=7)).strftime('%Y-%m-%d')
            elif date_range == '30d':
                filters['date_from'] = (now - timedelta(days=30)).strftime('%Y-%m-%d')
            elif date_range == '3m':
                filters['date_from'] = (now - timedelta(days=90)).strftime('%Y-%m-%d')
            elif date_range == '1y':
                filters['date_from'] = (now - timedelta(days=365)).strftime('%Y-%m-%d')
            elif date_range == 'custom':
                if date_from:
                    filters['date_from'] = date_from
                if date_to:
                    filters['date_to'] = date_to
        else:
            if date_from:
                filters['date_from'] = date_from
            if date_to:
                filters['date_to'] = date_to
        
        offset = (page - 1) * per_page
        iocs, total = db.get_all_iocs(filters=filters, limit=per_page, offset=offset)
        
        # Enrich each IOC with its query URLs
        for ioc in iocs:
            ioc['query_urls'] = get_query_urls(ioc['ioc_type'], ioc['ioc_value'])
        
        # Retrieve unique IOC types present in the database
        unique_types = db.get_unique_ioc_types()
        
        # Retrieve unique source names
        unique_source_names = db.get_unique_source_names()
        
        # Retrieve all groups for the filter
        all_groups = db.get_all_groups()
        
        return render_template('iocs_list.html', 
                             iocs=iocs, 
                             total=total,
                             page=page,
                             per_page=per_page,
                             filters=filters,
                             unique_types=unique_types,
                             unique_source_names=unique_source_names,
                             all_groups=all_groups)
    except Exception as e:
        logger.error(f"IOCs list error: {e}")
        flash("Error loading IOCs", "error")
        return render_template('iocs_list.html', iocs=[], total=0, page=1, per_page=30, 
                             filters={}, unique_types=[], unique_source_names=[])

@app.route('/ioc/<int:ioc_id>')
@require_auth
def ioc_detail(ioc_id):
    """IOC detail page"""
    try:
        ioc = db.get_ioc(ioc_id)
        if not ioc:
            flash("IOC not found", "error")
            return redirect(url_for('iocs_list'))
        
        source = db.get_source(ioc['source_id'])
        tag_history = db.get_tag_history(ioc_id)
        all_tags = db.get_all_tags(only_with_iocs=False)

        current_tag_ids = {t['id'] for t in ioc.get('tags', [])}
        manual_tags = [t for t in all_tags if not t.get('is_auto') and t['id'] not in current_tag_ids]
        auto_tags   = [t for t in ioc.get('tags', []) if t.get('is_auto')]
        ttp_links = db.get_ioc_ttp_links(ioc_id)
        manual_applied = [t for t in ioc.get('tags', []) if not t.get('is_auto')]

        return render_template('ioc_detail.html',
                             ioc=ioc,
                             source=source,
                             tag_history=tag_history,
                             manual_tags=manual_tags,
                             auto_tags=auto_tags,
                             manual_applied=manual_applied,
                             ttp_links=ttp_links)
    except Exception as e:
        logger.error(f"IOC detail error: {e}")
        flash("Error loading IOC", "error")
        return redirect(url_for('iocs_list'))

@app.route('/sources')
@require_auth
def sources_list():
    """Sources list"""
    try:
        sources = db.get_all_sources(limit=1000)
        all_groups = db.get_all_groups()
        ioc_type_counts = db.get_ioc_type_counts_by_source()
        for s in sources:
            s['ioc_type_counts'] = ioc_type_counts.get(s['id'], [])
        return render_template('sources_list.html', sources=sources, all_groups=all_groups)
    except Exception as e:
        logger.error(f"Sources list error: {e}")
        flash("Error loading sources", "error")
        return render_template('sources_list.html', sources=[], all_groups=[])


@app.route('/export')
@require_auth
def export():
    """Export page"""
    try:
        # Limit sources to reasonable number for dropdown (most recent first)
        sources = db.get_all_sources(limit=500)
        all_groups = db.get_all_groups()
        _, total = db.get_all_iocs(limit=1, offset=0)
        unique_types = db.get_unique_ioc_types()
        ioc_type_counts = db.get_ioc_type_counts_by_source()
        for s in sources:
            s['ioc_type_counts'] = ioc_type_counts.get(s['id'], [])
        all_tags = db.get_all_tags(only_with_iocs=True)
        manual_tags = [t for t in all_tags if not t.get('is_auto')]
        return render_template('export.html', sources=sources, all_groups=all_groups,
                               total_iocs=total, unique_types=unique_types, manual_tags=manual_tags)
    except Exception as e:
        logger.error(f"Export error: {e}")
        return render_template('export.html', sources=[], all_groups=[], total_iocs=0,
                               unique_types=[], manual_tags=[])

@app.route('/stix-graph')
@require_auth
def stix_graph():
    """STIX Graph Analyzer page - Interactive visualization of STIX bundles"""
    return render_template('stix_graph.html', stix_producer_name=db.get_setting('stix_producer_name', 'ODYSAFE CTI'))

@app.route('/settings')
@require_auth
def settings():
    """Settings page"""
    return render_template('settings.html')

@app.route('/onboarding')
@app.route('/welcome')
@require_auth
def onboarding():
    """Product guide · overview of all modules and workflows."""
    return render_template('onboarding.html')

@app.route('/intelligence')
@app.route('/cti-toolkit')
@require_auth
def cti_toolkit():
    """Analysis hub · central page for analysis and research tools"""
    return render_template('intelligence_hub.html')


@app.route('/cti-resources')
@app.route('/cti-resources/deepdarkcti')
@require_auth
def cti_resources():
    """CTI Resources page - Main page with resource selector
    
    IMPORTANT: This route ONLY reads existing repository data.
    It does NOT trigger automatic download. Download must be done
    explicitly via /api/cti-resources/download or /api/cti-resources/update endpoints.
    """
    try:
        repo_exists = github_repo_manager.repo_exists()
        
        # Load categories once (optimization)
        # NOTE: fetch_all_categories() only reads from local files, never downloads
        categories = github_repo_manager.fetch_all_categories()
        
        # Get category info (includes last_update from cache, not current time)
        category_info = github_repo_manager.get_category_info()
        
        return render_template('deepdarkcti.html', 
                             repo_exists=repo_exists,
                             categories=categories, 
                             category_info=category_info)
    except Exception as e:
        logger.error(f"DeepDarkCTI error: {e}")
        flash("Error loading DeepDarkCTI data", "error")
        return render_template('deepdarkcti.html', 
                             repo_exists=False,
                             categories={}, 
                             category_info={})

@app.route('/cti-resources/ransomware-matrix')
@require_auth
def cti_resources_ransomware_matrix():
    """Ransomware Tool Matrix page within CTI Resources
    
    IMPORTANT: This route ONLY reads existing repository data.
    It does NOT trigger automatic download. Download must be done
    explicitly via /api/ransomware-tools/download or /api/ransomware-tools/update endpoints.
    """
    try:
        repo_exists = rtm_repo_manager.repo_exists()
        
        # Always try to load data, even if repo_exists is False (in case repo was just downloaded)
        tools = {}
        threat_intel = {}
        group_profiles = []
        community_reports = []
        
        if repo_exists:
            try:
                tools = rtm_repo_manager.fetch_tools()
            except Exception as e:
                logger.error(f"Error fetching tools: {e}")
            
            try:
                threat_intel = rtm_repo_manager.fetch_threat_intel()
            except Exception as e:
                logger.error(f"Error fetching threat intel: {e}")
            
            try:
                group_profiles = rtm_repo_manager.fetch_group_profiles()
            except Exception as e:
                logger.error(f"Error fetching group profiles: {e}")
            
            try:
                community_reports = rtm_repo_manager.fetch_community_reports()
            except Exception as e:
                logger.error(f"Error fetching community reports: {e}")
        
        # Re-check repo_exists after loading data (in case it was just downloaded)
        repo_exists = rtm_repo_manager.repo_exists()
        
        # Get category info (includes last_update from cache)
        category_info = rtm_repo_manager.get_category_info()
        
        return render_template('ransomware_tools.html', 
                             repo_exists=repo_exists,
                             tools=tools,
                             threat_intel=threat_intel,
                             group_profiles=group_profiles,
                             community_reports=community_reports,
                             category_info=category_info)
    except Exception as e:
        logger.error(f"Ransomware Tools error: {e}", exc_info=True)
        flash("Error loading Ransomware Tool Matrix data", "error")
        # Try to return with empty data but still show the page
        try:
            repo_exists = rtm_repo_manager.repo_exists()
            category_info = rtm_repo_manager.get_category_info()
        except:
            repo_exists = False
            category_info = {}
        return render_template('ransomware_tools.html', 
                             repo_exists=repo_exists,
                             tools={},
                             threat_intel={},
                             group_profiles=[],
                             community_reports=[],
                             category_info=category_info)

@app.route('/ransomware-tools')
@require_auth
def ransomware_tools():
    """Ransomware Tool Matrix page
    
    IMPORTANT: This route ONLY reads existing repository data.
    It does NOT trigger automatic download. Download must be done
    explicitly via /api/ransomware-tools/download or /api/ransomware-tools/update endpoints.
    """
    try:
        repo_exists = rtm_repo_manager.repo_exists()
        
        # Always try to load data, even if repo_exists is False (in case repo was just downloaded)
        tools = {}
        threat_intel = {}
        group_profiles = []
        community_reports = []
        
        if repo_exists:
            try:
                tools = rtm_repo_manager.fetch_tools()
            except Exception as e:
                logger.error(f"Error fetching tools: {e}")
            
            try:
                threat_intel = rtm_repo_manager.fetch_threat_intel()
            except Exception as e:
                logger.error(f"Error fetching threat intel: {e}")
            
            try:
                group_profiles = rtm_repo_manager.fetch_group_profiles()
            except Exception as e:
                logger.error(f"Error fetching group profiles: {e}")
            
            try:
                community_reports = rtm_repo_manager.fetch_community_reports()
            except Exception as e:
                logger.error(f"Error fetching community reports: {e}")
        
        # Re-check repo_exists after loading data (in case it was just downloaded)
        repo_exists = rtm_repo_manager.repo_exists()
        
        # Get category info (includes last_update from cache)
        category_info = rtm_repo_manager.get_category_info()
        
        return render_template('ransomware_tools.html', 
                             repo_exists=repo_exists,
                             tools=tools,
                             threat_intel=threat_intel,
                             group_profiles=group_profiles,
                             community_reports=community_reports,
                             category_info=category_info)
    except Exception as e:
        logger.error(f"Ransomware Tools error: {e}", exc_info=True)
        flash("Error loading Ransomware Tool Matrix data", "error")
        # Try to return with empty data but still show the page
        try:
            repo_exists = rtm_repo_manager.repo_exists()
            category_info = rtm_repo_manager.get_category_info()
        except:
            repo_exists = False
            category_info = {}
        return render_template('ransomware_tools.html', 
                             repo_exists=repo_exists,
                             tools={},
                             threat_intel={},
                             group_profiles=[],
                             community_reports=[],
                             category_info=category_info)

@app.route('/cti-resources/mitre-attack')
@require_auth
def cti_resources_mitre_attack():
    """MITRE ATT&CK browser page within CTI Resources"""
    file_exists = mitre_json_exists()
    matrix_view = None
    stats = {}
    groups = []
    
    if file_exists:
        parser = MitreStixParser()
        if parser.load():
            matrix_view = parser.get_matrix_view()
            stats = parser.get_stats()
            groups = parser.get_groups()
    
    from modules.pinned_sources import get_mitre_enterprise_attack_url

    return render_template('mitre_attack.html',
                         file_exists=file_exists,
                         matrix_view=matrix_view,
                         stats=stats,
                         groups=groups,
                         mitre_pinned_url=get_mitre_enterprise_attack_url())


# ========== ROUTES API ==========

@app.route('/api/upload', methods=['POST'])
@require_auth
def api_upload():
    """API: File upload"""
    try:
        if 'file' not in request.files:
            return api_error('No file provided', 400)
        
        file = request.files['file']
        if file.filename == '':
            return api_error('No file selected', 400)
        
        if not allowed_file(file.filename):
            return api_error('File type not allowed', 400)
        
        name = request.form.get('name', '').strip()
        context = request.form.get('context', '').strip()
        
        if not name:
            return api_error('Source name is required', 400)
        
        # Auto-generate context if empty
        if not context:
            context = generate_default_context('file_upload')
        
        # Validate file before saving to disk
        filename = secure_filename(file.filename)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        unique_filename = f"{timestamp}_{filename}"
        file_path = UPLOAD_FOLDER / unique_filename
        
        # Save to temporary file first for validation
        temp_file_path = None
        try:
            import tempfile
            temp_file_path = Path(tempfile.mkstemp(suffix=file_path.suffix)[1])
            file.save(str(temp_file_path))
            
            # Validate MIME type before moving to final location
            if not validate_file_mime(temp_file_path):
                return api_error('File type validation failed: MIME type not allowed', 400)
            
            # Move validated file to final location
            temp_file_path.rename(file_path)
            temp_file_path = None  # Prevent cleanup after successful move
            
        except Exception as e:
            logger.error(f"Error during file validation: {e}")
            return api_error(f'File processing error: {str(e)}', 500)
        finally:
            # Cleanup temporary file if it still exists
            if temp_file_path and temp_file_path.exists():
                try:
                    temp_file_path.unlink()
                except Exception as e:
                    logger.warning(f"Failed to delete temporary file {temp_file_path}: {e}")
        
        # Create source in database
        source_id = db.create_source(
            name=name,
            context=context,
            source_type='file_upload',
            file_path=str(file_path),
            original_filename=filename
        )
        
        # Check if auto-rotation is enabled and apply if needed
        check_and_apply_auto_rotation(db)
        
        # Update status
        db.update_source_status(source_id, 'processing')
        
        # Start progress tracking
        progress_tracker.start_task(f"source_{source_id}", "source_processing", total_steps=100)
        progress_tracker.update_progress(f"source_{source_id}", percentage=10, message="File uploaded, starting extraction...")
        
        # Extract IOCs in background
        thread = threading.Thread(
            target=_extract_iocs_background,
            args=(
                source_id,
                extract_iocs,
                (str(file_path), None, True),
                "Extracting IOCs...",
                20,
            ),
            daemon=True,
        )
        thread.start()
        
        return api_success(
            {'source_id': source_id},
            'File uploaded successfully, extraction in progress...'
        )
    
    except RequestEntityTooLarge:
        return api_error('File too large', 413)
    except Exception as e:
        logger.error(f"api_upload error: {e}")
        return api_error(str(e), 500)

@app.route('/api/paste', methods=['POST'])
@require_auth
def api_paste():
    """API: Paste text processing"""
    try:
        data = request.get_json()
        text_content = data.get('text', '')
        name = data.get('name', '').strip()
        context = data.get('context', '').strip()
        
        if not text_content:
            return api_error('No text provided', 400)
        
        if not name:
            return api_error('Source name is required', 400)
        
        # Auto-generate context if empty
        if not context:
            context = generate_default_context('paste')
        
        if not IOCSEARCHER_AVAILABLE:
            return api_error('iocsearcher is not available', 503)
        
        # Create temporary file for text
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        temp_filename = f"paste_{timestamp}.txt"
        file_path = UPLOAD_FOLDER / temp_filename
        
        with open(file_path, 'w', encoding='utf-8') as f:
            f.write(text_content)
        
        # Create the source in the database
        source_id = db.create_source(
            name=name,
            context=context,
            source_type='paste',
            file_path=str(file_path),
            original_filename=temp_filename
        )
        
        # Check if auto-rotation is enabled
        auto_rotation = db.get_setting('auto_rotation_enabled', 'false').lower() == 'true'
        if auto_rotation:
            max_sources = int(db.get_setting('max_sources', '20'))
            deleted_count = db.rotate_sources_if_needed(max_sources)
            if deleted_count > 0:
                logger.info(f"Auto-rotation: {deleted_count} oldest source(s) deleted")
        
        db.update_source_status(source_id, 'processing')
        
        # Start progress tracking
        progress_tracker.start_task(f"source_{source_id}", "source_processing", total_steps=100)
        progress_tracker.update_progress(f"source_{source_id}", percentage=10, message="Text received, starting extraction...")
        
        # Extract IOCs in background
        thread = threading.Thread(
            target=_extract_iocs_background,
            args=(source_id, extract_from_text, (text_content,), "Extracting IOCs...", 20),
            daemon=True,
        )
        thread.start()
        
        return api_success(
            {'source_id': source_id},
            'Text processed successfully, extraction in progress...'
        )
    
    except Exception as e:
        logger.error(f"api_paste error: {e}")
        return api_error(str(e), 500)

@app.route('/api/url', methods=['POST'])
@require_auth
def api_url():
    """API: URL processing"""
    try:
        data = request.get_json()
        url = data.get('url', '')
        name = data.get('name', '').strip()
        context = data.get('context', '').strip()
        
        if not url:
            return api_error('No URL provided', 400)
        
        if not name:
            return api_error('Source name is required', 400)
        
        # Auto-generate context if empty
        if not context:
            context = generate_default_context('url')
        
        if not IOCSEARCHER_AVAILABLE:
            return api_error('iocsearcher is not available', 503)
        
        # Create source in database
        source_id = db.create_source(
            name=name,
            context=context,
            source_type='url',
            file_path=None,
            original_filename=url
        )
        
        # Check if auto-rotation is enabled and apply if needed
        check_and_apply_auto_rotation(db)
        
        db.update_source_status(source_id, 'processing')
        
        # Start progress tracking
        progress_tracker.start_task(f"source_{source_id}", "source_processing", total_steps=100)
        progress_tracker.update_progress(f"source_{source_id}", percentage=10, message="URL received, downloading...")
        
        # Extract IOCs in background
        thread = threading.Thread(
            target=_extract_iocs_background,
            args=(
                source_id,
                extract_from_url,
                (url, None, True),
                "Download completed, extracting IOCs...",
                30,
            ),
            daemon=True,
        )
        thread.start()
        
        return api_success(
            {'source_id': source_id},
            'URL processed successfully, extraction in progress...'
        )
    
    except Exception as e:
        logger.error(f"api_url error: {e}")
        return api_error(str(e), 500)

@app.route('/api/ioc/<int:ioc_id>/tag', methods=['POST', 'DELETE'])
@require_auth
def api_ioc_tag(ioc_id):
    """API: Add/remove a tag from an IOC"""
    try:
        data = request.get_json()
        tag_name = data.get('tag_name', '')
        
        if not tag_name:
            return api_error('Tag name required', 400)
        
        # Retrieve or create the tag
        all_tags = db.get_all_tags(only_with_iocs=False)
        tag = next((t for t in all_tags if t['name'] == tag_name), None)

        if not tag:
            # Create a new custom tag
            tag_id = db.create_tag(tag_name, category='custom')
            tag = {'id': tag_id, 'name': tag_name, 'category': 'custom', 'is_auto': 0}
        else:
            tag_id = tag['id']

        if request.method == 'POST':
            db.add_tag_to_ioc(ioc_id, tag_id)
            return api_success({'tag': tag}, message='Tag added')
        else:  # DELETE
            db.remove_tag_from_ioc(ioc_id, tag_id)
            return api_success(message='Tag removed')
    
    except Exception as e:
        logger.error(f"api_ioc_tag error: {e}")
        return api_error(str(e), 500)

@app.route('/api/ioc/<int:ioc_id>/notes', methods=['POST'])
@require_auth
def api_ioc_notes(ioc_id):
    """API: Update IOC notes"""
    try:
        data = request.get_json()
        notes = data.get('notes', '')
        
        db.update_ioc_notes(ioc_id, notes)

        memory_index_status = 'failed'
        try:
            from modules import zettelforge_bridge as zf
            ioc = db.get_ioc(ioc_id)
            if ioc:
                memory_index_status = zf.remember_ioc_note(
                    ioc_id,
                    ioc.get('ioc_value', ''),
                    ioc.get('ioc_type', ''),
                    notes,
                )
            if memory_index_status == 'failed':
                logger.warning('CTI memory IOC note indexing failed for IOC %s', ioc_id)
        except Exception as mem_err:
            logger.warning('CTI memory IOC note indexing failed for IOC %s: %s', ioc_id, mem_err)
            try:
                db.upsert_cross_ref(
                    f'odysafe:ioc:{ioc_id}',
                    'ioc',
                    str(ioc_id),
                    status='failed',
                )
            except Exception as status_error:
                logger.warning(
                    'Failed to persist CTI memory indexing status for IOC %s: %s',
                    ioc_id,
                    status_error,
                )
        
        return api_success({
            'memory_indexed': memory_index_status == 'indexed',
            'memory_index_status': memory_index_status,
        }, message='Notes updated')
    
    except Exception as e:
        logger.error(f"api_ioc_notes error: {e}")
        return api_error(str(e), 500)


@app.route('/api/ioc/<int:ioc_id>/correlations', methods=['GET'])
@require_auth
def api_ioc_correlations(ioc_id):
    """Cross-module index: MITRE links, memory status, duplicate instances."""
    try:
        ioc = db.get_ioc(ioc_id)
        if not ioc or ioc.get('is_deleted'):
            return api_error('IOC not found', 404)
        from modules import zettelforge_bridge as zf
        data = zf.get_ioc_correlations(
            ioc_id,
            ioc.get('ioc_type', ''),
            ioc.get('ioc_value', ''),
        )
        cross_ref_status = (data.get('index') or {}).get('status')
        if cross_ref_status == 'active':
            memory_index_status = 'indexed'
        elif cross_ref_status == 'failed':
            memory_index_status = 'failed'
        else:
            memory_index_status = 'never_indexed'
        data['memory_indexed'] = memory_index_status == 'indexed'
        data['memory_index_status'] = memory_index_status
        data['ioc_id'] = ioc_id
        data['ioc'] = {
            'id': ioc_id,
            'ioc_type': ioc.get('ioc_type'),
            'ioc_value': ioc.get('ioc_value'),
            'tlp': ioc.get('tlp'),
            'validation_status': ioc.get('validation_status'),
            'confidence': ioc.get('confidence'),
            'notes': ioc.get('notes'),
            'first_seen': ioc.get('first_seen'),
            'last_seen': ioc.get('last_seen'),
        }
        data['source'] = {
            'id': ioc.get('source_id'),
            'name': (db.get_source(ioc['source_id']) or {}).get('name'),
        }
        return api_success(data)
    except Exception as e:
        logger.error(f"api_ioc_correlations error: {e}", exc_info=True)
        return api_error(str(e), 500)

@app.route('/api/export/txt', methods=['POST'])
@require_auth
def api_export_txt():
    """API: Export TXT"""
    try:
        data = request.get_json()
        output_file = export_txt(db, data, OUTPUT_FOLDER)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        return send_file(str(output_file), as_attachment=True, download_name=f"iocs_export_{timestamp}.txt")
    except Exception as e:
        logger.error(f"api_export_txt error: {e}")
        return api_error(str(e), 500)

@app.route('/api/export/json', methods=['POST'])
@require_auth
def api_export_json():
    """API: Export JSON interne"""
    try:
        data = request.get_json()
        output_file = export_json(db, data, OUTPUT_FOLDER)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        return send_file(str(output_file), as_attachment=True, download_name=f"iocs_export_{timestamp}.json")
    except Exception as e:
        logger.error(f"api_export_json error: {e}")
        return api_error(str(e), 500)

@app.route('/api/export/csv', methods=['POST'])
@require_auth
def api_export_csv():
    """API: Export CSV"""
    try:
        data = request.get_json()
        output_file = export_csv(db, data, OUTPUT_FOLDER)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        return send_file(str(output_file), as_attachment=True, download_name=f"iocs_export_{timestamp}.csv")
    except Exception as e:
        logger.error(f"api_export_csv error: {e}", exc_info=True)
        return api_error(str(e), 500)

@app.route('/api/export/txt-simple', methods=['POST'])
@require_auth
def api_export_txt_simple():
    """API: Export TXT - values only (Firewall/EDR compatible)"""
    try:
        data = request.get_json()
        output_file = export_txt_simple(db, data, OUTPUT_FOLDER)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        return send_file(str(output_file), as_attachment=True, download_name=f"iocs_export_simple_{timestamp}.txt")
    except Exception as e:
        logger.error(f"api_export_txt_simple error: {e}")
        return api_error(str(e), 500)

@app.route('/api/export/csv-firewall', methods=['POST'])
@require_auth
def api_export_csv_firewall():
    """API: Export CSV - simplified format (Firewall/EDR compatible)"""
    try:
        data = request.get_json()
        output_file = export_csv_firewall(db, data, OUTPUT_FOLDER)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        return send_file(str(output_file), as_attachment=True, download_name=f"iocs_export_firewall_{timestamp}.csv")
    except Exception as e:
        logger.error(f"api_export_csv_firewall error: {e}")
        return api_error(str(e), 500)

@app.route('/api/export/json-simple', methods=['POST'])
@require_auth
def api_export_json_simple():
    """API: Export JSON - simplified format grouped by type"""
    try:
        data = request.get_json()
        output_file = export_json_simple(db, data, OUTPUT_FOLDER)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        return send_file(str(output_file), as_attachment=True, download_name=f"iocs_export_simple_{timestamp}.json")
    except Exception as e:
        logger.error(f"api_export_json_simple error: {e}")
        return api_error(str(e), 500)

@app.route('/api/export/xlsx', methods=['POST'])
@require_auth
def api_export_xlsx():
    """API: Export XLSX with elegant formatting"""
    try:
        data = request.get_json()
        output_file = export_xlsx(db, data, OUTPUT_FOLDER)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        return send_file(str(output_file), as_attachment=True, download_name=f"iocs_export_{timestamp}.xlsx")
    except Exception as e:
        logger.error(f"api_export_xlsx error: {e}")
        return api_error(str(e), 500)

@app.route('/api/quick-export', methods=['POST'])
@require_auth
def api_quick_export():
    """API: Quick export - extract IOCs and export without saving to database"""
    try:
        if not IOCSEARCHER_AVAILABLE:
            return api_error('iocsearcher is not available', 503)
        
        input_type = request.form.get('input_type', '').strip()
        format_type = request.form.get('format', '').strip()
        
        if not input_type or not format_type:
            return api_error('Input type and format are required', 400)
        
        # Extract IOCs based on input type
        ioc_results = []
        temp_file_path = None
        
        try:
            if input_type == 'file':
                if 'file' not in request.files:
                    return api_error('No file provided', 400)
                
                file = request.files['file']
                if file.filename == '':
                    return api_error('No file selected', 400)
                
                if not allowed_file(file.filename):
                    return api_error('File type not allowed', 400)
                
                # Save to temporary file
                filename = secure_filename(file.filename)
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                temp_filename = f"quick_export_{timestamp}_{filename}"
                temp_file_path = UPLOAD_FOLDER / temp_filename
                file.save(str(temp_file_path))
                
                # Validate MIME type
                if not validate_file_mime(temp_file_path):
                    temp_file_path.unlink()
                    return api_error('File type validation failed', 400)
                
                # Extract IOCs
                ioc_results = extract_iocs(str(temp_file_path))
                
            elif input_type == 'paste':
                text_content = request.form.get('text', '').strip()
                if not text_content:
                    return api_error('No text provided', 400)
                
                # Extract IOCs from text
                ioc_results = extract_from_text(text_content)
                
            elif input_type == 'url':
                url = request.form.get('url', '').strip()
                if not url:
                    return api_error('No URL provided', 400)
                
                # Extract IOCs from URL
                ioc_results = extract_from_url(url)
                
            else:
                return api_error('Invalid input type', 400)
            
            if not ioc_results:
                return api_error('No IOCs found in the provided content', 404)
            
            # Convert extraction results to IOC dict format compatible with export functions
            iocs = []
            for ioc_type, ioc_value, raw_value, offset in ioc_results:
                ioc_dict = {
                    'id': len(iocs) + 1,  # Temporary ID
                    'ioc_type': ioc_type,
                    'ioc_value': ioc_value,
                    'raw_value': raw_value or ioc_value,
                    'source_id': 0,  # No source since not saved
                    'source_name': 'Quick Export',
                    'created_at': datetime.now().isoformat(),
                    'first_seen': datetime.now().isoformat(),
                    'last_seen': datetime.now().isoformat(),
                    'tags': []
                }
                iocs.append(ioc_dict)
            
            # Create temporary data structure for export
            export_data = {'ioc_ids': [ioc['id'] for ioc in iocs]}
            
            # Export based on format
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_file = None
            
            # Create a mock database object for export functions
            class MockDB:
                def get_ioc(self, ioc_id):
                    return next((ioc for ioc in iocs if ioc['id'] == ioc_id), None)
                
                def get_source(self, source_id):
                    return None
                
                def get_iocs_by_source(self, source_id):
                    return []
            
            mock_db = MockDB()
            
            # Use export functions but with direct IOC list
            if format_type == 'txt':
                output_file = _quick_export_txt(iocs, OUTPUT_FOLDER, timestamp)
            elif format_type == 'txt-simple':
                output_file = _quick_export_txt_simple(iocs, OUTPUT_FOLDER, timestamp)
            elif format_type == 'csv':
                output_file = _quick_export_csv(iocs, OUTPUT_FOLDER, timestamp)
            elif format_type == 'csv-firewall':
                output_file = _quick_export_csv_firewall(iocs, OUTPUT_FOLDER, timestamp)
            elif format_type == 'json':
                output_file = _quick_export_json(iocs, OUTPUT_FOLDER, timestamp)
            elif format_type == 'json-simple':
                output_file = _quick_export_json_simple(iocs, OUTPUT_FOLDER, timestamp)
            elif format_type == 'xlsx':
                output_file = _quick_export_xlsx(iocs, OUTPUT_FOLDER, timestamp)
            else:
                return api_error('Invalid export format', 400)
            
            # Determine file extension
            ext_map = {
                'txt': 'txt',
                'txt-simple': 'txt',
                'csv': 'csv',
                'csv-firewall': 'csv',
                'json': 'json',
                'json-simple': 'json',
                'xlsx': 'xlsx'
            }
            ext = ext_map.get(format_type, 'txt')
            
            return send_file(str(output_file), as_attachment=True, download_name=f"quick_export_{timestamp}.{ext}")
            
        finally:
            # Clean up temporary file if created
            if temp_file_path and temp_file_path.exists():
                try:
                    temp_file_path.unlink()
                except Exception as e:
                    logger.warning(f"Failed to delete temp file {temp_file_path}: {e}")
    
    except RequestEntityTooLarge:
        return api_error('File too large', 413)
    except Exception as e:
        logger.error(f"api_quick_export error: {e}")
        return api_error(str(e), 500)

# Helper functions for quick export (without database)
def _quick_export_txt(iocs: List[Dict], output_folder: Path, timestamp: str) -> Path:
    """Quick export to TXT format"""
    output_file = output_folder / "iocs" / f"quick_export_{timestamp}.txt"
    output_file.parent.mkdir(parents=True, exist_ok=True)
    with open(output_file, 'w', encoding='utf-8') as f:
        for ioc in iocs:
            f.write(f"{ioc['ioc_type']}\t{ioc['ioc_value']}\n")
    return output_file

def _quick_export_txt_simple(iocs: List[Dict], output_folder: Path, timestamp: str) -> Path:
    """Quick export to TXT simple format"""
    output_file = output_folder / "iocs" / f"quick_export_simple_{timestamp}.txt"
    output_file.parent.mkdir(parents=True, exist_ok=True)
    with open(output_file, 'w', encoding='utf-8') as f:
        for ioc in iocs:
            f.write(f"{ioc['ioc_value']}\n")
    return output_file

def _quick_export_csv(iocs: List[Dict], output_folder: Path, timestamp: str) -> Path:
    """Quick export to CSV format"""
    output_file = output_folder / "iocs" / f"quick_export_{timestamp}.csv"
    output_file.parent.mkdir(parents=True, exist_ok=True)
    with open(output_file, 'w', encoding='utf-8', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(['Type', 'Value', 'Source', 'Created At', 'Tags'])
        for ioc in iocs:
            writer.writerow([
                ioc.get('ioc_type', ''),
                ioc.get('ioc_value', ''),
                'Quick Export',
                ioc.get('created_at', ''),
                ''
            ])
    return output_file

def _quick_export_csv_firewall(iocs: List[Dict], output_folder: Path, timestamp: str) -> Path:
    """Quick export to CSV firewall format"""
    output_file = output_folder / "iocs" / f"quick_export_firewall_{timestamp}.csv"
    output_file.parent.mkdir(parents=True, exist_ok=True)
    with open(output_file, 'w', encoding='utf-8', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(['value'])
        for ioc in iocs:
            writer.writerow([ioc.get('ioc_value', '')])
    return output_file

def _quick_export_json(iocs: List[Dict], output_folder: Path, timestamp: str) -> Path:
    """Quick export to JSON format"""
    output_file = output_folder / "iocs" / f"quick_export_{timestamp}.json"
    output_file.parent.mkdir(parents=True, exist_ok=True)
    export_data = {
        'export_date': datetime.now().isoformat(),
        'total_iocs': len(iocs),
        'source': 'Quick Export',
        'iocs': iocs
    }
    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(export_data, f, indent=2, ensure_ascii=False, default=str)
    return output_file

def _quick_export_json_simple(iocs: List[Dict], output_folder: Path, timestamp: str) -> Path:
    """Quick export to JSON simple format"""
    output_file = output_folder / "iocs" / f"quick_export_simple_{timestamp}.json"
    output_file.parent.mkdir(parents=True, exist_ok=True)
    
    # Group by type
    grouped = {}
    for ioc in iocs:
        ioc_type = ioc.get('ioc_type', 'unknown')
        if ioc_type not in grouped:
            grouped[ioc_type] = []
        grouped[ioc_type].append(ioc.get('ioc_value', ''))
    
    export_data = {
        'export_date': datetime.now().isoformat(),
        'total_iocs': len(iocs),
        'source': 'Quick Export',
        'iocs_by_type': grouped
    }
    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(export_data, f, indent=2, ensure_ascii=False, default=str)
    return output_file

def _quick_export_xlsx(iocs: List[Dict], output_folder: Path, timestamp: str) -> Path:
    """Quick export to XLSX format"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise RuntimeError("openpyxl is not available. Please install it: pip install openpyxl")
    
    output_file = output_folder / "iocs" / f"quick_export_{timestamp}.xlsx"
    output_file.parent.mkdir(parents=True, exist_ok=True)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "IOCs"
    
    # Header
    header_fill = PatternFill(start_color="F5F5F7", end_color="F5F5F7", fill_type="solid")
    header_font = Font(name='Calibri', size=11, bold=True, color="1D1D1F")
    
    headers = ['Type', 'Value', 'Source', 'Created At']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='left', vertical='center')
    
    # Data
    for row_idx, ioc in enumerate(iocs, 2):
        ws.cell(row=row_idx, column=1, value=ioc.get('ioc_type', ''))
        ws.cell(row=row_idx, column=2, value=ioc.get('ioc_value', ''))
        ws.cell(row=row_idx, column=3, value='Quick Export')
        ws.cell(row=row_idx, column=4, value=ioc.get('created_at', ''))
    
    # Auto-adjust column widths
    for col in range(1, len(headers) + 1):
        col_letter = get_column_letter(col)
        ws.column_dimensions[col_letter].width = 20
    
    wb.save(str(output_file))
    return output_file

# ========== ROUTES API STIX IMPORT > EXPORT ==========


@app.route('/api/stix/files', methods=['GET'])
@require_auth
def api_stix_files():
    """API: List all STIX files from stix and stix_conversions directories"""
    try:
        files = []
        
        # List files from main stix directory
        stix_dir = OUTPUT_FOLDER / "stix"
        if stix_dir.exists():
            for file_path in stix_dir.glob("*.json"):
                try:
                    stat = file_path.stat()
                    files.append({
                        'name': file_path.name,
                        'path': file_path.name,  # Relative path for download
                        'source': 'export',
                        'size': stat.st_size,
                        'size_formatted': format_bytes(stat.st_size),
                        'converted_at': datetime.fromtimestamp(stat.st_mtime).isoformat()
                    })
                except Exception as e:
                    logger.warning(f"Error reading file {file_path}: {e}")
                    continue
        
        # List files from stix_conversions directory
        stix_conv_dir = OUTPUT_FOLDER / "stix_conversions"
        if stix_conv_dir.exists():
            for file_path in stix_conv_dir.glob("*.json"):
                try:
                    stat = file_path.stat()
                    files.append({
                        'name': file_path.name,
                        'path': f"stix_conversions/{file_path.name}",  # Relative path for download
                        'source': 'import',
                        'size': stat.st_size,
                        'size_formatted': format_bytes(stat.st_size),
                        'converted_at': datetime.fromtimestamp(stat.st_mtime).isoformat()
                    })
                except Exception as e:
                    logger.warning(f"Error reading file {file_path}: {e}")
                    continue
        
        # Sort by modification time, newest first
        files.sort(key=lambda x: x['converted_at'], reverse=True)
        
        return api_success({'files': files})
    
    except Exception as e:
        logger.error(f"api_stix_files error: {e}")
        return api_error(str(e), 500)

@app.route('/api/stix/files/<path:filepath>/download', methods=['GET'])
@require_auth
def api_stix_files_download(filepath):
    """API: Download a STIX file from stix or stix_conversions directory"""
    try:
        # Check if path contains stix_conversions prefix
        if filepath.startswith('stix_conversions/'):
            base_dir = OUTPUT_FOLDER / "stix_conversions"
            relative_path = filepath[len('stix_conversions/'):]
        else:
            base_dir = OUTPUT_FOLDER / "stix"
            relative_path = filepath
        
        file_path = base_dir / relative_path
        
        # Prevent directory traversal - ensure file is within allowed directories
        outputs_dir = OUTPUT_FOLDER.resolve()
        if not str(file_path.resolve()).startswith(str(outputs_dir)):
            return api_error('Invalid file path', 400)
        
        if file_path.exists() and file_path.is_file():
            # Return file content as JSON (for graph loading) or as download
            if request.args.get('raw', 'false').lower() == 'true':
                return send_file(str(file_path), as_attachment=True, download_name=file_path.name)
            else:
                # Return JSON content directly for loading in graph
                return send_file(str(file_path), mimetype='application/json')
        else:
            return api_not_found('File')
    
    except Exception as e:
        logger.error(f"api_stix_files_download error: {e}")
        return api_error(str(e), 500)

@app.route('/api/stix/files/<path:filepath>/delete', methods=['DELETE'])
@require_auth
def api_stix_files_delete(filepath):
    """API: Delete a STIX file from stix or stix_conversions directory"""
    try:
        # Check if path contains stix_conversions prefix
        if filepath.startswith('stix_conversions/'):
            base_dir = OUTPUT_FOLDER / "stix_conversions"
            relative_path = filepath[len('stix_conversions/'):]
        else:
            base_dir = OUTPUT_FOLDER / "stix"
            relative_path = filepath
        
        file_path = base_dir / relative_path
        
        # Prevent directory traversal - ensure file is within allowed directories
        outputs_dir = OUTPUT_FOLDER.resolve()
        if not str(file_path.resolve()).startswith(str(outputs_dir)):
            return api_error('Invalid file path', 400)
        
        if file_path.exists() and file_path.is_file():
            file_path.unlink()
            return api_success(message='File deleted successfully')
        else:
            return api_not_found('File')
    
    except Exception as e:
        logger.error(f"api_stix_files_delete error: {e}")
        return api_error(str(e), 500)

@app.route('/api/stix/files/upload', methods=['POST'])
@require_auth
def api_stix_files_upload():
    """Upload STIX, create a source, and import its supported observables."""
    try:
        if 'file' not in request.files:
            return api_error('No file provided', 400)
        
        file = request.files['file']
        if file.filename == '':
            return api_error('No file selected', 400)
        
        if not file.filename.endswith('.json'):
            return api_error('Only JSON files are allowed', 400)
        
        raw_content = file.read()
        try:
            stix_text = raw_content.decode('utf-8-sig')
            from modules.stix_ioc_importer import extract_stix_iocs, source_metadata
            bundle, extracted_iocs = extract_stix_iocs(stix_text)
        except (UnicodeDecodeError, json.JSONDecodeError, ValueError) as exc:
            return api_error(f'Invalid STIX JSON: {exc}', 400)

        content_hash = hashlib.sha256(raw_content).hexdigest()
        hash_marker = f'STIX SHA256: {content_hash}'
        with db.connection() as conn:
            existing = conn.execute(
                """SELECT id, original_filename
                   FROM sources
                   WHERE source_type = 'stix' AND is_deleted = 0 AND context LIKE ?
                   LIMIT 1""",
                (f'%{hash_marker}%',),
            ).fetchone()
            if existing:
                count_row = conn.execute(
                    'SELECT COUNT(*) AS count FROM iocs WHERE source_id = ? AND is_deleted = 0',
                    (existing['id'],),
                ).fetchone()
                return api_success({
                    'filename': existing['original_filename'],
                    'source_id': existing['id'],
                    'ioc_count': count_row['count'] if count_row else 0,
                    'already_imported': True,
                }, 'This STIX content is already imported')

        # Save to stix_conversions directory
        stix_conv_dir = OUTPUT_FOLDER / "stix_conversions"
        stix_conv_dir.mkdir(parents=True, exist_ok=True)
        
        # Use original filename or generate one
        filename = file.filename
        file_path = stix_conv_dir / filename
        
        # If file exists, add timestamp
        if file_path.exists():
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            name_parts = filename.rsplit('.', 1)
            filename = f"{name_parts[0]}_{timestamp}.{name_parts[1]}"
            file_path = stix_conv_dir / filename
        
        file.stream.seek(0)
        file.save(str(file_path))

        metadata = source_metadata(bundle, filename)
        metadata['context'] += f'\n{hash_marker}'
        source_id = None
        imported_count = 0
        try:
            source_id = db.create_source(
                name=metadata['name'],
                context=metadata['context'],
                source_type='stix',
                file_path=str(file_path),
                original_filename=filename,
            )
            source_info = db.get_source(source_id)
            for item in extracted_iocs:
                ioc_id = db.create_ioc(
                    source_id=source_id,
                    ioc_type=item['ioc_type'],
                    ioc_value=item['value'],
                    raw_value=item['value'],
                    source_info=source_info,
                    tlp=item['tlp'],
                    validation_status='Unverified',
                    confidence=item['confidence'],
                )
                if item.get('notes'):
                    db.update_ioc_notes(ioc_id, item['notes'])
                if item.get('first_seen') or item.get('last_seen'):
                    with db.connection() as conn:
                        conn.execute(
                            """UPDATE iocs
                               SET first_seen = COALESCE(NULLIF(?, ''), first_seen),
                                   last_seen = COALESCE(NULLIF(?, ''), last_seen)
                               WHERE id = ?""",
                            (item.get('first_seen', ''), item.get('last_seen', ''), ioc_id),
                        )
                        conn.commit()
                imported_count += 1
            db.update_source_status(source_id, 'completed')
        except Exception:
            if source_id:
                db.update_source_status(source_id, 'error')
            raise
        
        return api_success({
            'filename': filename,
            'path': f"stix_conversions/{filename}",
            'source_id': source_id,
            'ioc_count': imported_count,
        }, f'File uploaded and {imported_count} IOC(s) imported')
    
    except Exception as e:
        logger.error(f"api_stix_files_upload error: {e}")
        return api_error(str(e), 500)

# ========== ROUTES API SAVED STIX MODELS ==========

@app.route('/api/stix/models', methods=['GET'])
@require_auth
def api_stix_models_list():
    """API: List all saved STIX models"""
    try:
        models = db.get_all_saved_stix_models()
        return api_success({'models': models})
    except Exception as e:
        logger.error(f"api_stix_models_list error: {e}")
        return api_error(str(e), 500)

@app.route('/api/stix/models', methods=['POST'])
@require_auth
def api_stix_models_create():
    """API: Create a new saved STIX model"""
    try:
        data = request.get_json()
        name = data.get('name', '').strip()
        stix_content = data.get('stix_content', '')
        description = data.get('description', '').strip()
        node_count = data.get('node_count', 0)
        edge_count = data.get('edge_count', 0)
        local_metadata = data.get('local_metadata')
        
        if not name:
            return api_error('Name is required', 400)
        if not stix_content:
            return api_error('STIX content is required', 400)
        
        model_id = db.create_saved_stix_model(
            name=name,
            stix_content=stix_content,
            description=description,
            node_count=node_count,
            edge_count=edge_count,
            local_metadata=json.dumps(local_metadata) if isinstance(local_metadata, dict) else local_metadata,
        )
        
        return api_success({'id': model_id}, 'Model saved successfully')
    except Exception as e:
        logger.error(f"api_stix_models_create error: {e}")
        return api_error(str(e), 500)


@app.route('/api/stix/models/from-iocs', methods=['POST'])
@require_auth
def api_stix_model_from_iocs():
    """Create one STIX 2.1 model from selected IOC rows."""
    try:
        data = request.get_json(silent=True) or {}
        raw_ids = data.get('ioc_ids') or []
        model_name = str(data.get('name') or '').strip()
        if not model_name:
            return api_error('Model name is required', 400)
        try:
            ioc_ids = list(dict.fromkeys(int(value) for value in raw_ids))
        except (TypeError, ValueError):
            return api_error('Invalid IOC selection', 400)
        if not ioc_ids:
            return api_error('Select at least one IOC', 400)

        selected_iocs = [ioc for ioc_id in ioc_ids if (ioc := db.get_ioc(ioc_id)) and not ioc.get('is_deleted')]
        for ioc in selected_iocs:
            source = db.get_source(ioc.get('source_id')) if ioc.get('source_id') else None
            ioc['source_name'] = source.get('name') if source else ''
            with db.connection() as conn:
                rows = conn.execute(
                    """SELECT g.name FROM groups g
                       JOIN ioc_groups ig ON ig.group_id = g.id WHERE ig.ioc_id = ?""",
                    (ioc['id'],),
                ).fetchall()
                ioc['groups'] = [row['name'] for row in rows]
        from modules.ioc_stix_builder import bundle_json
        producer_name = db.get_setting('stix_producer_name', 'ODYSAFE CTI') or 'ODYSAFE CTI'
        stix_content, skipped, local_metadata = bundle_json(
            selected_iocs, include_extensions=True, producer_name=producer_name,
        )
        parsed = json.loads(stix_content)
        if not parsed['objects']:
            return api_error('None of the selected IOC types can be converted to standard STIX', 400)

        model_id = db.create_saved_stix_model(
            name=model_name,
            description=f'Created from {len(selected_iocs)} selected ODYSAFE IOC(s).',
            stix_content=stix_content,
            node_count=len(parsed['objects']),
            edge_count=0,
            local_metadata=json.dumps(local_metadata),
        )
        return api_success({
            'model_id': model_id,
            'created_count': len(parsed['objects']),
            'skipped': skipped,
        }, 'STIX model created')
    except Exception as exc:
        logger.error('api_stix_model_from_iocs error: %s', exc, exc_info=True)
        return api_error(str(exc), 500)


@app.route('/api/stix/objects/from-iocs', methods=['POST'])
@require_auth
def api_stix_objects_from_iocs():
    """Convert database or manually entered IOCs for an open STIX 2.1 model."""
    try:
        data = request.get_json(silent=True) or {}
        raw_ids = data.get('ioc_ids') or []
        try:
            ioc_ids = list(dict.fromkeys(int(value) for value in raw_ids))
        except (TypeError, ValueError):
            return api_error('Invalid IOC selection', 400)
        selected_iocs = [ioc for ioc_id in ioc_ids if (ioc := db.get_ioc(ioc_id)) and not ioc.get('is_deleted')]
        for ioc in selected_iocs:
            source = db.get_source(ioc.get('source_id')) if ioc.get('source_id') else None
            ioc['source_name'] = source.get('name') if source else ''
        manual = data.get('manual')
        if isinstance(manual, dict) and manual.get('ioc_type') and manual.get('ioc_value'):
            selected_iocs.append({
                'ioc_type': str(manual['ioc_type']).strip(),
                'ioc_value': str(manual['ioc_value']).strip(),
                'tlp': str(manual.get('tlp') or 'WHITE'),
                'validation_status': 'Unverified',
                'confidence': 'Unknown',
                'source_name': 'STIX Builder manual entry',
            })
        if not selected_iocs:
            return api_error('Select or enter at least one IOC', 400)
        from modules.ioc_stix_builder import build_bundle
        bundle, skipped, metadata = build_bundle(
            selected_iocs,
            include_extensions=True,
            producer_name=db.get_setting('stix_producer_name', 'ODYSAFE CTI') or 'ODYSAFE CTI',
        )
        return api_success({'objects': bundle['objects'], 'skipped': skipped, 'local_metadata': metadata})
    except Exception as exc:
        logger.error('api_stix_objects_from_iocs error: %s', exc, exc_info=True)
        return api_error(str(exc), 500)


@app.route('/api/stix/models/<int:model_id>', methods=['PUT'])
@require_auth
def api_stix_models_update(model_id):
    """Update an existing STIX model without changing object IDs."""
    try:
        if not db.get_saved_stix_model(model_id):
            return api_not_found('Model')
        data = request.get_json(silent=True) or {}
        content = data.get('stix_content')
        if content:
            from modules.stix_ioc_importer import parse_bundle
            parse_bundle(content)
        db.update_saved_stix_model(
            model_id,
            name=data.get('name'), description=data.get('description'), stix_content=content,
            node_count=data.get('node_count'), edge_count=data.get('edge_count'),
            local_metadata=json.dumps(data['local_metadata']) if isinstance(data.get('local_metadata'), dict) else data.get('local_metadata'),
        )
        return api_success({'id': model_id}, 'Model updated')
    except (ValueError, json.JSONDecodeError) as exc:
        return api_error(f'Invalid STIX content: {exc}', 400)
    except Exception as exc:
        logger.error('api_stix_models_update error: %s', exc, exc_info=True)
        return api_error(str(exc), 500)

@app.route('/api/stix/models/<int:model_id>', methods=['GET'])
@require_auth
def api_stix_models_get(model_id):
    """API: Get a saved STIX model by ID"""
    try:
        model = db.get_saved_stix_model(model_id)
        if not model:
            return api_not_found('Model')
        
        # Update last_loaded_at
        db.update_saved_stix_model_last_loaded(model_id)
        
        return api_success({'model': model})
    except Exception as e:
        logger.error(f"api_stix_models_get error: {e}")
        return api_error(str(e), 500)

@app.route('/api/stix/models/<int:model_id>', methods=['DELETE'])
@require_auth
def api_stix_models_delete(model_id):
    """API: Delete a saved STIX model"""
    try:
        success = db.delete_saved_stix_model(model_id)
        if not success:
            return api_not_found('Model')
        
        return api_success(message='Model deleted successfully')
    except Exception as e:
        logger.error(f"api_stix_models_delete error: {e}")
        return api_error(str(e), 500)

@app.route('/api/tags', methods=['GET'])
@require_auth
def api_tags():
    """API: Retrieve all tags"""
    try:
        tags = db.get_all_tags()
        return api_success({'tags': tags})
    except Exception as e:
        logger.error(f"api_tags error: {e}")
        return api_error(str(e), 500)

@app.route('/api/stats', methods=['GET'])
@require_auth
def api_stats():
    """API: Retrieve global statistics"""
    try:
        stats = db.get_statistics()
        recent_limit = int(db.get_setting('recent_sources_limit', '20'))
        recent_sources = db.get_all_sources(limit=recent_limit)
        return api_success({
            'stats': stats,
            'recent_sources': recent_sources
        })
    except Exception as e:
        logger.error(f"api_stats error: {e}")
        return api_error(str(e), 500)

# ========== ROUTES API RAPPORTS ==========

@app.route('/api/reports/<int:source_id>', methods=['GET'])
@require_auth
def api_reports_list(source_id):
    """API: List of generated reports for a source"""
    try:
        reports = db.get_reports_by_source(source_id)
        return api_success({'reports': reports})
    except Exception as e:
        logger.error(f"api_reports_list error: {e}")
        return api_error(str(e), 500)

@app.route('/api/reports/<int:report_id>/download', methods=['GET'])
@require_auth
def api_report_download(report_id):
    """API: Download a generated report"""
    try:
        report = db.get_report(report_id)
        if not report:
            return api_not_found('Report')
        
        file_path = Path(report['file_path'])
        if not file_path.exists():
            return api_not_found('Report file')
        
        # Determine download name
        report_type = report['report_type']
        extension = file_path.suffix or '.json'
        download_name = f"{report_type}_{file_path.stem}{extension}"
        
        return send_file(
            str(file_path),
            as_attachment=True,
            download_name=download_name
        )
    except Exception as e:
        logger.error(f"api_report_download error: {e}")
        return api_error(str(e), 500)

@app.route('/api/iocs/export', methods=['GET'])
@require_auth
def api_iocs_export():
    """API: Get IOC IDs with applied filters"""
    try:
        # Get filtering parameters (same logic as iocs_list)
        filters = {}
        if request.args.get('type'):
            filters['ioc_type'] = request.args.get('type')
        if request.args.get('search'):
            filters['search'] = request.args.get('search')
        if request.args.getlist('tag'):
            filters['tags'] = request.args.getlist('tag')
            filters['tag_logic'] = request.args.get('tag_logic', 'AND')
        if request.args.get('date_from'):
            filters['date_from'] = request.args.get('date_from')
        if request.args.get('date_to'):
            filters['date_to'] = request.args.get('date_to')
        
        # Get all IOCs with these filters using streaming for large datasets
        iocs = []
        for batch in db.get_all_iocs_streaming(filters=filters, limit=None):
            iocs.extend(batch)
        
        return api_success({
            'ioc_ids': [ioc['id'] for ioc in iocs],
            'count': len(iocs)
        })
    except Exception as e:
        logger.error(f"api_iocs_export error: {e}")
        return api_error(str(e), 500)


@app.route('/api/progress/<task_id>', methods=['GET'])
@require_auth
def api_get_progress(task_id):
    """API: Get task progress"""
    try:
        progress = progress_tracker.get_progress(task_id)
        if progress:
            progress_data = dict(progress)
            if (
                task_id.startswith('source_')
                and progress_data.get('status') == 'completed'
                and 'memory_index_status' not in progress_data
            ):
                source_id = task_id[len('source_'):]
                memory_status_available = True
                try:
                    cross_ref = db.get_cross_ref(f'odysafe:source:{source_id}')
                    cross_ref_status = (cross_ref or {}).get('status')
                except Exception as memory_status_error:
                    memory_status_available = False
                    logger.warning(
                        'Unable to read CTI Memory status for completed source %s: %s',
                        source_id,
                        memory_status_error,
                    )
                    cross_ref_status = None
                if cross_ref_status == 'active':
                    memory_index_status = 'indexed'
                elif cross_ref_status == 'failed':
                    memory_index_status = 'failed'
                elif memory_status_available:
                    memory_index_status = 'never_indexed'
                else:
                    memory_index_status = 'unknown'
                progress_data['memory_indexed'] = memory_index_status == 'indexed'
                progress_data['memory_index_status'] = memory_index_status
            return api_success({'progress': progress_data})
        else:
            return api_not_found('Task')
    except Exception as e:
        logger.error(f"api_get_progress error: {e}")
        return api_error(str(e), 500)

@app.route('/api/progress/<task_id>/stop', methods=['POST'])
@require_auth
def api_stop_task(task_id):
    """API: Stop a running task"""
    try:
        success = progress_tracker.stop_task(task_id)
        if success:
            return api_success({'message': 'Stop request sent'})
        else:
            return api_error('Task not found or cannot be stopped', 404)
    except Exception as e:
        logger.error(f"api_stop_task error: {e}")
        return api_error(str(e), 500)

@app.route('/api/progress/<task_id>/download', methods=['GET'])
@require_auth
def api_download_complete_report(task_id):
    """API: Download STIX report once generation is complete"""
    try:
        progress = progress_tracker.get_progress(task_id)
        if not progress:
            return api_not_found('Task')
        
        if progress.get('status') != 'completed':
            return api_error(
                'Report not yet completed',
                202,
                {
                'status': progress.get('status'),
                'progress': progress.get('percentage', 0)
                }
            )
        
        results = progress.get('results', {})
        if not results:
            return api_error('No results available', 404)
        
        # Use generated STIX file
        stix_file = results.get('stix_file')
        if not stix_file:
            return api_error('STIX file not found. The report may not have been fully generated.', 404)
        
        stix_path = Path(stix_file)
        if not stix_path.exists():
            return api_error('STIX file not found on filesystem', 404)
        
        # Send STIX file
        return send_file(
            str(stix_path),
            as_attachment=True,
            download_name=stix_path.name,
            mimetype='application/json'
        )
        
    except Exception as e:
        logger.error(f"api_download_complete_report error: {e}", exc_info=True)
        return api_error(str(e), 500)

# ========== ROUTES API SUPPRESSION IOCs ==========

@app.route('/api/iocs/bulk-delete', methods=['POST'])
@require_auth
def api_iocs_bulk_delete():
    """API: Delete multiple IOCs in bulk"""
    try:
        data = request.get_json()
        ioc_ids = data.get('ioc_ids', [])
        
        if not ioc_ids:
            return api_error('No IOCs selected', 400)
        
        count = 0
        for ioc_id in ioc_ids:
            if db.soft_delete_ioc(ioc_id):
                count += 1
        
        return api_success(
            {'count': count},
            f'{count} IOC(s) deleted successfully'
        )
    except Exception as e:
        logger.error(f"api_iocs_bulk_delete error: {e}")
        return api_error(str(e), 500)

@app.route('/api/iocs/all-ids', methods=['GET'])
@require_auth
def api_iocs_all_ids():
    """API: Retrieve all IOC IDs with current filters"""
    try:
        # Retrieve the same filters as the page
        ioc_type = request.args.get('type', '').strip()
        search = request.args.get('search', '').strip()
        source_name = request.args.get('source', '').strip()
        group_id = request.args.get('group', '').strip()
        date_from = request.args.get('date_from', '').strip()
        date_to = request.args.get('date_to', '').strip()
        show_duplicates = request.args.get('duplicates', '').strip()
        
        filters = {}
        if ioc_type:
            filters['ioc_type'] = ioc_type
        if search:
            filters['search'] = search
        if source_name:
            filters['source_name'] = source_name
        if group_id:
            try:
                filters['group_id'] = int(group_id)
            except ValueError:
                pass
        if date_from:
            filters['date_from'] = date_from
        if date_to:
            filters['date_to'] = date_to
        if show_duplicates == 'true':
            filters['show_duplicates'] = True
        
        # Retrieve all IOCs with these filters using streaming
        iocs = []
        for batch in db.get_all_iocs_streaming(filters=filters, limit=None):
            iocs.extend(batch)
        ioc_ids = [ioc['id'] for ioc in iocs]
        
        return api_success({'ioc_ids': ioc_ids})
    except Exception as e:
        logger.error(f"api_iocs_all_ids error: {e}")
        return api_error(str(e), 500)

@app.route('/api/iocs/load', methods=['GET'])
@require_auth
def api_iocs_load():
    """API: Load IOCs with pagination for lazy loading"""
    try:
        from modules.ioc_query_urls import get_query_urls
        
        # Retrieve filtering parameters (same as iocs_list)
        ioc_type = request.args.get('type', '').strip()
        search = request.args.get('search', '').strip()
        source_name = request.args.get('source', '').strip()
        date_range = request.args.get('date_range', '')
        date_from = request.args.get('date_from', '').strip()
        date_to = request.args.get('date_to', '').strip()
        page = int(request.args.get('page', 1))
        per_page = int(request.args.get('per_page', 30))
        
        # Handle hashtag filters
        hashtag_filters = request.args.get('hashtag_filters', '')
        if hashtag_filters:
            try:
                hashtag_filters_list = json.loads(hashtag_filters)
                # Convert hashtag filters to regular filters
                for filter_key in hashtag_filters_list:
                    if ':' in filter_key:
                        filter_type, filter_value = filter_key.split(':', 1)
                        if filter_type == 'type':
                            ioc_type = filter_value
                        elif filter_type == 'source':
                            source_name = filter_value
                        elif filter_type == 'group':
                            # Extract group ID from group name
                            groups = db.get_all_groups()
                            for group in groups:
                                group_name_clean = group['name'].replace(' ', '_').replace('/', '_').replace('\\', '_').replace('#', '').replace(':', '_')
                                if group_name_clean == filter_value:
                                    group_id = group['id']
                                    break
            except json.JSONDecodeError:
                pass
        
        filters = {}
        if ioc_type:
            filters['ioc_type'] = ioc_type
        if search:
            filters['search'] = search
        if source_name:
            filters['source_name'] = source_name
        group_id = request.args.get('group', '').strip()
        if group_id:
            try:
                filters['group_id'] = int(group_id)
            except ValueError:
                pass
        show_duplicates = request.args.get('duplicates', '').strip()
        if show_duplicates == 'true':
            filters['show_duplicates'] = True
        if date_range:
            filters['date_range'] = date_range
            from datetime import datetime, timedelta
            now = datetime.now()
            if date_range == '24h':
                filters['date_from'] = (now - timedelta(hours=24)).strftime('%Y-%m-%d')
            elif date_range == '7d':
                filters['date_from'] = (now - timedelta(days=7)).strftime('%Y-%m-%d')
            elif date_range == '30d':
                filters['date_from'] = (now - timedelta(days=30)).strftime('%Y-%m-%d')
            elif date_range == '3m':
                filters['date_from'] = (now - timedelta(days=90)).strftime('%Y-%m-%d')
            elif date_range == '1y':
                filters['date_from'] = (now - timedelta(days=365)).strftime('%Y-%m-%d')
            elif date_range == 'custom':
                if date_from:
                    filters['date_from'] = date_from
                if date_to:
                    filters['date_to'] = date_to
        else:
            if date_from:
                filters['date_from'] = date_from
            if date_to:
                filters['date_to'] = date_to
        
        offset = (page - 1) * per_page
        iocs, total = db.get_all_iocs(filters=filters, limit=per_page, offset=offset)
        
        # Enrich each IOC with its query URLs
        for ioc in iocs:
            ioc['query_urls'] = get_query_urls(ioc['ioc_type'], ioc['ioc_value'])
        
        return api_success({
            'iocs': iocs,
            'total': total,
            'page': page,
            'per_page': per_page,
            'has_more': (page * per_page) < total
        })
    except Exception as e:
        logger.error(f"api_iocs_load error: {e}")
        return api_error(str(e), 500)

@app.route('/api/iocs/get-sources', methods=['POST'])
@require_auth
def api_iocs_get_sources():
    """API: Retrieve source_ids from ioc_ids"""
    try:
        data = request.get_json()
        ioc_ids = data.get('ioc_ids', [])
        
        if not ioc_ids:
            return api_error('No IOCs provided', 400)
        
        source_ids = []
        for ioc_id in ioc_ids:
            ioc = db.get_ioc(ioc_id)
            if ioc and ioc.get('source_id'):
                source_ids.append(ioc['source_id'])
        
        # Remove duplicates
        source_ids = list(set(source_ids))
        
        return api_success({'source_ids': source_ids})
    except Exception as e:
        logger.error(f"api_iocs_get_sources error: {e}")
        return api_error(str(e), 500)

@app.route('/api/iocs/bulk-add-group', methods=['POST'])
@require_auth
def api_iocs_bulk_add_group():
    """API: Add multiple IOCs to a group"""
    try:
        data = request.get_json()
        ioc_ids = data.get('ioc_ids', [])
        group_id = data.get('group_id')
        
        if not ioc_ids:
            return api_error('No IOCs selected', 400)
        if not group_id:
            return api_error('Group ID is required', 400)
        
        count = db.bulk_add_iocs_to_group(ioc_ids, group_id)
        
        return api_success(
            {'count': count},
            f'{count} IOC(s) added to group'
        )
    except Exception as e:
        logger.error(f"api_iocs_bulk_add_group error: {e}")
        return api_error(str(e), 500)

@app.route('/api/iocs/remove-group', methods=['POST'])
@require_auth
def api_iocs_remove_group():
    """API: Remove an IOC from a group (direct group or source group exclusion)"""
    try:
        data = request.get_json()
        ioc_id = data.get('ioc_id')
        group_id = data.get('group_id')
        is_source_group = data.get('is_source_group', False)  # Indicates if it's a source group
        
        if not ioc_id:
            return api_error('IOC ID is required', 400)
        if not group_id:
            return api_error('Group ID is required', 400)
        
        if is_source_group:
            # Exclude the source group for this IOC
            success = db.exclude_ioc_from_source_group(ioc_id, group_id)
            if success:
                return api_success(message='IOC excluded from source group successfully')
            else:
                return api_error('IOC was already excluded from this source group', 400)
        else:
            # Remove the direct group
            success = db.remove_ioc_from_group(ioc_id, group_id)
            if success:
                return api_success(message='IOC removed from group successfully')
            else:
                return api_error('IOC was not in this group', 400)
    except Exception as e:
        logger.error(f"api_iocs_remove_group error: {e}")
        return api_error(str(e), 500)

@app.route('/api/groups/create', methods=['POST'])
@require_auth
def api_groups_create():
    """API: Create a new group"""
    try:
        data = request.get_json()
        name = data.get('name', '').strip()
        color = data.get('color', '#8B5CF6')
        
        if not name:
            return api_error('Group name is required', 400)
        
        group_id = db.create_group(name, color=color)
        return api_success(
            {'group_id': group_id},
            'Group created successfully'
        )
    except Exception as e:
        logger.error(f"api_groups_create error: {e}")
        return api_error(str(e), 500)

@app.route('/api/sources/active', methods=['GET'])
@require_auth
def api_sources_active():
    """API: Get active sources list"""
    try:
        sources = db.get_all_sources(limit=1000)
        return api_success({
            'sources': [{'id': s['id'], 'name': s['name'], 'source_type': s['source_type'], 'created_at': s['created_at']} for s in sources]
        })
    except Exception as e:
        logger.error(f"api_sources_active error: {e}")
        return api_error(str(e), 500)

@app.route('/api/sources/<int:source_id>/iocs', methods=['GET'])
@require_auth
def api_source_iocs(source_id):
    """API: Get all IOCs for a source"""
    try:
        iocs = db.get_iocs_by_source(source_id)
        formatted_iocs = []
        for ioc in iocs:
            ioc_id = ioc.get('id')
            ttp_links = db.get_ioc_ttp_links(ioc_id) if ioc_id else []
            formatted_iocs.append({
                'id': ioc.get('id'),
                'ioc_type': ioc.get('ioc_type'),
                'ioc_value': ioc.get('ioc_value'),
                'raw_value': ioc.get('raw_value'),
                'first_seen': ioc.get('first_seen'),
                'last_seen': ioc.get('last_seen'),
                'tlp': ioc.get('tlp'),
                'validation_status': ioc.get('validation_status'),
                'confidence': ioc.get('confidence'),
                'notes': ioc.get('notes'),
                'is_whitelisted': ioc.get('is_whitelisted'),
                'tags': ioc.get('tags', []),
                'mitre_techniques': [l['technique_id'] for l in ttp_links],
                'mitre_links': ttp_links,
            })
        return api_success({'iocs': formatted_iocs})
    except Exception as e:
        logger.error(f"api_source_iocs error: {e}")
        return api_error(str(e), 500)

@app.route('/api/mitre/groups', methods=['GET'])
@require_auth
def api_mitre_groups():
    """API: Get MITRE ATT&CK groups list"""
    try:
        search = request.args.get('search', '').strip()
        groups = db.get_mitre_groups(search=search if search else None)
        return api_success({'groups': groups})
    except Exception as e:
        logger.error(f"api_mitre_groups error: {e}")
        return api_error(str(e), 500)

@app.route('/api/cti-resources/urls', methods=['GET'])
@require_auth
def api_cti_resources_urls():
    """API: Get a list of URLs from CTI resources"""
    try:
        urls = []
        try:
            if github_repo_manager.repo_exists():
                categories = github_repo_manager.fetch_all_categories()
                for cat_name, cat_data in categories.items():
                    for src in cat_data.get('sources', []):
                        if src.get('url'):
                            urls.append({
                                'name': src.get('name') or src.get('url'),
                                'url': src.get('url'),
                                'source': 'DeepDarkCTI',
                                'category': cat_name
                            })
        except Exception as e:
            logger.error(f"Error loading DeepDarkCTI URLs: {e}")
            
        try:
            if rtm_repo_manager.repo_exists():
                group_profiles = rtm_repo_manager.fetch_group_profiles()
                for group in group_profiles:
                    if group.get('profile_url'):
                        urls.append({
                            'name': group.get('name') or group.get('profile_url'),
                            'url': group.get('profile_url'),
                            'source': 'Ransomware Matrix',
                            'category': 'Group Profiles'
                        })
        except Exception as e:
            logger.error(f"Error loading Ransomware Matrix URLs: {e}")

        try:
            apt_annotations_manager.ensure_available()
            for src in apt_annotations_manager.get_all_sources():
                if src.get('url'):
                    urls.append({
                        'name': src.get('name') or src.get('url'),
                        'url': src.get('url'),
                        'source': 'APT Research Base',
                        'category': 'Trusted CTI Sources'
                    })
        except Exception as e:
            logger.error(f"Error loading APT annotation URLs: {e}")
            
        return api_success({'urls': urls})
    except Exception as e:
        logger.error(f"api_cti_resources_urls error: {e}")
        return api_error(str(e), 500)


@app.route('/api/apt-annotations/status', methods=['GET'])
@require_auth
def api_apt_annotations_status():
    """API: APT Research Base status"""
    try:
        return api_success(apt_annotations_manager.get_status())
    except Exception as e:
        logger.error(f"api_apt_annotations_status error: {e}")
        return api_error(str(e), 500)


@app.route('/api/apt-annotations/refresh', methods=['POST'])
@require_auth
def api_apt_annotations_refresh():
    """API: Download/update APT annotations from Neo23x0 gist"""
    try:
        success = apt_annotations_manager.update_annotations()
        if success:
            status = apt_annotations_manager.get_status()
            return api_success(status, f"Downloaded {status['source_count']} reference sources")
        return api_error("Failed to download APT annotations", 500)
    except Exception as e:
        logger.error(f"api_apt_annotations_refresh error: {e}")
        return api_error(str(e), 500)


@app.route('/api/apt-annotations/sources', methods=['GET'])
@require_auth
def api_apt_annotations_sources():
    """API: List APT reference sources"""
    try:
        query = request.args.get('q', '').strip()
        sources = apt_annotations_manager.search_sources(query) if query else apt_annotations_manager.get_all_sources()
        return api_success({'sources': sources, 'count': len(sources)})
    except Exception as e:
        logger.error(f"api_apt_annotations_sources error: {e}")
        return api_error(str(e), 500)


@app.route('/api/apt-annotations/search-actor', methods=['GET'])
@require_auth
def api_apt_annotations_search_actor():
    """API: Build scoped search links for a threat actor across reference sources"""
    try:
        actor = request.args.get('actor', '').strip()
        if not actor:
            return api_error("Actor name required", 400)
        results = apt_annotations_manager.get_search_urls_for_actor(actor)
        return api_success({'actor': actor, 'sources': results, 'count': len(results)})
    except Exception as e:
        logger.error(f"api_apt_annotations_search_actor error: {e}")
        return api_error(str(e), 500)


@app.route('/api/apt-annotations/manual', methods=['POST'])
@require_auth
def api_apt_annotations_add_manual():
    """API: Add a manual APT reference source"""
    try:
        data = request.get_json() or {}
        url = data.get('url', '').strip()
        name = data.get('name', '').strip() or None
        if not url:
            return api_error("URL is required", 400)
        if not apt_annotations_manager.add_manual_source(url, name):
            return api_error("Source URL already exists", 400)
        return api_success(message="Source added")
    except Exception as e:
        logger.error(f"api_apt_annotations_add_manual error: {e}")
        return api_error(str(e), 500)


@app.route('/api/apt-annotations/manual', methods=['DELETE'])
@require_auth
def api_apt_annotations_delete_manual():
    """API: Delete a manual APT reference source"""
    try:
        data = request.get_json() or {}
        url = data.get('url', '').strip()
        if not url:
            return api_error("URL is required", 400)
        if not apt_annotations_manager.delete_manual_source(url):
            return api_error("Manual source not found", 404)
        return api_success(message="Source removed")
    except Exception as e:
        logger.error(f"api_apt_annotations_delete_manual error: {e}")
        return api_error(str(e), 500)


@app.route('/api/apt-annotations/favorite', methods=['POST'])
@require_auth
def api_apt_annotations_toggle_favorite():
    """API: Toggle favorite on an APT reference source"""
    try:
        data = request.get_json() or {}
        url = data.get('url', '').strip()
        if not url:
            return api_error("URL is required", 400)
        is_favorite = apt_annotations_manager.toggle_favorite(url)
        return api_success({'is_favorite': is_favorite})
    except Exception as e:
        logger.error(f"api_apt_annotations_toggle_favorite error: {e}")
        return api_error(str(e), 500)


@app.route('/api/sources/bulk-add-group', methods=['POST'])
@require_auth
def api_sources_bulk_add_group():
    """API: Add multiple sources to a group"""
    try:
        data = request.get_json()
        source_ids = data.get('source_ids', [])
        group_id = data.get('group_id')
        
        if not source_ids:
            return api_error('No sources selected', 400)
        if not group_id:
            return api_error('Group ID is required', 400)
        
        count = 0
        for source_id in source_ids:
            if db.add_source_to_group(source_id, group_id):
                count += 1
        
        return api_success(
            {'count': count},
            f'{count} source(s) added to group'
        )
    except Exception as e:
        logger.error(f"api_sources_bulk_add_group error: {e}")
        return api_error(str(e), 500)

@app.route('/api/sources/remove-group', methods=['POST'])
@require_auth
def api_sources_remove_group():
    """API: Remove a source from a group"""
    try:
        data = request.get_json()
        source_id = data.get('source_id')
        group_id = data.get('group_id')
        
        if not source_id:
            return api_error('Source ID is required', 400)
        if not group_id:
            return api_error('Group ID is required', 400)
        
        success = db.remove_source_from_group(source_id, group_id)
        if success:
            return api_success(message='Source removed from group successfully')
        else:
            return api_error('Source was not in this group', 400)
    except Exception as e:
        logger.error(f"api_sources_remove_group error: {e}")
        return api_error(str(e), 500)

@app.route('/api/groups/<int:group_id>', methods=['DELETE'])
@require_auth
def api_groups_delete(group_id):
    """API: Delete a group"""
    try:
        # Vérifier que ce n'est pas un groupe système
        group = db.get_group_by_id(group_id)
        if not group:
            return api_not_found('Group')
        
        group_name = group.get('name', '')
        # Empêcher la suppression des groupes système
        system_groups = ['default', 'True Positive', 'False Positive']
        is_tlp_group = group_name.startswith('TLP:')
        
        if group_name in system_groups or is_tlp_group:
            return api_error('Cannot delete system groups (default, TLP groups, True/False Positive)', 400)
        
        success = db.delete_group(group_id)
        if success:
            return api_success(message='Group deleted successfully')
        else:
            return api_not_found('Group')
    except Exception as e:
        logger.error(f"api_groups_delete error: {e}")
        return api_error(str(e), 500)

@app.route('/api/groups/get-by-name', methods=['GET'])
@require_auth
def api_groups_get_by_name():
    """API: Retrieve a group by its name"""
    try:
        name = request.args.get('name', '').strip()
        
        if not name:
            return api_error('Group name is required', 400)
        
        group = db.get_group_by_name(name)
        if group:
            return api_success({'group_id': group['id'], 'group': group})
        else:
            return api_not_found('Group')
    except Exception as e:
        logger.error(f"api_groups_get_by_name error: {e}")
        return api_error(str(e), 500)

@app.route('/api/sources/bulk-remove-group', methods=['POST'])
@require_auth
def api_sources_bulk_remove_group():
    """API: Remove multiple sources from a group"""
    try:
        data = request.get_json()
        source_ids = data.get('source_ids', [])
        group_id = data.get('group_id')
        
        if not source_ids:
            return api_error('No sources selected', 400)
        if not group_id:
            return api_error('Group ID is required', 400)
        
        count = 0
        for source_id in source_ids:
            if db.remove_source_from_group(source_id, group_id):
                count += 1
        
        return api_success(
            {'count': count},
            f'{count} source(s) removed from group'
        )
    except Exception as e:
        logger.error(f"api_sources_bulk_remove_group error: {e}")
        return api_error(str(e), 500)

@app.route('/api/ioc/<int:ioc_id>/delete', methods=['POST'])
@require_auth
def api_ioc_delete(ioc_id):
    """API: Permanently delete an IOC"""
    try:
        success = db.hard_delete_ioc(ioc_id)
        if success:
            # Clean up orphaned tags
            db.cleanup_orphaned_tags()
            return api_success(message='IOC permanently deleted')
        else:
            return api_not_found('IOC')
    except Exception as e:
        logger.error(f"api_ioc_delete error: {e}")
        return api_error(str(e), 500)

@app.route('/api/tags/cleanup', methods=['POST'])
@require_auth
def api_tags_cleanup():
    """API: Force cleanup of orphaned tags"""
    try:
        db.cleanup_orphaned_tags()
        return api_success(message='Orphaned tags cleaned up')
    except Exception as e:
        logger.error(f"api_tags_cleanup error: {e}")
        return api_error(str(e), 500)

# ========== ROUTES API SUPPRESSION SOURCES ==========

@app.route('/api/sources/<int:source_id>/delete', methods=['POST', 'DELETE'])
@require_auth
def api_source_delete(source_id):
    """API: Delete a single source (soft delete)"""
    try:
        if db.soft_delete_source(source_id):
            return api_success({'source_id': source_id}, 'Source moved to trash')
        else:
            return api_error('Source not found or already deleted', 404)
    except Exception as e:
        logger.error(f"api_source_delete error: {e}")
        return api_error(str(e), 500)

@app.route('/api/sources/bulk-delete', methods=['POST'])
@require_auth
def api_sources_bulk_delete():
    """API: Permanently delete multiple sources"""
    try:
        data = request.get_json()
        source_ids = data.get('source_ids', [])
        
        if not source_ids:
            return api_error('No source selected', 400)
        
        count = 0
        for source_id in source_ids:
            if db.hard_delete_source(source_id):
                count += 1
        
        return api_success(
            {'count': count},
            f'{count} source(s) permanently deleted'
        )
    except Exception as e:
        logger.error(f"api_sources_bulk_delete error: {e}")
        return api_error(str(e), 500)

@app.route('/api/settings/cleanup/all-sources', methods=['POST'])
@require_auth
def api_cleanup_all_sources():
    """API: Delete ALL sources and their IOCs"""
    try:
        count = db.delete_all_sources()
        return api_success(
            {'count': count},
            f'All sources ({count}) and their IOCs have been permanently deleted'
        )
    except Exception as e:
        logger.error(f"api_cleanup_all_sources error: {e}")
        return api_error(str(e), 500)

@app.route('/api/settings/cleanup/all-iocs', methods=['POST'])
@require_auth
def api_cleanup_all_iocs():
    """API: Delete ALL IOCs"""
    try:
        count = db.delete_all_iocs()
        return api_success(
            {'count': count},
            f'All IOCs ({count}) have been permanently deleted'
        )
    except Exception as e:
        logger.error(f"api_cleanup_all_iocs error: {e}")
        return api_error(str(e), 500)

@app.route('/api/settings/cleanup/all-stix-files', methods=['POST'])
@require_auth
def api_cleanup_all_stix_files():
    """API: Delete all STIX files from stix and stix_conversions directories"""
    try:
        deleted_count = 0
        
        # Delete files from stix directory
        stix_dir = OUTPUT_FOLDER / "stix"
        if stix_dir.exists():
            for file_path in stix_dir.glob("*.json"):
                try:
                    file_path.unlink()
                    deleted_count += 1
                except Exception as e:
                    logger.warning(f"Error deleting file {file_path}: {e}")
        
        # Delete files from stix_conversions directory
        stix_conv_dir = OUTPUT_FOLDER / "stix_conversions"
        if stix_conv_dir.exists():
            for file_path in stix_conv_dir.glob("*.json"):
                try:
                    file_path.unlink()
                    deleted_count += 1
                except Exception as e:
                    logger.warning(f"Error deleting file {file_path}: {e}")
        
        return api_success({'deleted_count': deleted_count}, f'Successfully deleted {deleted_count} STIX files')
    except Exception as e:
        logger.error(f"api_cleanup_all_stix_files error: {e}")
        return api_error(str(e), 500)

@app.route('/api/settings/cleanup/all-data', methods=['POST'])
@require_auth
def api_cleanup_all_data():
    """API: Comprehensive cleanup - delete all data and reset application"""
    try:
        import shutil
        
        results = {
            'uploads_deleted': 0,
            'outputs_deleted': 0,
            'stix_files_deleted': 0,
            'sources_deleted': 0,
            'iocs_deleted': 0,
            'groups_deleted': 0,
            'saved_stix_models_deleted': 0,
            'zettelforge_cleared': False,
            'settings_reset': False
        }
        
        # 1. Clean uploads directory
        if UPLOAD_FOLDER.exists():
            for item in UPLOAD_FOLDER.iterdir():
                try:
                    if item.is_file():
                        item.unlink()
                        results['uploads_deleted'] += 1
                    elif item.is_dir():
                        shutil.rmtree(item)
                        results['uploads_deleted'] += 1
                except Exception as e:
                    logger.warning(f"Error deleting upload item {item}: {e}")
        
        # 2. Clean outputs directory (except stix subdirectories)
        if OUTPUT_FOLDER.exists():
            for item in OUTPUT_FOLDER.iterdir():
                try:
                    if item.is_file():
                        item.unlink()
                        results['outputs_deleted'] += 1
                except Exception as e:
                    logger.warning(f"Error deleting output item {item}: {e}")
        
        # 3. Delete STIX files
        stix_dirs = [OUTPUT_FOLDER / "stix", OUTPUT_FOLDER / "stix_conversions"]
        for stix_dir in stix_dirs:
            if stix_dir.exists():
                for file_path in stix_dir.glob("*.json"):
                    try:
                        file_path.unlink()
                        results['stix_files_deleted'] += 1
                    except Exception as e:
                        logger.warning(f"Error deleting STIX file {file_path}: {e}")
        
        # 4. Delete relational analysis data and core database content
        with db.connection() as conn:
            cursor = conn.execute("SELECT COUNT(*) as count FROM saved_stix_models")
            results['saved_stix_models_deleted'] = cursor.fetchone()['count']
            conn.execute("DELETE FROM saved_stix_models")

            conn.execute("DELETE FROM ioc_ttp_links")
            conn.execute("DELETE FROM cross_refs")
            conn.execute("DELETE FROM ioc_tags")
            conn.execute("DELETE FROM tag_history")
            conn.execute("DELETE FROM tags")
            conn.execute("DELETE FROM generated_reports")

            # Delete all IOCs
            cursor = conn.execute("SELECT COUNT(*) as count FROM iocs")
            results['iocs_deleted'] = cursor.fetchone()['count']
            conn.execute("DELETE FROM iocs")

            # Delete all sources
            cursor = conn.execute("SELECT COUNT(*) as count FROM sources")
            results['sources_deleted'] = cursor.fetchone()['count']
            conn.execute("DELETE FROM sources")

            # Delete all groups
            cursor = conn.execute("SELECT COUNT(*) as count FROM groups")
            results['groups_deleted'] = cursor.fetchone()['count']
            conn.execute("DELETE FROM groups")

            # Delete all settings except essential ones
            conn.execute("DELETE FROM settings WHERE key NOT IN ('secret_key', 'auth_enabled')")
            results['settings_reset'] = True

            conn.commit()

        # 6. Clear CTI Memory local store (ZettelForge)
        try:
            from config import ZETTELFORGE_MEMORY_DIR
            if ZETTELFORGE_MEMORY_DIR.exists():
                shutil.rmtree(ZETTELFORGE_MEMORY_DIR, ignore_errors=True)
                ZETTELFORGE_MEMORY_DIR.mkdir(parents=True, exist_ok=True)
                results['zettelforge_cleared'] = True
        except Exception as e:
            logger.warning(f"Error clearing CTI Memory data: {e}")
        
        # Reset auth to disabled
        db.set_setting('auth_enabled', 'false')
        
        total_items = sum([
            results['uploads_deleted'],
            results['outputs_deleted'],
            results['stix_files_deleted'],
            results['sources_deleted'],
            results['iocs_deleted'],
            results['groups_deleted']
        ])
        
        message = (
            f"Complete reset successful. Deleted: "
            f"{results['uploads_deleted']} uploads, "
            f"{results['outputs_deleted']} outputs, "
            f"{results['stix_files_deleted']} STIX files, "
            f"{results['sources_deleted']} sources, "
            f"{results['iocs_deleted']} IOCs, "
            f"{results['groups_deleted']} groups, "
            f"{results['saved_stix_models_deleted']} STIX models. "
            f"Settings reset to defaults."
            + (" CTI Memory cleared." if results.get('zettelforge_cleared') else "")
        )
        
        return api_success(results, message)
    except Exception as e:
        logger.error(f"api_cleanup_all_data error: {e}")
        return api_error(str(e), 500)

# ========== ROUTES API SETTINGS ==========

@app.route('/api/settings/stix-producer', methods=['GET', 'POST'])
@require_auth
def api_settings_stix_producer():
    """Manage the producer identity used for new STIX domain objects."""
    try:
        if request.method == 'GET':
            return api_success({'name': db.get_setting('stix_producer_name', 'ODYSAFE CTI')})
        data = request.get_json(silent=True) or {}
        name = str(data.get('name') or '').strip()
        if not name or len(name) > 120:
            return api_error('Producer name must contain 1 to 120 characters', 400)
        db.set_setting('stix_producer_name', name)
        return api_success({'name': name}, 'Setting updated')
    except Exception as e:
        logger.error(f"api_settings_stix_producer error: {e}")
        return api_error(str(e), 500)

@app.route('/api/settings/auto-tag', methods=['GET', 'POST'])
@require_auth
def api_settings_auto_tag():
    """API: Manage auto-tagging"""
    try:
        if request.method == 'GET':
            enabled = db.get_setting('auto_tag_enabled', 'true')
            return api_success({'enabled': enabled.lower() == 'true'})
        
        elif request.method == 'POST':
            data = request.get_json()
            enabled = data.get('enabled', True)
            db.set_setting('auto_tag_enabled', 'true' if enabled else 'false')
            return api_success(message='Setting updated')
    except Exception as e:
        logger.error(f"api_settings_auto_tag error: {e}")
        return api_error(str(e), 500)

@app.route('/api/settings/storage', methods=['GET'])
@require_auth
def api_settings_storage():
    """API: Get storage information"""
    try:
        storage_info = get_storage_info()
        return api_success(storage_info)
    except Exception as e:
        logger.error(f"api_settings_storage error: {e}")
        return api_error(str(e), 500)

@app.route('/api/settings/source-rotation', methods=['GET', 'POST'])
@require_auth
def api_settings_source_rotation():
    """API: Manage automatic source rotation"""
    try:
        if request.method == 'GET':
            enabled = db.get_setting('auto_rotation_enabled', 'false')
            return api_success({'enabled': enabled.lower() == 'true'})
        
        elif request.method == 'POST':
            data = request.get_json()
            enabled = data.get('enabled', False)
            db.set_setting('auto_rotation_enabled', 'true' if enabled else 'false')
            return api_success(message='Setting updated')
    except Exception as e:
        logger.error(f"api_settings_source_rotation error: {e}")
        return api_error(str(e), 500)

@app.route('/api/settings/max-sources', methods=['GET', 'POST'])
@require_auth
def api_settings_max_sources():
    """API: Manage maximum sources limit"""
    try:
        if request.method == 'GET':
            max_sources = db.get_setting('max_sources', '20')
            return api_success({'max_sources': int(max_sources)})
        
        elif request.method == 'POST':
            data = request.get_json()
            max_sources = data.get('max_sources', 20)
            if max_sources < 1:
                return api_error('Maximum sources must be at least 1', 400)
            db.set_setting('max_sources', str(max_sources))
            return api_success(message='Setting updated')
    except Exception as e:
        logger.error(f"api_settings_max_sources error: {e}")
        return api_error(str(e), 500)

@app.route('/api/settings/recent-sources-limit', methods=['GET', 'POST'])
@require_auth
def api_settings_recent_sources_limit():
    """API: Manage recent sources display limit"""
    try:
        if request.method == 'GET':
            limit = db.get_setting('recent_sources_limit', '20')
            return api_success({'limit': int(limit)})
        
        elif request.method == 'POST':
            data = request.get_json()
            limit = data.get('limit', 20)
            if limit < 1:
                return api_error('Display limit must be at least 1', 400)
            db.set_setting('recent_sources_limit', str(limit))
            return api_success(message='Setting updated')
    except Exception as e:
        logger.error(f"api_settings_recent_sources_limit error: {e}")
        return api_error(str(e), 500)

@app.route('/api/settings/trash-cleanup-days', methods=['GET', 'POST'])
@require_auth
def api_settings_trash_cleanup_days():
    """API: Manage trash cleanup days"""
    try:
        if request.method == 'GET':
            days = db.get_setting('trash_cleanup_days', '5')
            return api_success({'days': int(days)})
        
        elif request.method == 'POST':
            data = request.get_json()
            days = data.get('days', 5)
            if days < 1:
                return api_error('Cleanup days must be at least 1', 400)
            db.set_setting('trash_cleanup_days', str(days))
            return api_success(message='Setting updated')
    except Exception as e:
        logger.error(f"api_settings_trash_cleanup_days error: {e}")
        return api_error(str(e), 500)

@app.route('/api/settings/cleanup-trash-now', methods=['POST'])
@require_auth
def api_settings_cleanup_trash_now():
    """API: Manually trigger trash cleanup"""
    try:
        days = int(db.get_setting('trash_cleanup_days', '5'))
        deleted_count = db.cleanup_trash(days)
        return api_success(
            {'deleted_count': deleted_count},
            f'{deleted_count} source(s) permanently deleted from trash'
        )
    except Exception as e:
        logger.error(f"api_settings_cleanup_trash_now error: {e}")
        return api_error(str(e), 500)

@app.route('/api/settings/export-zip', methods=['POST'])
@require_auth
def api_settings_export_zip():
    """API: Export platform backup (IOCs, sources, tags, config, analysis data, CTI Memory)."""
    try:
        from modules.backup_manager import build_export_zip

        temp_zip = tempfile.NamedTemporaryFile(delete=False, suffix='.zip')
        temp_zip_path = temp_zip.name
        temp_zip.close()

        with zipfile.ZipFile(temp_zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
            build_export_zip(zipf, db, github_repo_manager, rtm_repo_manager)
        
        # Generate filename with timestamp
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'cti-export-{timestamp}.zip'
        
        # Return ZIP file and schedule cleanup
        # Use a longer delay (5 minutes) to ensure file is downloaded
        def remove_file():
            try:
                if os.path.exists(temp_zip_path):
                    os.unlink(temp_zip_path)
                    logger.debug(f"Temporary ZIP file cleaned up: {temp_zip_path}")
            except Exception as e:
                logger.warning(f"Failed to clean up temporary ZIP file {temp_zip_path}: {e}")
        
        # Schedule cleanup after response is sent (5 minutes should be enough for download)
        threading.Timer(300.0, remove_file).start()
        
        return send_file(
            temp_zip_path,
            mimetype='application/zip',
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        logger.error(f"api_settings_export_zip error: {e}")
        return api_error(str(e), 500)

@app.route('/api/settings/import-zip', methods=['POST'])
@require_auth
def api_settings_import_zip():
    """API: Import platform backup from ZIP file."""
    try:
        from modules.backup_manager import import_from_zip_path

        if 'file' not in request.files:
            return api_error('No file provided', 400)

        file = request.files['file']
        if file.filename == '':
            return api_error('No file selected', 400)

        if not file.filename.endswith('.zip'):
            return api_error('File must be a ZIP file', 400)

        temp_zip = tempfile.NamedTemporaryFile(delete=False, suffix='.zip')
        temp_zip_path = temp_zip.name
        file.save(temp_zip_path)
        temp_zip.close()

        try:
            stats = import_from_zip_path(
                temp_zip_path,
                db,
                github_repo_manager,
                rtm_repo_manager,
            )
        finally:
            try:
                if os.path.exists(temp_zip_path):
                    os.unlink(temp_zip_path)
            except Exception as e:
                logger.warning(f"Failed to clean up temporary ZIP file {temp_zip_path}: {e}")

        parts = [
            f"{stats.get('iocs_imported', 0)} IOCs",
            f"{stats.get('sources_imported', 0)} sources",
            f"{stats.get('groups_imported', 0)} groups",
        ]
        if stats.get('tags_imported'):
            parts.append(f"{stats['tags_imported']} tag links")
        if stats.get('ioc_ttp_links_imported'):
            parts.append(f"{stats['ioc_ttp_links_imported']} MITRE IOC links")
        if stats.get('cross_refs_imported'):
            parts.append(f"{stats['cross_refs_imported']} cross_refs")
        if stats.get('source_templates_imported'):
            parts.append(f"{stats['source_templates_imported']} source templates")
        if stats.get('group_exclusions_imported'):
            parts.append(f"{stats['group_exclusions_imported']} group exclusions")
        if stats.get('saved_stix_models_imported'):
            parts.append(f"{stats['saved_stix_models_imported']} STIX models")
        if stats.get('zettelforge_files_imported'):
            parts.append(f"{stats['zettelforge_files_imported']} CTI Memory files")
        if stats.get('rtm_favorites_imported'):
            parts.append(f"{stats['rtm_favorites_imported']} RTM favorites")

        return api_success(
            stats,
            f"Import completed: {', '.join(parts)}",
        )
    except zipfile.BadZipFile:
        return api_error('Invalid ZIP file', 400)
    except json.JSONDecodeError as e:
        return api_error(f'Invalid JSON in ZIP file: {str(e)}', 400)
    except Exception as e:
        logger.error(f"api_settings_import_zip error: {e}")
        return api_error(str(e), 500)

@app.route('/api/outputs/<path:filename>/delete', methods=['DELETE'])
@require_auth
def api_output_delete(filename):
    """API: Delete an export file directly"""
    try:
        file_path = OUTPUT_FOLDER / filename
        if file_path.exists() and file_path.is_file():
            file_path.unlink()
            return api_success(message='File deleted')
        else:
            return api_not_found('File')
    except Exception as e:
        logger.error(f"api_output_delete error: {e}")
        return api_error(str(e), 500)

@app.route('/api/uploads/<path:filename>/delete', methods=['DELETE'])
@require_auth
def api_upload_delete(filename):
    """API: Delete an uploaded file directly"""
    try:
        file_path = UPLOAD_FOLDER / filename
        if file_path.exists() and file_path.is_file():
            file_path.unlink()
            return api_success(message='File deleted')
        else:
            return api_not_found('File')
    except Exception as e:
        logger.error(f"api_upload_delete error: {e}")
        return api_error(str(e), 500)

@app.route('/api/settings/outputs/recent', methods=['GET'])
@require_auth
def api_settings_outputs_recent():
    """API: Get recent output files"""
    try:
        from modules.storage_monitor import format_bytes
        
        recent_outputs = []
        max_files = 5
        
        # Get all output files from subfolders
        for subfolder in ['iocs', 'stix', 'reports']:
            subfolder_path = OUTPUT_FOLDER / subfolder
            if subfolder_path.exists():
                for file_path in subfolder_path.iterdir():
                    if file_path.is_file():
                        try:
                            mtime = file_path.stat().st_mtime
                            size = file_path.stat().st_size
                            recent_outputs.append({
                                'path': str(file_path),
                                'name': file_path.name,
                                'folder': subfolder,
                                'size': size,
                                'size_formatted': format_bytes(size),
                                'modified': datetime.fromtimestamp(mtime).isoformat(),
                                'modified_timestamp': mtime
                            })
                        except OSError:
                            continue
        
        # Sort by modification date (most recent first) and limit to max_files
        recent_outputs.sort(key=lambda x: x['modified_timestamp'], reverse=True)
        recent_outputs = recent_outputs[:max_files]
        
        return api_success({'outputs': recent_outputs})
    except Exception as e:
        logger.error(f"api_settings_outputs_recent error: {e}")
        return api_error(str(e), 500)

@app.route('/api/settings/cleanup/uploads', methods=['POST'])
@require_auth
def api_settings_cleanup_uploads():
    """API: Delete all uploaded files"""
    try:
        from modules.storage_monitor import format_bytes
        
        deleted_count = 0
        total_size = 0
        
        if UPLOAD_FOLDER.exists():
            for file_path in UPLOAD_FOLDER.iterdir():
                if file_path.is_file():
                    try:
                        size = file_path.stat().st_size
                        file_path.unlink()
                        deleted_count += 1
                        total_size += size
                        logger.info(f"Upload file deleted: {file_path.name}")
                    except Exception as e:
                        logger.warning(f"Unable to delete {file_path}: {e}")
        
        return api_success(
            {'deleted_count': deleted_count, 'total_size': total_size},
            f'Cleanup completed: {deleted_count} file(s) deleted ({format_bytes(total_size)})'
        )
    except Exception as e:
        logger.error(f"api_settings_cleanup_uploads error: {e}")
        return api_error(str(e), 500)

@app.route('/api/settings/cleanup/outputs', methods=['POST'])
@require_auth
def api_settings_cleanup_outputs():
    """API: Delete all output files"""
    try:
        from modules.storage_monitor import format_bytes
        
        deleted_count = 0
        total_size = 0
        
        # Delete files from all output subfolders
        for subfolder in ['iocs', 'stix', 'reports']:
            subfolder_path = OUTPUT_FOLDER / subfolder
            if subfolder_path.exists():
                for file_path in subfolder_path.iterdir():
                    if file_path.is_file():
                        try:
                            size = file_path.stat().st_size
                            file_path.unlink()
                            deleted_count += 1
                            total_size += size
                            logger.info(f"Output file deleted: {file_path.name}")
                        except Exception as e:
                            logger.warning(f"Unable to delete {file_path}: {e}")
        
        return api_success(
            {'deleted_count': deleted_count, 'total_size': total_size},
            f'Cleanup completed: {deleted_count} file(s) deleted ({format_bytes(total_size)})'
        )
    except Exception as e:
        logger.error(f"api_settings_cleanup_outputs error: {e}")
        return api_error(str(e), 500)

@app.route('/api/settings/outputs/<path:filepath>/download', methods=['GET'])
@require_auth
def api_settings_outputs_download(filepath):
    """API: Download an output file"""
    try:
        file_path = OUTPUT_FOLDER / filepath
        if file_path.exists() and file_path.is_file():
            return send_file(str(file_path), as_attachment=True, download_name=file_path.name)
        else:
            return api_not_found('File')
    except Exception as e:
        logger.error(f"api_settings_outputs_download error: {e}")
        return api_error(str(e), 500)

# ========== ROUTES API DeepDarkCTI ==========

@app.route('/api/cti-resources/download', methods=['POST'])
@require_auth
def api_cti_resources_download():
    """API: Download deepdarkcti repository"""
    try:
        logger.info("Downloading DeepDarkCTI repository...")
        
        success = github_repo_manager.download_repo()
        
        if success:
            return api_success({'message': 'Repository downloaded successfully'})
        else:
            return api_error("Repository download error", 500)
    except Exception as e:
        logger.error(f"api_cti_resources_download error: {e}", exc_info=True)
        return api_error(str(e), 500)

@app.route('/api/cti-resources/update', methods=['POST'])
@require_auth
def api_cti_resources_update():
    """API: Update repository (deletes cache + old repo + downloads new one)"""
    try:
        logger.info("Updating DeepDarkCTI repository...")
        
        success = github_repo_manager.update_repo()
        
        if success:
            return api_success({'message': 'Repository updated successfully'})
        else:
            return api_error("Repository update failed. Please check the logs for details.", 500)
    except RuntimeError as e:
        logger.error(f"Error api_cti_resources_update: {e}", exc_info=True)
        return api_error(str(e), 500)
    except Exception as e:
        logger.error(f"Error api_cti_resources_update: {e}", exc_info=True)
        return api_error(f"Repository update error: {str(e)}", 500)

@app.route('/api/cti-resources/source/delete', methods=['POST'])
@require_auth
def api_cti_resources_delete_source():
    """API: Delete a source"""
    try:
        data = request.get_json()
        category = data.get('category')
        source_url = data.get('url')
        is_manual = data.get('is_manual', False)
        
        if not source_url:
            return api_error("URL required", 400)
        
        # If it's a manual source, use specific method
        if is_manual or category == '_manual_sources':
            success = github_repo_manager.delete_manual_source(source_url)
        else:
            if not category:
                return api_error("Category required for repository sources", 400)
            success = github_repo_manager.delete_source(category, source_url)
        
        if success:
            return api_success({'message': 'Source deleted successfully'})
        else:
            return api_error("Deletion error", 500)
    except Exception as e:
        logger.error(f"api_cti_resources_delete_source error: {e}")
        return api_error(str(e), 500)

@app.route('/api/cti-resources/manual-source/add', methods=['POST'])
@require_auth
def api_cti_resources_add_manual_source():
    """API: Add a source manually"""
    try:
        data = request.get_json()
        url = data.get('url')
        name = data.get('name')
        description = data.get('description')
        
        if not url:
            return api_error("URL required", 400)
        
        # Validate that it's a valid URL
        if not url.startswith('http://') and not url.startswith('https://'):
            return api_error("Invalid URL. Must start with http:// or https://", 400)
        
        success = github_repo_manager.add_manual_source(url, name, description)
        
        if success:
            return api_success({'message': 'Source added successfully'})
        else:
            return api_error("This URL already exists", 400)
    except Exception as e:
        logger.error(f"api_cti_resources_add_manual_source error: {e}")
        return api_error(str(e), 500)

@app.route('/api/ransomware-tools/download', methods=['POST'])
@require_auth
def api_ransomware_tools_download():
    """API: Download Ransomware Tool Matrix repository"""
    try:
        logger.info("Downloading Ransomware Tool Matrix repository...")
        
        success = rtm_repo_manager.download_repo()
        
        if success:
            return api_success({'message': 'Repository downloaded successfully'})
        else:
            return api_error("Repository download error", 500)
    except Exception as e:
        logger.error(f"api_ransomware_tools_download error: {e}", exc_info=True)
        return api_error(str(e), 500)

@app.route('/api/ransomware-tools/update', methods=['POST'])
@require_auth
def api_ransomware_tools_update():
    """API: Update repository (deletes cache + old repo + downloads new one)"""
    try:
        logger.info("Updating Ransomware Tool Matrix repository...")
        
        success = rtm_repo_manager.update_repo()
        
        if success:
            return api_success({'message': 'Repository updated successfully'})
        else:
            return api_error("Repository update failed. Please check the logs for details.", 500)
    except RuntimeError as e:
        logger.error(f"Error api_ransomware_tools_update: {e}", exc_info=True)
        return api_error(str(e), 500)
    except Exception as e:
        logger.error(f"Error api_ransomware_tools_update: {e}", exc_info=True)
        return api_error(f"Repository update error: {str(e)}", 500)

@app.route('/api/ransomware-tools/favorite/toggle', methods=['POST'])
@require_auth
def api_ransomware_tools_toggle_favorite():
    """API: Add/remove a source from favorites"""
    try:
        data = request.get_json()
        url = data.get('url')
        
        if not url:
            return api_error("URL required", 400)
        
        is_favorite = rtm_repo_manager.toggle_favorite(url)
        
        return api_success({
            'message': 'Favorite updated successfully',
            'is_favorite': is_favorite
        })
    except Exception as e:
        logger.error(f"api_ransomware_tools_toggle_favorite error: {e}")
        return api_error(str(e), 500)

@app.route('/api/cti-resources/favorite/toggle', methods=['POST'])
@require_auth
def api_cti_resources_toggle_favorite():
    """API: Add/remove a source from favorites"""
    try:
        data = request.get_json()
        url = data.get('url')
        
        if not url:
            return api_error("URL required", 400)
        
        is_favorite = github_repo_manager.toggle_favorite(url)
        
        return api_success({
            'message': 'Favorite updated successfully',
            'is_favorite': is_favorite
        })
    except Exception as e:
        logger.error(f"api_cti_resources_toggle_favorite error: {e}")
        return api_error(str(e), 500)

@app.route('/api/iocs/add-manual', methods=['POST'])
@require_auth
def api_iocs_add_manual():
    """API: Add a single IOC manually"""
    try:
        data = request.get_json()
        ioc_type = data.get('ioc_type', '').strip()
        ioc_value = data.get('ioc_value', '').strip()
        context = data.get('context', 'Manual IOC entry').strip()
        tlp = (data.get('tlp') or 'WHITE').strip()
        validation_status = (data.get('validation_status') or 'Unverified').strip()
        confidence = (data.get('confidence') or 'Unknown').strip()
        
        if not ioc_type or not ioc_value:
            return api_error("IOC type and value are required", 400)
        
        # Get or create "Manual" source
        source_context = context if context else 'Manual IOC entries'
        
        with db.connection() as conn:
            ody = conn.cursor()
            
            # Check if "Manual" source exists
            ody.execute("SELECT id FROM sources WHERE name = 'Manual' AND is_deleted = 0 LIMIT 1")
            source_row = ody.fetchone()
            
            if source_row:
                source_id = source_row[0]
            else:
                # Create "Manual" source with explicit local timestamp
                from database import get_local_timestamp
                ody.execute("""
                    INSERT INTO sources (name, context, source_type, created_at)
                    VALUES (?, ?, ?, ?)
                """, ('Manual', source_context, 'manual', get_local_timestamp()))
                source_id = ody.lastrowid
                
                # Add to default group
                default_group = db.get_group_by_name("default", conn)
                if default_group:
                    try:
                        ody.execute("""
                            INSERT OR IGNORE INTO source_groups (source_id, group_id)
                            VALUES (?, ?)
                        """, (source_id, default_group['id']))
                    except Exception as e:
                        logger.warning(f"Failed to add manual source to default group: {e}")
                conn.commit()
        
        # Create the IOC
        ioc_id = db.create_ioc(
            source_id,
            ioc_type,
            ioc_value,
            ioc_value,
            tlp=tlp,
            validation_status=validation_status,
            confidence=confidence,
        )
        
        return api_success({
            'message': 'IOC added successfully',
            'ioc_id': ioc_id
        })
    except Exception as e:
        logger.error(f"api_iocs_add_manual error: {e}")
        return api_error(str(e), 500)

# ========== ROUTES API Data-Shield IPv4 Blocklist ==========

@app.route('/api/data-shield/status', methods=['GET'])
@require_auth
def api_data_shield_status():
    """API: Get Data-Shield repository status"""
    try:
        status = data_shield_repo_manager.get_status()
        
        # Count imported IPs from database
        imported_count = 0
        try:
            with db.connection() as conn:
                ody = conn.cursor()
                ody.execute("""
                    SELECT COUNT(*) as count FROM iocs i
                    JOIN sources s ON i.source_id = s.id
                    WHERE s.name = 'data-shield' AND s.is_deleted = 0
                """)
                row = ody.fetchone()
                if row:
                    imported_count = row['count']
        except Exception as e:
            logger.warning(f"Error counting imported IPs: {e}")
        
        status['imported_ip_count'] = imported_count
        
        return api_success(status)
    except Exception as e:
        logger.error(f"api_data_shield_status error: {e}")
        return api_error(str(e), 500)

@app.route('/api/data-shield/download', methods=['POST'])
@require_auth
def api_data_shield_download():
    """API: Download/update Data-Shield repository"""
    try:
        logger.info("Downloading malicious IPv4 IP list from Data-Shield...")
        
        success = data_shield_repo_manager.download_blocklist()
        
        if success:
            return api_success({'message': 'Repository downloaded successfully'})
        else:
            return api_error("Repository download error", 500)
    except Exception as e:
        logger.error(f"api_data_shield_download error: {e}", exc_info=True)
        return api_error(str(e), 500)

def _import_data_shield_ips() -> Dict:
    """Helper function to import IPs from Data-Shield blocklist"""
    # Check if blocklist file exists
    if not data_shield_repo_manager.blocklist_file_exists():
        raise ValueError("Blocklist file not found. Please download it first.")
    
    # Get IPs from blocklist
    ips = data_shield_repo_manager.get_blocklist_ips()
    
    if not ips:
        raise ValueError("No valid IPs found in blocklist file")
    
    logger.info(f"Importing {len(ips)} IPs from Data-Shield blocklist...")
    
    # Get or create source "data-shield"
    source_name = "data-shield"
    source_context = "Data-Shield IPv4 Blocklist - Malicious IP addresses blocklist"
    
    # Check if source already exists
    source_id = None
    with db.connection() as conn:
        ody = conn.cursor()
        ody.execute("SELECT id FROM sources WHERE name = ? AND is_deleted = 0 LIMIT 1", (source_name,))
        source_row = ody.fetchone()
        
        if source_row:
            source_id = source_row['id']
            logger.info(f"Using existing source: {source_id}")
    
    if source_id:
        # Delete all existing IOCs from this source before importing new ones
        logger.info(f"Deleting existing IOCs from source {source_id}...")
        with db.connection() as conn:
            ody = conn.cursor()
            try:
                # Delete IOC group associations
                ody.execute("""
                    DELETE FROM ioc_groups 
                    WHERE ioc_id IN (SELECT id FROM iocs WHERE source_id = ?)
                """, (source_id,))
                
                # Delete IOC source group exclusions
                ody.execute("""
                    DELETE FROM ioc_source_group_exclusions 
                    WHERE ioc_id IN (SELECT id FROM iocs WHERE source_id = ?)
                """, (source_id,))
                
                # Delete IOC tags
                ody.execute("""
                    DELETE FROM ioc_tags 
                    WHERE ioc_id IN (SELECT id FROM iocs WHERE source_id = ?)
                """, (source_id,))
                
                # Delete IOCs
                ody.execute("DELETE FROM iocs WHERE source_id = ?", (source_id,))
                deleted_count = ody.rowcount
                logger.info(f"Deleted {deleted_count} existing IOCs from source")
            except Exception as e:
                logger.error(f"Error deleting existing IOCs: {e}")
                raise
    else:
        # Create new source
        source_id = db.create_source(
            name=source_name,
            context=source_context,
            source_type='file_upload',
            file_path=str(data_shield_repo_manager.blocklist_file),
            original_filename="prod_data-shield_ipv4_blocklist.txt"
        )
        logger.info(f"Created new source: {source_id}")
    
    # Get True Positive group
    true_positive_group = db.get_group_by_name("True Positive")
    if not true_positive_group:
        raise ValueError("True Positive group not found. Please ensure the database is properly initialized.")
    
    # Remove source from default group if it exists (using direct SQL query)
    default_group = db.get_group_by_name("default")
    if default_group:
        try:
            with db.connection() as conn:
                ody = conn.cursor()
                ody.execute("""
                    SELECT 1 FROM source_groups 
                    WHERE source_id = ? AND group_id = ?
                """, (source_id, default_group['id']))
                if ody.fetchone():
                    db.remove_source_from_group(source_id, default_group['id'])
                    logger.info("Removed source from default group")
        except Exception as e:
            logger.warning(f"Could not remove source from default group: {e}")
    
    # Add source to True Positive group
    # Note: add_source_to_group() automatically:
    # - Removes from other positive groups (True/False Positive) before adding
    # - Uses INSERT OR IGNORE, so it's safe to call even if already in the group
    db.add_source_to_group(source_id, true_positive_group['id'])
    logger.info("Source added to True Positive group (or already was)")
    
    # Import IPs using batch operations for performance
    from database import get_local_timestamp
    
    local_ts = get_local_timestamp()
    batch_size = 1000
    
    logger.info(f"Importing {len(ips)} IPs using batch operations...")
    
    try:
        with db.connection() as conn:
            ody = conn.cursor()
            
            # Batch insert IOCs using INSERT OR IGNORE
            # Since we deleted all IOCs from this source before, all should insert successfully
            for i in range(0, len(ips), batch_size):
                batch = ips[i:i + batch_size]
                ioc_values = [(source_id, 'ipv4', ip, ip, local_ts, local_ts, local_ts) for ip in batch]
                
                ody.executemany("""
                    INSERT OR IGNORE INTO iocs (source_id, ioc_type, ioc_value, raw_value, first_seen, last_seen, created_at)
                    VALUES (?, ?, ?, ?, ?, ?, ?)
                """, ioc_values)
                
                # Commit every batch
                conn.commit()
                
                # Log progress
                if (i + batch_size) % 10000 == 0 or (i + batch_size) >= len(ips):
                    logger.info(f"Processing IPs... ({i + batch_size}/{len(ips)})")
            
            # Get all IOC IDs that were just inserted for this source
            ody.execute("""
                SELECT id FROM iocs 
                WHERE source_id = ? AND ioc_type = 'ipv4'
            """, (source_id,))
            ioc_ids = [row[0] for row in ody.fetchall()]
            imported_count = len(ioc_ids)
            
            # Note: IOCs inherit "True Positive" group from their source (already assigned at line 3738)
            # No need to assign directly to each IOC to avoid duplication
            
            skipped_count = len(ips) - imported_count
    except Exception as e:
        logger.error(f"Error during batch import: {e}")
        raise
    
    logger.info(f"Import completed: {imported_count} imported, {skipped_count} skipped")
    
    return {
        'imported_count': imported_count,
        'skipped_count': skipped_count,
        'total_count': len(ips)
    }

@app.route('/api/data-shield/update', methods=['POST'])
@require_auth
def api_data_shield_update():
    """API: Update repository (deletes cache + old repo + downloads new one) and auto-import IPs"""
    try:
        logger.info("Downloading malicious IPv4 IP list from Data-Shield...")
        
        success = data_shield_repo_manager.update_blocklist()
        
        if not success:
            return api_error("Repository update failed. Please check the logs for details.", 500)
        
        # Automatically import IPs after successful download
        try:
            import_result = _import_data_shield_ips()
            return api_success({
                'message': f'IP list downloaded and {import_result["imported_count"]} IPs imported successfully',
                'imported_count': import_result['imported_count'],
                'skipped_count': import_result['skipped_count'],
                'total_count': import_result['total_count']
            })
        except Exception as import_error:
            logger.warning(f"IP list downloaded but import failed: {import_error}")
            return api_success({
                'message': 'IP list downloaded successfully, but IP import failed. You can import manually.',
                'import_error': str(import_error)
            })
            
    except RuntimeError as e:
        logger.error(f"Error api_data_shield_update: {e}", exc_info=True)
        return api_error(str(e), 500)
    except Exception as e:
        logger.error(f"Error api_data_shield_update: {e}", exc_info=True)
        return api_error(f"Repository update error: {str(e)}", 500)

@app.route('/api/data-shield/import', methods=['POST'])
@require_auth
def api_data_shield_import():
    """API: Import IPs from Data-Shield blocklist"""
    try:
        import_result = _import_data_shield_ips()
        
        return api_success({
            'message': f'Import completed: {import_result["imported_count"]} IPs imported, {import_result["skipped_count"]} skipped',
            'imported_count': import_result['imported_count'],
            'skipped_count': import_result['skipped_count'],
            'total_count': import_result['total_count']
        })
        
    except ValueError as e:
        return api_error(str(e), 400)
    except Exception as e:
        logger.error(f"api_data_shield_import error: {e}", exc_info=True)
        return api_error(str(e), 500)

# ========== ROUTES API MITRE ATT&CK ==========

@app.route('/api/mitre-attack/status', methods=['GET'])
@require_auth
def api_mitre_attack_status():
    """API: Get MITRE ATT&CK data status"""
    try:
        file_exists = mitre_json_exists()
        stats = {}
        
        if file_exists:
            parser = MitreStixParser()
            if parser.load():
                stats = parser.get_stats()
        
        return api_success({
            'file_exists': file_exists,
            'file_path': str(get_mitre_json_path()),
            'counts': stats
        })
    except Exception as e:
        logger.error(f"api_mitre_attack_status error: {e}", exc_info=True)
        return api_error(str(e), 500)

@app.route('/api/mitre-attack/import', methods=['POST'])
@require_auth
def api_mitre_attack_import():
    """API: Reload MITRE ATT&CK data from JSON file"""
    try:
        if not mitre_json_exists():
            return api_error('MITRE ATT&CK JSON file not found', 404)
        
        parser = MitreStixParser()
        if parser.load():
            stats = parser.get_stats()
            return api_success({
                'message': 'MITRE ATT&CK data loaded successfully',
                'loaded': stats
            })
        else:
            return api_error('Failed to load MITRE ATT&CK data', 500)
    except Exception as e:
        logger.error(f"api_mitre_attack_import error: {e}", exc_info=True)
        return api_error(str(e), 500)

@app.route('/api/mitre-attack/download', methods=['POST'])
@require_auth
def api_mitre_attack_download():
    """API: Download MITRE ATT&CK JSON (pinned snapshot by default)."""
    try:
        import urllib.request
        import shutil
        from modules.pinned_sources import get_mitre_enterprise_attack_url

        use_latest = request.args.get('latest', '').lower() in ('1', 'true', 'yes')
        mitre_url = get_mitre_enterprise_attack_url(use_latest=use_latest)
        mitre_path = get_mitre_json_path()

        # Backup existing file if present
        if mitre_path.exists():
            backup_path = mitre_path.with_suffix('.json.backup')
            shutil.copy(mitre_path, backup_path)
            logger.info(f"Existing MITRE file backed up to: {backup_path}")

        # Download new file
        logger.info(f"Downloading MITRE ATT&CK from: {mitre_url}")
        urllib.request.urlretrieve(mitre_url, mitre_path)

        file_size = mitre_path.stat().st_size
        logger.info(f"MITRE ATT&CK downloaded successfully: {file_size} bytes")

        return api_success({
            'message': 'MITRE ATT&CK data downloaded successfully',
            'file_path': str(mitre_path),
            'file_size': file_size,
            'reload_needed': True
        })
    except Exception as e:
        logger.error(f"api_mitre_attack_download error: {e}", exc_info=True)
        return api_error(f'Failed to download MITRE ATT&CK data: {str(e)}', 500)

@app.route('/api/mitre-attack/tactics', methods=['GET'])
@require_auth
def api_mitre_attack_tactics():
    """API: Get all MITRE ATT&CK tactics"""
    try:
        parser = MitreStixParser()
        if not parser.load():
            return api_error('MITRE ATT&CK data not available', 503)
        tactics = parser.get_tactics()
        return api_success({'tactics': tactics, 'count': len(tactics)})
    except Exception as e:
        logger.error(f"api_mitre_attack_tactics error: {e}", exc_info=True)
        return api_error(str(e), 500)

@app.route('/api/mitre-attack/techniques', methods=['GET'])
@require_auth
def api_mitre_attack_techniques():
    """API: Get MITRE ATT&CK techniques (optionally filtered by tactic)"""
    try:
        parser = MitreStixParser()
        if not parser.load():
            return api_error('MITRE ATT&CK data not available', 503)
        
        tactic_shortname = request.args.get('tactic', '').strip() or None
        search = request.args.get('search', '').strip() or None
        
        if search:
            results = parser.search(search)
            techniques = results.get('techniques', [])
        elif tactic_shortname:
            techniques = parser.get_techniques_for_tactic(tactic_shortname)
        else:
            techniques = parser.get_techniques()
        
        return api_success({'techniques': techniques, 'count': len(techniques)})
    except Exception as e:
        logger.error(f"api_mitre_attack_techniques error: {e}", exc_info=True)
        return api_error(str(e), 500)

@app.route('/api/mitre-attack/techniques/<technique_id>', methods=['GET'])
@require_auth
def api_mitre_attack_technique_detail(technique_id):
    """API: Get single MITRE ATT&CK technique with related data including mitigations and IOCs"""
    try:
        parser = MitreStixParser()
        if not parser.load():
            return api_error('MITRE ATT&CK data not available', 503)
        
        tech_id = technique_id.upper()
        all_techniques = parser.get_techniques()
        technique = next((t for t in all_techniques if t['id'] == tech_id), None)
        
        if not technique:
            return api_not_found('Technique not found')
        
        # Get related data from MITRE JSON
        mitigations = parser.get_mitigations_for_technique(technique['stix_id'])
        groups = parser.get_groups_for_technique(technique['stix_id'])
        
        # Get IOCs linked to this technique from database
        ioc_count = 0
        linked_iocs = []
        try:
            with db.connection() as conn:
                # Count IOCs with TTP matching this technique
                cursor = conn.execute("""
                    SELECT COUNT(DISTINCT i.id) as count 
                    FROM iocs i
                    JOIN ioc_ttp_links itt ON i.id = itt.ioc_id
                    WHERE itt.technique_id = ? AND i.is_deleted = 0
                """, (tech_id,))
                row = cursor.fetchone()
                ioc_count = row['count'] if row else 0
                
                # Get linked IOC details
                cursor = conn.execute("""
                    SELECT DISTINCT i.id, i.ioc_value, i.ioc_type, i.validation_status AS verdict
                    FROM iocs i
                    JOIN ioc_ttp_links itt ON i.id = itt.ioc_id
                    WHERE itt.technique_id = ? AND i.is_deleted = 0
                    LIMIT 10
                """, (tech_id,))
                linked_iocs = [dict(r) for r in cursor.fetchall()]
        except Exception as e:
            logger.warning(f"Could not fetch IOC links for technique {tech_id}: {e}")
        
        # Get ransomware groups using this technique
        ransomware_groups = []
        try:
            # Map technique to tools (static mapping for now)
            TOOL_TECHNIQUE_MAP = {
                'T1003.001': ['Mimikatz', 'Seatbelt'],
                'T1003.004': ['Mimikatz'],
                'T1055': ['Cobalt Strike', 'Metasploit'],
                'T1059.001': ['PowerShell Empire', ' Cobalt Strike'],
                'T1569.002': ['PsExec'],
                'T1021.006': ['PsExec', 'Windows Admin Shares'],
                'T1078': ['Valid Accounts'],
                'T1486': ['Ransomware'],
            }
            tools = TOOL_TECHNIQUE_MAP.get(tech_id, [])
            ransomware_groups = tools  # Simplified - would need actual ransomware matrix integration
        except Exception as e:
            logger.warning(f"Could not fetch ransomware groups for technique {tech_id}: {e}")
        
        return api_success({
            'technique': technique,
            'mitigations': mitigations,
            'groups': groups,
            'ioc_count': ioc_count,
            'linked_iocs': linked_iocs,
            'ransomware_tools': ransomware_groups
        })
    except Exception as e:
        logger.error(f"api_mitre_attack_technique_detail error: {e}", exc_info=True)
        return api_error(str(e), 500)


@app.route('/api/mitre-attack/search', methods=['GET'])
@require_auth
def api_mitre_attack_search():
    """API: Search MITRE ATT&CK techniques by keyword"""
    try:
        parser = MitreStixParser()
        if not parser.load():
            return api_error('MITRE ATT&CK data not available', 503)
        
        query = request.args.get('q', '').strip().lower()
        if not query:
            return api_error('Search query required', 400)
        
        results = parser.search(query)
        
        return api_success({
            'query': query,
            'techniques': results.get('techniques', [])[:20],
            'groups': results.get('groups', [])[:10],
            'total_techniques': len(results.get('techniques', [])),
            'total_groups': len(results.get('groups', []))
        })
    except Exception as e:
        logger.error(f"api_mitre_attack_search error: {e}", exc_info=True)
        return api_error(str(e), 500)


@app.route('/api/mitre-attack/iocs-by-technique/<technique_id>', methods=['GET'])
@require_auth
def api_mitre_attack_iocs_by_technique(technique_id):
    """API: Get all IOCs linked to a specific MITRE technique"""
    try:
        tech_id = technique_id.upper()
        with db.connection() as conn:
            cursor = conn.cursor()
            cursor.execute("""
                SELECT i.id, i.ioc_value, i.ioc_type, i.tlp, s.name AS source_name
                FROM ioc_ttp_links itl
                JOIN iocs i ON itl.ioc_id = i.id
                JOIN sources s ON i.source_id = s.id
                WHERE itl.technique_id = ?
                ORDER BY i.created_at DESC
            """, (tech_id,))
            iocs = [dict(zip([c[0] for c in cursor.description], row)) for row in cursor.fetchall()]
            return api_success({
                'technique_id': tech_id,
                'iocs': iocs
            })
    except Exception as e:
        logger.error(f"api_mitre_attack_iocs_by_technique error: {e}", exc_info=True)
        return api_error(str(e), 500)


@app.route('/api/mitre-attack/linked-iocs', methods=['GET'])
@require_auth
def api_mitre_attack_linked_iocs():
    """API: Get all IOCs linked to MITRE ATT&CK techniques with full details"""
    try:
        with db.connection() as conn:
            cursor = conn.cursor()
            cursor.execute("""
                SELECT i.id as ioc_id, i.ioc_value, i.ioc_type, i.tlp, s.name AS source_name,
                       itl.technique_id
                FROM ioc_ttp_links itl
                JOIN iocs i ON itl.ioc_id = i.id
                JOIN sources s ON i.source_id = s.id
                WHERE i.is_deleted = 0
                ORDER BY itl.technique_id, i.created_at DESC
            """)
            rows = cursor.fetchall()
            data = [dict(zip([c[0] for c in cursor.description], row)) for row in rows]
            return api_success({
                'count': len(data),
                'data': data
            })
    except Exception as e:
        logger.error(f"api_mitre_attack_linked_iocs error: {e}", exc_info=True)
        return api_error(str(e), 500)


@app.route('/api/mitre-attack/my-ttps', methods=['GET'])
@require_auth
def api_mitre_attack_my_ttps():
    """API: Get MITRE techniques that have linked IOCs in the database"""
    try:
        with db.connection() as conn:
            cursor = conn.execute("""
                SELECT DISTINCT itt.technique_id, COUNT(i.id) as ioc_count
                FROM ioc_ttp_links itt
                JOIN iocs i ON itt.ioc_id = i.id
                WHERE i.is_deleted = 0
                GROUP BY itt.technique_id
                ORDER BY ioc_count DESC
            """)
            
            my_ttps = [dict(r) for r in cursor.fetchall()]
            
            # Enrich with technique details from MITRE
            parser = MitreStixParser()
            if parser.load():
                all_techniques = {t['id']: t for t in parser.get_techniques()}
                for ttp in my_ttps:
                    tech = all_techniques.get(ttp['technique_id'], {})
                    ttp['name'] = tech.get('name', 'Unknown')
                    ttp['tactics'] = tech.get('tactics', [])
            
            return api_success({
                'ttp_count': len(my_ttps),
                'ttps': my_ttps
            })
    except Exception as e:
        logger.error(f"api_mitre_attack_my_ttps error: {e}", exc_info=True)
        return api_error(str(e), 500)


@app.route('/api/mitre-attack/link-ioc-to-technique', methods=['POST'])
@require_auth
def api_mitre_attack_link_ioc_to_technique():
    """API: Link an IOC to a MITRE ATT&CK technique"""
    try:
        data = request.get_json()
        if not data:
            return api_error('JSON data required', 400)
        
        ioc_id = data.get('ioc_id')
        technique_id = data.get('technique_id')
        confidence = data.get('confidence', 'medium')
        notes = data.get('notes', '')
        
        if not ioc_id or not technique_id:
            return api_error('ioc_id and technique_id required', 400)
        
        # Validate technique exists
        parser = MitreStixParser()
        if parser.load():
            all_techniques = parser.get_techniques()
            tech = next((t for t in all_techniques if t['id'] == technique_id.upper()), None)
            if not tech:
                return api_error(f'Technique {technique_id} not found', 404)
        
        # Create link
        with db.connection() as conn:
            conn.execute("""
                INSERT OR REPLACE INTO ioc_ttp_links 
                (ioc_id, technique_id, confidence, notes, created_at)
                VALUES (?, ?, ?, ?, datetime('now'))
            """, (ioc_id, technique_id.upper(), confidence, notes))
            conn.commit()
        
        return api_success({
            'message': f'IOC linked to technique {technique_id}',
            'ioc_id': ioc_id,
            'technique_id': technique_id,
            'technique_name': tech.get('name', '') if tech else ''
        })
    except Exception as e:
        logger.error(f"api_mitre_attack_link_ioc_to_technique error: {e}", exc_info=True)
        return api_error(str(e), 500)
        
        technique['mitigations'] = mitigations
        technique['groups'] = groups
        
        return api_success({'technique': technique})
    except Exception as e:
        logger.error(f"api_mitre_attack_technique_detail error: {e}", exc_info=True)
        return api_error(str(e), 500)

@app.route('/api/mitre-attack/groups', methods=['GET'])
@require_auth
def api_mitre_attack_groups():
    """API: Get MITRE ATT&CK groups (APTs)"""
    try:
        parser = MitreStixParser()
        if not parser.load():
            return api_error('MITRE ATT&CK data not available', 503)
        
        groups = parser.get_groups()
        search = request.args.get('search', '').strip() or None
        
        if search:
            query = search.lower()
            groups = [g for g in groups if query in g['id'].lower() 
                     or query in g['name'].lower()
                     or any(query in a.lower() for a in g.get('aliases', []))]
        
        return api_success({'groups': groups, 'count': len(groups)})
    except Exception as e:
        logger.error(f"api_mitre_attack_groups error: {e}", exc_info=True)
        return api_error(str(e), 500)

@app.route('/api/mitre-attack/software', methods=['GET'])
@require_auth
def api_mitre_attack_software():
    """API: Get MITRE ATT&CK software"""
    try:
        parser = MitreStixParser()
        if not parser.load():
            return api_error('MITRE ATT&CK data not available', 503)
        
        software = parser.get_software()
        s_type = request.args.get('type', '').strip() or None
        
        if s_type:
            software = [s for s in software if s.get('type') == s_type]
        
        return api_success({'software': software, 'count': len(software)})
    except Exception as e:
        logger.error(f"api_mitre_attack_software error: {e}", exc_info=True)
        return api_error(str(e), 500)

@app.route('/api/mitre-attack/matrix', methods=['GET'])
@require_auth
def api_mitre_attack_matrix():
    """API: Get full ATT&CK matrix with 3 levels (tactics->techniques->subtechniques)"""
    try:
        parser = MitreStixParser()
        if not parser.load():
            return api_error('MITRE ATT&CK data not available', 503)
        
        matrix = parser.get_matrix_view()
        return api_success(matrix)
    except Exception as e:
        logger.error(f"api_mitre_attack_matrix error: {e}", exc_info=True)
        return api_error(str(e), 500)

@app.route('/api/mitre-attack/groups/<group_id>/techniques', methods=['GET'])
@require_auth
def api_mitre_attack_group_techniques(group_id):
    """API: Get techniques used by a specific threat group"""
    try:
        parser = MitreStixParser()
        if not parser.load():
            return api_error('MITRE ATT&CK data not available', 503)
        
        # Find group by ID
        groups = parser.get_groups()
        group = next((g for g in groups if g['id'] == group_id.upper()), None)
        
        if not group:
            return api_not_found('Group not found')
        
        techniques = parser.get_techniques_for_group(group['stix_id'])
        return api_success({
            'group': group,
            'techniques': techniques,
            'count': len(techniques)
        })
    except Exception as e:
        logger.error(f"api_mitre_attack_group_techniques error: {e}", exc_info=True)
        return api_error(str(e), 500)

@app.route('/api/ioc/count-by-ttp', methods=['GET'])
@require_auth
def api_ioc_count_by_ttp():
    """GET /api/ioc/count-by-ttp?technique_id=T1486
    Returns the number of TTP-type IOCs in the database that match the given technique ID.
    """
    technique_id = request.args.get('technique_id', '').strip().upper()
    if not technique_id:
        return api_error('technique_id is required', 400)
    try:
        count = db.count_iocs_by_ttp(technique_id)
        return api_success({'technique_id': technique_id, 'count': count})
    except Exception as e:
        logger.error(f"api_ioc_count_by_ttp error: {e}", exc_info=True)
        return api_error(str(e), 500)


@app.route('/api/mitre-attack/groups/<group_id>/ioc-overlap', methods=['GET'])
@require_auth
def api_mitre_group_ioc_overlap(group_id):
    """GET /api/mitre-attack/groups/<gid>/ioc-overlap
    Returns the techniques used by this group that also have matching TTP IOCs in the DB.
    """
    try:
        parser = MitreStixParser()
        if not parser.load():
            return api_error('MITRE ATT&CK data not available', 503)
        groups = parser.get_groups()
        group = next((g for g in groups if g['id'] == group_id.upper()), None)
        if not group:
            return api_not_found('Group not found')
        techniques = parser.get_techniques_for_group(group['stix_id'])
        tech_ids = [t['id'] for t in techniques]
        matched_ids = db.get_workspace_technique_coverage(tech_ids)
        matched = [t for t in techniques if t['id'].upper() in matched_ids]
        return api_success({
            'group_id': group['id'],
            'group_name': group['name'],
            'total_techniques': len(techniques),
            'matched_count': len(matched),
            'matched_techniques': matched,
        })
    except Exception as e:
        logger.error(f"api_mitre_group_ioc_overlap error: {e}", exc_info=True)
        return api_error(str(e), 500)


# ========== GLOBAL SEARCH ==========

def _global_search_mitre(q: str) -> List[Dict]:
    """Fast MITRE lookup for Ctrl+K: SQLite first, cached STIX parser as fallback."""
    q = (q or '').strip()
    if not q:
        return []
    q_upper = q.upper()
    escaped_q = q.replace("%", "\\%").replace("_", "\\_")
    like = f"%{escaped_q}%"
    mitre_results: List[Dict] = []

    try:
        with db.connection() as conn:
            tech_count = conn.execute('SELECT COUNT(*) AS c FROM mitre_techniques').fetchone()['c']
            if tech_count:
                tech_rows = conn.execute("""
                    SELECT t.id, t.name,
                           (SELECT tt.tactic_id FROM mitre_technique_tactics tt
                            WHERE tt.technique_id = t.id LIMIT 1) AS tactic,
                           CASE WHEN t.id = ? THEN 0 WHEN t.id LIKE ? ESCAPE '\\' THEN 1 ELSE 2 END AS rank
                    FROM mitre_techniques t
                    WHERE t.deprecated = 0
                      AND (t.id = ? OR t.id LIKE ? ESCAPE '\\' OR t.name LIKE ? ESCAPE '\\')
                    ORDER BY rank ASC, t.id ASC
                    LIMIT 3
                """, (q_upper, f'{q_upper}%', q_upper, like, like)).fetchall()
                for row in tech_rows:
                    tid = row['id']
                    mitre_results.append({
                        'kind': 'technique',
                        'id': tid,
                        'name': row['name'],
                        'tactic': row['tactic'] or '',
                        'url': f'/cti-resources/mitre-attack?technique={tid}',
                    })

                remaining = max(0, 5 - len(mitre_results))
                if remaining:
                    group_rows = conn.execute("""
                        SELECT id, name,
                               CASE WHEN id = ? THEN 0 ELSE 1 END AS rank
                        FROM mitre_groups
                        WHERE id = ? OR id LIKE ? ESCAPE '\\'
                           OR name LIKE ? ESCAPE '\\' OR aliases LIKE ? ESCAPE '\\'
                        ORDER BY rank ASC, id ASC
                        LIMIT ?
                    """, (q_upper, q_upper, like, like, like, min(remaining, 2))).fetchall()
                    for row in group_rows:
                        gid = row['id']
                        mitre_results.append({
                            'kind': 'group',
                            'id': gid,
                            'name': row['name'],
                            'url': f'/cti-resources/mitre-attack?group={gid}',
                        })
                return mitre_results
    except Exception as exc:
        logger.debug('global search MITRE DB lookup failed: %s', exc)

    try:
        parser = get_loaded_parser(get_mitre_json_path())
        if parser:
            sr = parser.search_limited(q, max_techniques=3, max_groups=2)
            for t in (sr.get('techniques') or [])[:3]:
                tid = t.get('id', '')
                mitre_results.append({
                    'kind': 'technique',
                    'id': tid,
                    'name': t.get('name', ''),
                    'tactic': t.get('tactic', ''),
                    'url': f'/cti-resources/mitre-attack?technique={tid}' if tid else '/cti-resources/mitre-attack',
                })
            for g in (sr.get('groups') or [])[:2]:
                gid = g.get('id', '')
                mitre_results.append({
                    'kind': 'group',
                    'id': gid,
                    'name': g.get('name', ''),
                    'url': f'/cti-resources/mitre-attack?group={gid}' if gid else '/cti-resources/mitre-attack',
                })
    except Exception as exc:
        logger.debug('global search MITRE parser fallback failed: %s', exc)

    return mitre_results


@app.route('/api/global-search', methods=['GET'])
@require_auth
def api_global_search():
    """Global cross-entity search: IOCs, Sources, MITRE, CTI Memory."""
    try:
        q = request.args.get('q', '').strip()
        scope = request.args.get('scope', 'quick').strip().lower()
        workspace = scope == 'workspace'
        if not q or len(q) < 2:
            return api_success({'iocs': [], 'sources': [], 'mitre': [], 'memory': [], 'query': q})

        escaped_q = q.replace("%", "\\%").replace("_", "\\_")
        like = f"%{escaped_q}%"
        ioc_limit = 10 if workspace else 6
        source_limit = 8 if workspace else 5
        memory_k = 12 if workspace else 5
        results: Dict = {}

        def _fetch_memory():
            try:
                from modules import zettelforge_bridge as zf
                if zf.is_available():
                    notes = zf.recall_query(q, k=memory_k, enrich_db=workspace)
                    from modules.cross_index import memory_note_url
                    for note in notes:
                        nid = (note.get('id') or '').strip()
                        if nid:
                            note.setdefault('links', {})['memory_url'] = memory_note_url(nid, q)
                    return notes
            except Exception as mem_err:
                logger.warning('CTI memory global search failed for query %r: %s', q, mem_err)
            return []

        from concurrent.futures import ThreadPoolExecutor
        with ThreadPoolExecutor(max_workers=1) as pool:
            memory_future = pool.submit(_fetch_memory)

            with db.connection() as conn:
                rows = conn.execute("""
                    SELECT MIN(i.id) AS id, i.ioc_value, i.ioc_type,
                           COUNT(DISTINCT i.source_id) AS source_count,
                           (SELECT s2.name FROM sources s2
                            JOIN iocs i2 ON i2.source_id = s2.id
                            WHERE i2.ioc_value = i.ioc_value AND i2.ioc_type = i.ioc_type
                              AND i2.is_deleted = 0
                            ORDER BY i2.created_at ASC LIMIT 1) AS primary_source,
                           CASE WHEN i.ioc_value = ? THEN 0 ELSE 1 END AS exact_score
                    FROM iocs i
                    WHERE i.is_deleted = 0
                      AND (i.ioc_value = ? OR i.ioc_value LIKE ? ESCAPE '\\')
                    GROUP BY i.ioc_value, i.ioc_type
                    ORDER BY exact_score ASC, source_count DESC, i.created_at DESC
                    LIMIT ?
                """, (q, q, like, ioc_limit)).fetchall()
                results['iocs'] = [dict(r) for r in rows]
                from modules.cross_index import memory_attach_url
                for item in results['iocs']:
                    item['attach_url'] = memory_attach_url(
                        kind='ioc', id=item['id'], value=item['ioc_value'], ioc_type=item['ioc_type'],
                    )

                rows = conn.execute("""
                    SELECT id, name, context, source_type,
                           (SELECT COUNT(*) FROM iocs WHERE source_id = sources.id AND is_deleted = 0) AS ioc_count,
                           CASE WHEN name = ? THEN 0 ELSE 1 END AS exact_score
                    FROM sources
                    WHERE is_deleted = 0
                      AND (name = ? OR name LIKE ? ESCAPE '\\' OR context LIKE ? ESCAPE '\\')
                    ORDER BY exact_score ASC, ioc_count DESC
                    LIMIT ?
                """, (q, q, like, like, source_limit)).fetchall()
                results['sources'] = [
                    {
                        **dict(r),
                        'url': f'/sources?highlight={r["id"]}',
                        'attach_url': memory_attach_url(
                            kind='entity',
                            entity_type='source',
                            id=r['id'],
                            value=r['name'],
                            name=r['name'],
                        ),
                    }
                    for r in rows
                ]

            results['mitre'] = _global_search_mitre(q)
            from modules.cross_index import memory_attach_url
            for m in results['mitre']:
                if m.get('kind') == 'technique':
                    m['attach_url'] = memory_attach_url(kind='ttp', id=m.get('id', ''), name=m.get('name', ''))
                elif m.get('kind') == 'group':
                    m['attach_url'] = memory_attach_url(
                        kind='entity', entity_type='intrusion_set', value=m.get('id', ''), name=m.get('name', ''),
                    )
            try:
                results['memory'] = memory_future.result(timeout=12)
            except Exception:
                results['memory'] = []

        results['query'] = q
        return api_success(results)
    except Exception as e:
        logger.error(f"api_global_search error: {e}", exc_info=True)
        return api_error(str(e), 500)

# ========== ERROR HANDLERS ==========

# ========== ERROR HANDLERS ==========

@app.errorhandler(413)
def request_entity_too_large(error):
    return api_error('File too large', 413)

@app.errorhandler(400)
def bad_request(error):
    # Silently handle HTTPS connection attempts to HTTP server
    # These are common from bots/scanners and don't need to be logged
    try:
        request_data = request.get_data()
        if request_data and len(request_data) >= 2:
            # Check if it's a TLS handshake (starts with 0x16 0x03)
            if request_data[0] == 0x16 and request_data[1] == 0x03:
                # This is a TLS/HTTPS handshake attempt, silently ignore
                return '', 400
    except Exception:
        # Ignore errors when checking request data for TLS handshakes
        pass
    # For other 400 errors, return normal error response
    return api_error('Bad request', 400)

@app.errorhandler(404)
def not_found(error):
    return render_template('404.html'), 404

@app.errorhandler(500)
def internal_error(error):
    """Handle internal server errors"""
    logger.error(f"Internal error: {error}", exc_info=True)
    return api_error('Internal server error', 500)

# ========== MAIN ==========

if __name__ == '__main__':
    # Check SSL configuration
    ssl_context = None
    if USE_SSL:
        if not (SSL_CERT_FILE.exists() and SSL_KEY_FILE.exists()):
            logger.error(f"SSL enabled but certificates not found!")
            logger.error(f"Generate with: scripts/generate-ssl-cert.sh")
            logger.error(f"Or set CTI_USE_SSL=false")
            exit(1)
        ssl_context = (str(SSL_CERT_FILE), str(SSL_KEY_FILE))
    else:
        logger.warning("SSL disabled - enable with CTI_USE_SSL=true")
    
    # Enable SO_REUSEADDR to allow immediate port reuse on restart
    # This prevents "Address already in use" errors when restarting the service
    # We patch socket.socket to always enable SO_REUSEADDR for new sockets
    original_socket = socket.socket
    
    class SocketWithReuseAddr(original_socket):
        """Socket class that automatically enables SO_REUSEADDR"""
        def __init__(self, family=socket.AF_INET, type=socket.SOCK_STREAM, proto=0, fileno=None):
            super().__init__(family, type, proto, fileno)
            # Enable SO_REUSEADDR for all new sockets
            try:
                self.setsockopt(socket.SOL_SOCKET, socket.SO_REUSEADDR, 1)
            except (OSError, AttributeError):
                # If setsockopt fails (e.g., socket already closed), continue anyway
                pass
    
    # Temporarily replace socket.socket with our version
    socket.socket = SocketWithReuseAddr
    
    try:
        # Use run_simple which will now create sockets with SO_REUSEADDR enabled
        run_simple(
            hostname=HOST,
            port=PORT,
            application=app,
            use_reloader=False,  # Disable reloader in production
            use_debugger=DEBUG,
            ssl_context=ssl_context,
            threaded=True
        )
    finally:
        # Restore original socket class
        socket.socket = original_socket
