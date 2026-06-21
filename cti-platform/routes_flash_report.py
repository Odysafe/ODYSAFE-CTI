"""
Flash Report (FLINT) Routes - ODYSAFE CTI Platform
Flash INTelligence reporting system for rapid CTI response - Enhanced Version
"""

import logging
import json
import re
import io
from flask import Blueprint, render_template, request, jsonify, send_file
from datetime import datetime, timezone
import tempfile
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins
from openpyxl.drawing.image import Image as XLImage
from collections import Counter
from openpyxl.chart import PieChart, BarChart, Reference
from openpyxl.chart.label import DataLabelList

logger = logging.getLogger(__name__)

# Create Blueprint
flash_report_bp = Blueprint('flash_report', __name__, url_prefix='/flash-report')


def _write_section_header(ws, row, title, styles, cols=5):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = styles['header_font']
    cell.fill = styles['header_fill']
    cell.border = styles['thin_border']
    return row + 1


def _write_field_row(ws, row, label, value, styles, cols=5):
    ws.cell(row=row, column=1, value=label).font = styles['bold_font']
    ws.cell(row=row, column=1).border = styles['thin_border']
    value_cell = ws.cell(row=row, column=2, value=value or 'N/A')
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=cols)
    value_cell.border = styles['thin_border']
    value_cell.alignment = styles['wrap_align']
    return row + 1


def _write_table_headers(ws, row, headers, styles):
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles['header_font']
        cell.fill = styles['header_fill']
        cell.border = styles['thin_border']
        cell.alignment = styles['center_align']
    return row + 1


def get_tlp_color(tlp):
    """Get Excel fill color for TLP level"""
    colors = {
        'RED': 'DC3545',
        'AMBER': 'FFC107',
        'AMBER+STRICT': 'FF8C00',
        'GREEN': '28A745',
        'CLEAR': '17A2B8',
        'WHITE': '6C757D'
    }
    return colors.get(tlp, 'FFC107')


def get_priority_color(priority):
    """Get Excel fill color for priority"""
    colors = {
        'Critical': '7F1D1D',
        'High': 'DC2626',
        'Medium': 'F59E0B',
        'Low': '6B7280'
    }
    return colors.get(priority, '6B7280')


def defang_ioc(value, ioc_type):
    """Defang IOC value for safe sharing"""
    if not value:
        return value
    
    if ioc_type in ['URL', 'Domain']:
        # Replace http with hxxp and . with [.]
        value = re.sub(r'http', 'hxxp', value, flags=re.IGNORECASE)
        value = value.replace('.', '[.]')
    elif ioc_type == 'IP':
        # Replace . with [.]
        value = value.replace('.', '[.]')
    elif ioc_type in ['Email', 'Email Address']:
        # Replace @ with [@]
        value = value.replace('@', '[@]')
    
    return value


@flash_report_bp.route('/', methods=['GET'])
def flash_report_home():
    """Flash Report (FLINT) home page"""
    return render_template('flash_report/index.html', datetime=datetime)


@flash_report_bp.route('/export', methods=['POST'])
def export_flash_report():
    """Export Flash Report to Excel with all blocks"""
    try:
        # Parse JSON data
        data = request.get_json() if request.is_json else request.form.to_dict()
        
        if not isinstance(data, dict):
            return jsonify({'error': 'Invalid data format'}), 400
        
        # Create workbook with multiple sheets
        wb = Workbook()
        wb.remove(wb.active)

        # Dashboard with visuals (diamond, charts, KPIs)
        ws_dashboard = wb.create_sheet("Dashboard", 0)
        _create_dashboard_sheet(ws_dashboard, data)

        ws_summary = wb.create_sheet("Executive Summary", 1)
        _create_summary_sheet(ws_summary, data)

        ws_tech = wb.create_sheet("Technical Analysis", 2)
        _create_technical_sheet(ws_tech, data)

        ws_iocs = wb.create_sheet("IOCs", 3)
        _create_iocs_sheet(ws_iocs, data)

        ws_detect = wb.create_sheet("Detection", 4)
        _create_detection_sheet(ws_detect, data)

        ws_rec = wb.create_sheet("Recommendations", 5)
        _create_recommendations_sheet(ws_rec, data)

        ws_sources = wb.create_sheet("Sources", 6)
        _create_sources_sheet(ws_sources, data)
        
        # Save to temp file
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
            tmp_path = tmp.name
            wb.save(tmp_path)
        
        ref = data.get('reference', 'FLINT-UNKNOWN')
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"{ref}_{timestamp}.xlsx"
        
        logger.info(f"Generated Flash Report Excel: {filename}")

        try:
            from modules import zettelforge_bridge as zf
            zf.remember_flash_report(data)
        except Exception as mem_err:
            logger.debug('CTI memory flash report skipped: %s', mem_err)
        
        return send_file(
            tmp_path,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        logger.error(f"Error exporting Flash Report: {str(e)}", exc_info=True)
        return jsonify({'error': f'Export failed: {str(e)}'}), 500


def _diamond_content_text(content, fallback='Not specified'):
    text = (content or fallback).strip()
    if len(text) > 320:
        text = text[:317] + '...'
    return text


def _style_box_header(cell, title, fill_hex, styles):
    cell.value = title
    cell.fill = PatternFill(start_color=fill_hex, end_color=fill_hex, fill_type='solid')
    cell.font = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = styles['medium_border']


def _style_box_body(cell, content, fill_hex, styles):
    cell.value = content
    cell.fill = PatternFill(start_color=fill_hex, end_color=fill_hex, fill_type='solid')
    cell.font = Font(name='Calibri', size=10, color='1E1033')
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = styles['medium_border']


def _style_connector_cell(cell, symbol, styles):
    cell.value = symbol
    cell.fill = PatternFill(start_color='F5F3FF', end_color='F5F3FF', fill_type='solid')
    cell.font = Font(name='Calibri', size=13, bold=True, color='7C3AED')
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = styles['thin_border']


def _style_diamond_cell(cell, fill_hex, styles, title_size=10):
    cell.fill = PatternFill(start_color=fill_hex, end_color=fill_hex, fill_type='solid')
    cell.font = Font(name='Calibri', size=title_size, bold=True, color='1E1033')
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = styles['medium_border']


def _load_diagram_fonts():
    from PIL import ImageFont
    candidates = [
        ('/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf', 13),
        ('/usr/share/fonts/truetype/liberation/LiberationSans-Bold.ttf', 13),
        ('/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf', 11),
        ('/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf', 11),
    ]
    loaded = {}
    for path, size in candidates:
        key = 'title' if 'Bold' in path else 'body'
        if key in loaded:
            continue
        try:
            loaded[key] = ImageFont.truetype(path, size)
        except OSError:
            continue
    default = ImageFont.load_default()
    loaded.setdefault('title', default)
    loaded.setdefault('body', default)
    return loaded


def _wrap_diagram_text(draw, text, font, max_width, max_lines=5):
    words = (text or 'Not specified').split()
    lines = []
    current = []
    for word in words:
        trial = ' '.join(current + [word])
        width = draw.textlength(trial, font=font)
        if width <= max_width:
            current.append(word)
        else:
            if current:
                lines.append(' '.join(current))
            current = [word]
        if len(lines) >= max_lines:
            break
    if current and len(lines) < max_lines:
        lines.append(' '.join(current))
    if len(lines) == max_lines and len(' '.join(words)) > len(' '.join(lines)):
        lines[-1] = lines[-1][: max(0, len(lines[-1]) - 3)] + '...'
    return lines or ['Not specified']


def _draw_labeled_card(draw, box, title, content, colors, fonts):
    x0, y0, x1, y1 = box
    header_h = 30
    draw.rounded_rectangle(box, radius=10, fill=colors['body'], outline=colors['border'], width=2)
    draw.rectangle([x0, y0, x1, y0 + header_h], fill=colors['header'])
    draw.line([x0, y0 + header_h, x1, y0 + header_h], fill=colors['border'], width=1)
    draw.text(((x0 + x1) / 2, y0 + header_h / 2), title, fill='white', font=fonts['title'], anchor='mm')
    lines = _wrap_diagram_text(draw, _diamond_content_text(content), fonts['body'], x1 - x0 - 16)
    line_h = 15
    start_y = y0 + header_h + 12
    for i, line in enumerate(lines):
        draw.text((x0 + 10, start_y + i * line_h), line, fill=colors['text'], font=fonts['body'])


def _build_diamond_diagram_image(data):
    """Render Diamond Model as a PNG diagram for Excel embedding."""
    from PIL import Image, ImageDraw

    width, height = 980, 640
    img = Image.new('RGB', (width, height), '#FAFAFC')
    draw = ImageDraw.Draw(img)
    fonts = _load_diagram_fonts()

    cx, cy = width // 2, height // 2
    apex = (cx, 36)
    right_pt = (width - 36, cy)
    bottom = (cx, height - 36)
    left_pt = (36, cy)
    diamond = [apex, right_pt, bottom, left_pt]

    draw.polygon(diamond, fill='#F5F3FF', outline='#7C3AED', width=4)
    draw.line([apex, bottom], fill='#A78BFA', width=2)
    draw.line([left_pt, right_pt], fill='#A78BFA', width=2)

    event_w, event_h = 150, 96
    event_box = (cx - event_w // 2, cy - event_h // 2, cx + event_w // 2, cy + event_h // 2)
    draw.rounded_rectangle(event_box, radius=14, fill='#4C1D95', outline='#EDE9FE', width=3)
    draw.text((cx, cy - 10), 'EVENT', fill='white', font=fonts['title'], anchor='mm')
    draw.text((cx, cy + 14), 'Core incident', fill='#DDD6FE', font=fonts['body'], anchor='mm')

    card_w, card_h = 250, 118
    cards = [
        ('ADVERSARY', data.get('diamond_adversary'), (cx - card_w // 2, 52, cx + card_w // 2, 52 + card_h),
         {'header': '#5B21B6', 'body': '#EDE9FE', 'border': '#6D28D9', 'text': '#1E1033'}),
        ('CAPABILITY', data.get('diamond_capability'), (52, cy - card_h // 2, 52 + card_w, cy + card_h // 2),
         {'header': '#6D28D9', 'body': '#EDE9FE', 'border': '#7C3AED', 'text': '#1E1033'}),
        ('INFRASTRUCTURE', data.get('diamond_infrastructure'),
         (width - 52 - card_w, cy - card_h // 2, width - 52, cy + card_h // 2),
         {'header': '#7C3AED', 'body': '#DDD6FE', 'border': '#8B5CF6', 'text': '#1E1033'}),
        ('VICTIM', data.get('diamond_victim'),
         (cx - card_w // 2, height - 52 - card_h, cx + card_w // 2, height - 52),
         {'header': '#6D28D9', 'body': '#C4B5FD', 'border': '#7C3AED', 'text': '#1E1033'}),
    ]
    for title, content, box, colors in cards:
        _draw_labeled_card(draw, box, title, content, colors, fonts)

    draw.text((cx, height - 14), 'Diamond Model of Intrusion Analysis', fill='#6B7280', font=fonts['body'], anchor='mm')

    buf = io.BytesIO()
    img.save(buf, format='PNG')
    buf.seek(0)
    return buf


def _write_quadrant_box(ws, header_row, body_row, col_start, col_end, title, content,
                        header_color, body_color, styles, body_height=56):
    ws.merge_cells(start_row=header_row, start_column=col_start, end_row=header_row, end_column=col_end)
    ws.merge_cells(start_row=body_row, start_column=col_start, end_row=body_row, end_column=col_end)
    _style_box_header(ws.cell(row=header_row, column=col_start), title, header_color, styles)
    _style_box_body(
        ws.cell(row=body_row, column=col_start),
        _diamond_content_text(content),
        body_color,
        styles,
    )
    ws.row_dimensions[header_row].height = 22
    ws.row_dimensions[body_row].height = body_height


def _write_diamond_model_visual(ws, start_row, data, styles, col_start=2, col_end=9):
    """Embed Diamond Model as a PNG diagram (fallback: structured cell layout)."""
    row = _write_section_header(ws, start_row, "DIAMOND MODEL OF INTRUSION ANALYSIS", styles, cols=col_end)

    anchor_col = get_column_letter(col_start)
    try:
        buf = _build_diamond_diagram_image(data)
        xl_img = XLImage(buf)
        xl_img.width = 620
        xl_img.height = 405
        ws.add_image(xl_img, f'{anchor_col}{row}')
        for r in range(row, row + 22):
            ws.row_dimensions[r].height = 19
        return row + 24
    except Exception as exc:
        logger.warning(f"Diamond PNG render failed, using cell layout: {exc}")
        return _write_diamond_model_cell_layout(ws, row, data, styles, col_start, col_end)


def _write_diamond_model_cell_layout(ws, start_row, data, styles, col_start=2, col_end=9):
    """Fallback diamond layout using merged cells when Pillow is unavailable."""
    row = start_row
    top_header = row
    top_body = row + 1
    conn_top = row + 2
    mid_header = row + 3
    mid_body = row + 4
    conn_bottom = row + 5
    bottom_header = row + 6
    bottom_body = row + 7
    frame_end = row + 8

    left_mid_start = col_start
    left_mid_end = col_start + 2
    center_start = col_start + 3
    center_end = col_start + 4
    right_mid_start = col_start + 5
    right_mid_end = col_end

    top_start = col_start + 2
    top_end = col_end - 1

    palette = {
        'adversary': ('5B21B6', 'EDE9FE'),
        'capability': ('6D28D9', 'DDD6FE'),
        'event': ('4C1D95', '8B5CF6'),
        'infrastructure': ('7C3AED', 'C4B5FD'),
        'victim': ('6D28D9', 'A78BFA'),
    }

    _write_quadrant_box(
        ws, top_header, top_body, top_start, top_end,
        'ADVERSARY', data.get('diamond_adversary'),
        palette['adversary'][0], palette['adversary'][1], styles, body_height=48,
    )

    connectors_top = [
        (conn_top, col_start, '\\'),
        (conn_top, col_start + 1, '\\'),
        (conn_top, col_end - 1, '/'),
        (conn_top, col_end, '/'),
    ]
    for r, c, sym in connectors_top:
        _style_connector_cell(ws.cell(row=r, column=c), sym, styles)
    ws.row_dimensions[conn_top].height = 18

    _write_quadrant_box(
        ws, mid_header, mid_body, left_mid_start, left_mid_end,
        'CAPABILITY', data.get('diamond_capability'),
        palette['capability'][0], palette['capability'][1], styles, body_height=64,
    )

    ws.merge_cells(start_row=mid_header, start_column=center_start, end_row=mid_header, end_column=center_end)
    ws.merge_cells(start_row=mid_body, start_column=center_start, end_row=mid_body, end_column=center_end)
    _style_box_header(ws.cell(row=mid_header, column=center_start), 'EVENT', palette['event'][0], styles)
    event_body = ws.cell(row=mid_body, column=center_start, value='◆\nCore incident')
    event_body.fill = PatternFill(start_color=palette['event'][1], end_color=palette['event'][1], fill_type='solid')
    event_body.font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
    event_body.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    event_body.border = styles['medium_border']

    _write_quadrant_box(
        ws, mid_header, mid_body, right_mid_start, right_mid_end,
        'INFRASTRUCTURE', data.get('diamond_infrastructure'),
        palette['infrastructure'][0], palette['infrastructure'][1], styles, body_height=64,
    )

    connectors_bottom = [
        (conn_bottom, col_start, '/'),
        (conn_bottom, col_start + 1, '/'),
        (conn_bottom, col_end - 1, '\\'),
        (conn_bottom, col_end, '\\'),
    ]
    for r, c, sym in connectors_bottom:
        _style_connector_cell(ws.cell(row=r, column=c), sym, styles)
    ws.row_dimensions[conn_bottom].height = 18

    _write_quadrant_box(
        ws, bottom_header, bottom_body, top_start, top_end,
        'VICTIM', data.get('diamond_victim'),
        palette['victim'][0], palette['victim'][1], styles, body_height=48,
    )

    for c in range(col_start, col_end + 1):
        letter = get_column_letter(c)
        ws.column_dimensions[letter].width = 14 if c in (center_start, center_end) else 16

    for r in range(top_header, frame_end + 1):
        for c in range(col_start, col_end + 1):
            cell = ws.cell(row=r, column=c)
            if r == top_header and c == col_start:
                cell.border = Border(
                    left=Side(style='medium', color='7C3AED'),
                    top=Side(style='medium', color='7C3AED'),
                )
            elif r == top_header and c == col_end:
                cell.border = Border(
                    right=Side(style='medium', color='7C3AED'),
                    top=Side(style='medium', color='7C3AED'),
                )
            elif r == frame_end and c == col_start:
                cell.border = Border(
                    left=Side(style='medium', color='7C3AED'),
                    bottom=Side(style='medium', color='7C3AED'),
                )
            elif r == frame_end and c == col_end:
                cell.border = Border(
                    right=Side(style='medium', color='7C3AED'),
                    bottom=Side(style='medium', color='7C3AED'),
                )

    legend_row = frame_end + 2
    ws.merge_cells(start_row=legend_row, start_column=col_start, end_row=legend_row, end_column=col_end)
    legend = ws.cell(
        row=legend_row,
        column=col_start,
        value='Diamond Model mapping: Adversary (top), Capability (left), Infrastructure (right), Victim (bottom), Event (center).',
    )
    legend.font = Font(name='Calibri', size=9, italic=True, color='6B7280')
    legend.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    legend.fill = PatternFill(start_color='FAFAFA', end_color='FAFAFA', fill_type='solid')

    return legend_row + 2


def _write_cia_triad_visual(ws, row, data, styles):
    row = _write_section_header(ws, row, "IMPACT ASSESSMENT (CIA TRIAD)", styles, cols=7)

    impacts = [
        ('CONFIDENTIALITY', data.get('impact_confidentiality', 'None'), 'FECACA'),
        ('INTEGRITY', data.get('impact_integrity', 'None'), 'FED7AA'),
        ('AVAILABILITY', data.get('impact_availability', 'None'), 'BBF7D0'),
    ]
    col = 2
    for label, value, color in impacts:
        ws.merge_cells(start_row=row, start_column=col, end_row=row + 2, end_column=col + 1)
        cell = ws.cell(row=row, column=col, value=f"{label}\n{value or 'None'}")
        _style_diamond_cell(cell, color, styles, title_size=10)
        col += 2

    return row + 4


def _apply_tlp_fill(cell, tlp, styles):
    color = get_tlp_color(tlp or 'AMBER')
    cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
    cell.font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    cell.alignment = styles['center_align']
    cell.border = styles['thin_border']


def _write_hidden_chart_table(ws, start_row, start_col, headers, rows):
    for i, h in enumerate(headers):
        ws.cell(row=start_row, column=start_col + i, value=h)
    r = start_row + 1
    for row_data in rows:
        for i, val in enumerate(row_data):
            ws.cell(row=r, column=start_col + i, value=val)
        r += 1
    return start_row, r - 1


def _add_pie_chart(ws, anchor, title, cat_col, data_col, min_row, max_row, header_row):
    chart = PieChart()
    chart.title = title
    chart.style = 10
    labels = Reference(ws, min_col=cat_col, min_row=min_row, max_row=max_row)
    data = Reference(ws, min_col=data_col, min_row=header_row, max_row=max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(labels)
    chart.dataLabels = DataLabelList()
    chart.dataLabels.showPercent = True
    chart.dataLabels.showCatName = False
    chart.width = 14
    chart.height = 10
    ws.add_chart(chart, anchor)


def _add_bar_chart(ws, anchor, title, cat_col, data_col, min_row, max_row, header_row):
    chart = BarChart()
    chart.type = 'col'
    chart.title = title
    chart.style = 10
    labels = Reference(ws, min_col=cat_col, min_row=min_row, max_row=max_row)
    data = Reference(ws, min_col=data_col, min_row=header_row, max_row=max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(labels)
    chart.width = 14
    chart.height = 10
    ws.add_chart(chart, anchor)


def _create_dashboard_sheet(ws, data):
    """Overview sheet with KPI cards, diamond model, and charts."""
    styles = _get_excel_styles()
    row = 1

    ws.merge_cells('A1:H1')
    title = ws['A1']
    title.value = f"FLASH REPORT · {data.get('reference', 'N/A')}"
    title.font = styles['title_font']
    title.fill = styles['title_fill']
    title.alignment = styles['center_align']
    title.border = styles['thin_border']
    row = 3

    ws.merge_cells('A3:H3')
    sub = ws['A3']
    sub.value = data.get('subject') or 'No subject'
    sub.font = Font(name='Calibri', size=13, bold=True, color='4C1D95')
    sub.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    row = 5

    kpis = [
        ('TLP', data.get('tlp', 'AMBER'), get_tlp_color(data.get('tlp', 'AMBER'))),
        ('PRIORITY', data.get('priority', 'Medium'), get_priority_color(data.get('priority', 'Medium'))),
        ('STATUS', data.get('status', 'Draft'), '6D28D9'),
        ('IOCs', str(len(data.get('iocs', []))), '5B21B6'),
        ('TTPs', str(len(data.get('ttps', []))), '7C3AED'),
        ('ACTIONS', str(len(data.get('recommendations', []))), '8B5CF6'),
    ]
    col = 1
    for label, value, color in kpis:
        ws.merge_cells(start_row=row, start_column=col, end_row=row + 1, end_column=col)
        cell = ws.cell(row=row, column=col, value=f"{label}\n{value}")
        cell.border = styles['thin_border']
        cell.alignment = styles['center_align']
        if label == 'TLP':
            _apply_tlp_fill(cell, value, styles)
        else:
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
            cell.font = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
        col += 1

    row = 8
    row = _write_diamond_model_visual(ws, row, data, styles, col_start=1, col_end=8)

    meta_row = row
    ws.merge_cells(start_row=meta_row, start_column=1, end_row=meta_row, end_column=4)
    meta_left = ws.cell(row=meta_row, column=1, value=(
        f"Author: {data.get('author', 'N/A')}\n"
        f"Incident type: {data.get('incident_type', 'N/A')}\n"
        f"Threat actor: {data.get('threat_actor') or 'Unknown'}"
    ))
    meta_left.alignment = styles['wrap_align']
    meta_left.border = styles['thin_border']
    meta_left.font = styles['normal_font']

    ws.merge_cells(start_row=meta_row, start_column=5, end_row=meta_row, end_column=8)
    meta_right = ws.cell(row=meta_row, column=5, value=(
        f"Confidence: {data.get('incident_confidence', data.get('overall_confidence', 'N/A'))}\n"
        f"Created: {data.get('created_at', 'N/A')}\n"
        f"CVEs: {data.get('cves') or 'None'}"
    ))
    meta_right.alignment = styles['wrap_align']
    meta_right.border = styles['thin_border']
    meta_right.font = styles['normal_font']
    ws.row_dimensions[meta_row].height = 48
    row += 2

    chart_base_row = 55
    iocs = data.get('iocs', [])
    if iocs:
        type_counts = Counter(ioc.get('type', 'Unknown') for ioc in iocs)
        _, max_r = _write_hidden_chart_table(
            ws, chart_base_row, 10,
            ['Type', 'Count'],
            [[t, c] for t, c in type_counts.most_common()]
        )
        _add_pie_chart(ws, 'A32', 'IOC Types', 10, 11, chart_base_row + 1, max_r, chart_base_row)

    recs = data.get('recommendations', [])
    if recs:
        pri_counts = Counter(rec.get('priority', 'Medium') for rec in recs)
        base = chart_base_row + 12
        _, max_r = _write_hidden_chart_table(
            ws, base, 10,
            ['Priority', 'Count'],
            [[p, c] for p, c in pri_counts.most_common()]
        )
        _add_bar_chart(ws, 'E32', 'Actions by Priority', 10, 11, base + 1, max_r, base)

    ttps = data.get('ttps', [])
    if ttps:
        tactic_counts = Counter(ttp.get('tactic') or 'Unknown' for ttp in ttps)
        base = chart_base_row + 24
        _, max_r = _write_hidden_chart_table(
            ws, base, 10,
            ['Tactic', 'Count'],
            [[t, c] for t, c in tactic_counts.most_common()]
        )
        _add_bar_chart(ws, 'A48', 'TTPs by Tactic', 10, 11, base + 1, max_r, base)

    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 14
    ws.column_dimensions['G'].width = 14
    ws.column_dimensions['H'].width = 14
    ws.sheet_view.showGridLines = False


def _create_summary_sheet(ws, data):
    """Create Executive Summary sheet"""
    styles = _get_excel_styles()
    row = 1
    
    # Title
    ws.merge_cells(f'A{row}:E{row}')
    ws[f'A{row}'] = "FLASH REPORT - EXECUTIVE SUMMARY"
    ws[f'A{row}'].font = styles['title_font']
    ws[f'A{row}'].fill = styles['title_fill']
    ws[f'A{row}'].alignment = styles['center_align']
    ws[f'A{row}'].border = styles['thin_border']
    row += 2
    
    # Metadata section
    ws[f'A{row}'] = "METADATA"
    ws[f'A{row}'].font = styles['header_font']
    ws[f'A{row}'].fill = styles['header_fill']
    ws[f'A{row}'].border = styles['thin_border']
    ws.merge_cells(f'A{row}:B{row}')
    row += 1
    
    metadata_fields = [
        ('Reference', data.get('reference', 'N/A')),
        ('TLP Classification', data.get('tlp', 'AMBER')),
        ('PAP Classification', data.get('pap', 'AMBER')),
        ('Priority', data.get('priority', 'Medium')),
        ('Status', data.get('status', 'Draft')),
        ('Author', data.get('author', 'N/A')),
        ('Created', data.get('created_at', 'N/A')),
        ('Updated', data.get('updated_at', 'N/A')),
        ('Primary CVEs', data.get('cves', 'N/A')),
    ]
    
    for field, value in metadata_fields:
        ws[f'A{row}'] = field
        ws[f'B{row}'] = value
        ws[f'A{row}'].font = styles['bold_font']
        ws[f'A{row}'].border = styles['thin_border']
        ws[f'B{row}'].border = styles['thin_border']
        if field == 'TLP Classification':
            _apply_tlp_fill(ws[f'B{row}'], value, styles)
        row += 1
    
    row += 1
    
    # Subject & Classification
    ws[f'A{row}'] = "SUBJECT & CLASSIFICATION"
    ws[f'A{row}'].font = styles['header_font']
    ws[f'A{row}'].fill = styles['header_fill']
    ws[f'A{row}'].border = styles['thin_border']
    ws.merge_cells(f'A{row}:B{row}')
    row += 1
    
    ws[f'A{row}'] = "Subject"
    ws[f'B{row}'] = data.get('subject', 'N/A')
    ws[f'A{row}'].font = styles['bold_font']
    ws.merge_cells(f'B{row}:E{row}')
    ws[f'A{row}'].border = styles['thin_border']
    ws[f'B{row}'].border = styles['thin_border']
    ws[f'B{row}'].alignment = styles['wrap_align']
    row += 1
    
    ws[f'A{row}'] = "Incident Type"
    ws[f'B{row}'] = data.get('incident_type', 'N/A')
    ws[f'A{row}'].font = styles['bold_font']
    ws[f'A{row}'].border = styles['thin_border']
    ws[f'B{row}'].border = styles['thin_border']
    row += 1
    
    ws[f'A{row}'] = "Threat Actor"
    ws[f'B{row}'] = data.get('threat_actor', 'Unknown')
    ws[f'A{row}'].font = styles['bold_font']
    ws[f'A{row}'].border = styles['thin_border']
    ws[f'B{row}'].border = styles['thin_border']
    row += 1
    
    ws[f'A{row}'] = "Confidence Level (Incident)"
    ws[f'B{row}'] = data.get('incident_confidence', data.get('confidence', 'Unknown'))
    ws[f'A{row}'].font = styles['bold_font']
    ws[f'A{row}'].border = styles['thin_border']
    ws[f'B{row}'].border = styles['thin_border']
    row += 2
    
    # Executive Summary
    ws[f'A{row}'] = "EXECUTIVE SUMMARY"
    ws[f'A{row}'].font = styles['header_font']
    ws[f'A{row}'].fill = styles['header_fill']
    ws[f'A{row}'].border = styles['thin_border']
    ws.merge_cells(f'A{row}:E{row}')
    row += 1
    
    summary_parts = [
        ('Who is Affected', 'affected_entities'),
        ('What is the Threat', 'threat_nature'),
        ('Urgency & Action', 'urgency_action'),
    ]
    
    for label, key in summary_parts:
        ws[f'A{row}'] = label
        ws[f'B{row}'] = data.get(key, 'N/A')
        ws[f'A{row}'].font = styles['bold_font']
        ws.merge_cells(f'B{row}:E{row}')
        ws[f'A{row}'].border = styles['thin_border']
        ws[f'B{row}'].border = styles['thin_border']
        ws[f'B{row}'].alignment = styles['wrap_align']
        row += 1
    
    # Impact assessment (visual CIA triad)
    row += 1
    row = _write_cia_triad_visual(ws, row, data, styles)
    
    # Key Takeaways
    row += 1
    ws[f'A{row}'] = "KEY TAKEAWAYS"
    ws[f'A{row}'].font = styles['header_font']
    ws[f'A{row}'].fill = styles['header_fill']
    ws[f'A{row}'].border = styles['thin_border']
    ws.merge_cells(f'A{row}:E{row}')
    row += 1
    
    takeaways = data.get('takeaways', [])
    if takeaways:
        for i, takeaway in enumerate(takeaways, 1):
            ws[f'A{row}'] = f"{i}."
            ws[f'B{row}'] = takeaway
            ws.merge_cells(f'B{row}:E{row}')
            ws[f'A{row}'].border = styles['thin_border']
            ws[f'B{row}'].border = styles['thin_border']
            ws[f'B{row}'].alignment = styles['wrap_align']
            row += 1
    else:
        ws[f'A{row}'] = "No takeaways provided"
        ws.merge_cells(f'A{row}:E{row}')
        row += 1

    row += 1
    row = _write_section_header(ws, row, "ASSESSMENT & CONFIDENCE", styles)
    assessment_fields = [
        ('Overall Probability', data.get('overall_confidence', data.get('confidence', 'N/A'))),
        ('Source Reliability (Admiralty)', data.get('source_reliability', 'N/A')),
        ('Analysis Biases', data.get('biases', 'N/A')),
        ('Confidence by Section', data.get('section_confidence', 'N/A')),
    ]
    for label, value in assessment_fields:
        row = _write_field_row(ws, row, label, value, styles)
    
    # Set column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15


def _create_technical_sheet(ws, data):
    """Create Technical Analysis sheet"""
    styles = _get_excel_styles()
    row = 1
    
    # Title
    ws.merge_cells(f'A{row}:E{row}')
    ws[f'A{row}'] = "TECHNICAL ANALYSIS"
    ws[f'A{row}'].font = styles['title_font']
    ws[f'A{row}'].fill = styles['title_fill']
    ws[f'A{row}'].alignment = styles['center_align']
    ws[f'A{row}'].border = styles['thin_border']
    row += 2
    
    row = _write_diamond_model_visual(ws, row, data, styles, col_start=1, col_end=5)
    
    # Technical Details
    row += 1
    ws[f'A{row}'] = "TECHNICAL DETAILS"
    ws[f'A{row}'].font = styles['header_font']
    ws[f'A{row}'].fill = styles['header_fill']
    ws[f'A{row}'].border = styles['thin_border']
    ws.merge_cells(f'A{row}:E{row}')
    row += 1
    
    tech_fields = [
        ('Attack Vector', 'attack_vector'),
        ('Malware/Family', 'malware_family'),
        ('Persistence', 'persistence'),
        ('C2 Infrastructure', 'c2_infrastructure'),
    ]
    
    for label, key in tech_fields:
        ws[f'A{row}'] = label
        ws[f'B{row}'] = data.get(key, 'N/A')
        ws[f'A{row}'].font = styles['bold_font']
        ws.merge_cells(f'B{row}:E{row}')
        ws[f'A{row}'].border = styles['thin_border']
        ws[f'B{row}'].border = styles['thin_border']
        ws[f'B{row}'].alignment = styles['wrap_align']
        row += 1

    row += 1
    row = _write_section_header(ws, row, "CONTEXT & TIMELINE", styles, cols=5)
    timeline_context = [
        ('Threat History', data.get('threat_history')),
        ('Affected Scope', data.get('affected_scope')),
        ('First Observed ITW', data.get('first_observed')),
        ('Patch Available Date', data.get('patch_date')),
    ]
    for label, value in timeline_context:
        row = _write_field_row(ws, row, label, value, styles)

    row += 1
    row = _write_section_header(ws, row, "TIMELINE EVENTS", styles, cols=5)
    row = _write_table_headers(ws, row, ['Date/Time UTC', 'Event', 'Impact', 'Source'], styles)
    timeline = data.get('timeline', [])
    if timeline:
        for event in timeline:
            ws.cell(row=row, column=1, value=event.get('date', '')).border = styles['thin_border']
            ws.cell(row=row, column=2, value=event.get('event', '')).border = styles['thin_border']
            ws.cell(row=row, column=3, value=event.get('impact', '')).border = styles['thin_border']
            ws.cell(row=row, column=4, value=event.get('source', '')).border = styles['thin_border']
            row += 1
    else:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        ws.cell(row=row, column=1, value='No timeline events provided').border = styles['thin_border']
        row += 1

    ttps = data.get('ttps', [])
    if ttps:
        row += 1
        row = _write_section_header(ws, row, "MITRE ATT&CK TTPs", styles, cols=5)
        row = _write_table_headers(ws, row, ['Technique ID', 'Name', 'Tactic'], styles)
        for ttp in ttps:
            ws.cell(row=row, column=1, value=ttp.get('id', '')).border = styles['thin_border']
            ws.cell(row=row, column=2, value=ttp.get('name', '')).border = styles['thin_border']
            ws.cell(row=row, column=3, value=ttp.get('tactic', '')).border = styles['thin_border']
            row += 1

    row += 1
    row = _write_section_header(ws, row, "INTELLIGENCE GAPS", styles, cols=5)
    gaps_fields = [
        ('Known Unknowns', data.get('known_unknowns')),
        ('Open Hunt Questions', data.get('hunt_questions')),
        ('RFIs Generated', data.get('rfis')),
    ]
    for label, value in gaps_fields:
        row = _write_field_row(ws, row, label, value, styles)
    
    # Set column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 60


def _create_iocs_sheet(ws, data):
    """Create IOCs sheet with enriched data"""
    styles = _get_excel_styles()
    row = 1
    
    # Title
    ws.merge_cells(f'A{row}:H{row}')
    ws[f'A{row}'] = "INDICATORS OF COMPROMISE"
    ws[f'A{row}'].font = styles['title_font']
    ws[f'A{row}'].fill = styles['title_fill']
    ws[f'A{row}'].alignment = styles['center_align']
    ws[f'A{row}'].border = styles['thin_border']
    row += 2
    
    # Headers
    headers = ['Type', 'Value', 'Defanged Value', 'TLP', 'Confidence', 'First Seen', 'Last Seen', 'Active']
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles['header_font']
        cell.fill = styles['header_fill']
        cell.border = styles['thin_border']
        cell.alignment = styles['center_align']
    row += 1
    
    # IOC Data
    iocs = data.get('iocs', [])
    defang_enabled = data.get('defang_export', True)
    
    if not iocs:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        ws.cell(row=row, column=1, value='No IOCs provided').border = styles['thin_border']
        row += 1
    else:
        for ioc in iocs:
            ioc_type = ioc.get('type', 'Unknown')
            value = ioc.get('value', '')
            
            ws.cell(row=row, column=1, value=ioc_type).border = styles['thin_border']
            ws.cell(row=row, column=2, value=value).border = styles['thin_border']
            
            defanged = defang_ioc(value, ioc_type) if defang_enabled else value
            ws.cell(row=row, column=3, value=defanged).border = styles['thin_border']
            
            ws.cell(row=row, column=4, value=ioc.get('tlp', 'AMBER')).border = styles['thin_border']
            
            confidence = ioc.get('confidence', 'Medium')
            cell_conf = ws.cell(row=row, column=5, value=confidence)
            cell_conf.border = styles['thin_border']
            if confidence.lower() == 'high':
                cell_conf.fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
            elif confidence.lower() == 'medium':
                cell_conf.fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
            else:
                cell_conf.fill = PatternFill(start_color='FFB6C1', end_color='FFB6C1', fill_type='solid')
            
            ws.cell(row=row, column=6, value=ioc.get('first_seen', '')).border = styles['thin_border']
            ws.cell(row=row, column=7, value=ioc.get('last_seen', '')).border = styles['thin_border']
            ws.cell(row=row, column=8, value='Yes' if ioc.get('active', True) else 'No').border = styles['thin_border']
            
            row += 1

    if len(iocs) > 1:
        chart_row = max(row + 2, 20)
        type_counts = Counter(ioc.get('type', 'Unknown') for ioc in iocs)
        _, max_r = _write_hidden_chart_table(
            ws, chart_row, 10,
            ['Type', 'Count'],
            [[t, c] for t, c in type_counts.most_common()]
        )
        _add_pie_chart(ws, 'J2', 'IOC Type Breakdown', 10, 11, chart_row + 1, max_r, chart_row)
    
    # Set column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 10


def _create_detection_sheet(ws, data):
    """Create Detection Rules sheet"""
    styles = _get_excel_styles()
    row = 1
    
    # Title
    ws.merge_cells(f'A{row}:C{row}')
    ws[f'A{row}'] = "DETECTION RULES & HUNTING QUERIES"
    ws[f'A{row}'].font = styles['title_font']
    ws[f'A{row}'].fill = styles['title_fill']
    ws[f'A{row}'].alignment = styles['center_align']
    ws[f'A{row}'].border = styles['thin_border']
    row += 2
    
    # Detection rules
    rules = data.get('detection_rules', [])
    
    if not rules:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        ws.cell(row=row, column=1, value='No detection rules provided').border = styles['thin_border']
        row += 2
    else:
        for rule in rules:
            rule_type = rule.get('type', 'SIGMA')
            
            ws[f'A{row}'] = f"{rule_type} RULE"
            ws[f'A{row}'].font = styles['header_font']
            ws[f'A{row}'].fill = styles['header_fill']
            ws[f'A{row}'].border = styles['thin_border']
            ws.merge_cells(f'A{row}:C{row}')
            row += 1
            
            ws[f'A{row}'] = rule.get('name', 'Unnamed Rule')
            ws[f'A{row}'].font = styles['bold_font']
            ws.merge_cells(f'A{row}:C{row}')
            ws[f'A{row}'].border = styles['thin_border']
            row += 1
            
            ws[f'A{row}'] = rule.get('content', '')
            ws[f'A{row}'].font = Font(name='Courier New', size=10)
            ws[f'A{row}'].alignment = styles['wrap_align']
            ws.merge_cells(f'A{row}:C{row}')
            ws.row_dimensions[row].height = 100
            row += 2
    
    # Set column widths
    ws.column_dimensions['A'].width = 80
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15


def _create_recommendations_sheet(ws, data):
    """Create Recommendations sheet"""
    styles = _get_excel_styles()
    row = 1
    
    # Title
    ws.merge_cells(f'A{row}:F{row}')
    ws[f'A{row}'] = "RECOMMENDATIONS & ACTIONS"
    ws[f'A{row}'].font = styles['title_font']
    ws[f'A{row}'].fill = styles['title_fill']
    ws[f'A{row}'].alignment = styles['center_align']
    ws[f'A{row}'].border = styles['thin_border']
    row += 2
    
    # Headers
    headers = ['Priority', 'Action', 'Type', 'Owner', 'Due Date', 'Status']
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles['header_font']
        cell.fill = styles['header_fill']
        cell.border = styles['thin_border']
        cell.alignment = styles['center_align']
    row += 1
    
    # Recommendations data
    recs = data.get('recommendations', [])
    
    if not recs:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        ws.cell(row=row, column=1, value='No recommendations provided').border = styles['thin_border']
        row += 1
    else:
        for rec in recs:
            priority = rec.get('priority', 'Medium')
            cell_priority = ws.cell(row=row, column=1, value=priority)
            cell_priority.border = styles['thin_border']
            
            if 'Immediate' in priority or 'Critical' in priority:
                cell_priority.fill = PatternFill(start_color='DC2626', end_color='DC2626', fill_type='solid')
                cell_priority.font = Font(color='FFFFFF', bold=True)
            elif 'Short' in priority or 'High' in priority:
                cell_priority.fill = PatternFill(start_color='F59E0B', end_color='F59E0B', fill_type='solid')
                cell_priority.font = Font(color='212529', bold=True)
            else:
                cell_priority.fill = PatternFill(start_color='6B7280', end_color='6B7280', fill_type='solid')
                cell_priority.font = Font(color='FFFFFF', bold=True)
            
            ws.cell(row=row, column=2, value=rec.get('action', '')).border = styles['thin_border']
            ws.cell(row=row, column=3, value=rec.get('type', '')).border = styles['thin_border']
            ws.cell(row=row, column=4, value=rec.get('owner', '')).border = styles['thin_border']
            ws.cell(row=row, column=5, value=rec.get('due_date', '')).border = styles['thin_border']
            ws.cell(row=row, column=6, value=rec.get('status', 'Todo')).border = styles['thin_border']
            
            row += 1
    
    # Set column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12


def _create_sources_sheet(ws, data):
    """Create Sources & Distribution sheet"""
    styles = _get_excel_styles()
    row = 1
    
    # Title
    ws.merge_cells(f'A{row}:E{row}')
    ws[f'A{row}'] = "SOURCES & DISTRIBUTION"
    ws[f'A{row}'].font = styles['title_font']
    ws[f'A{row}'].fill = styles['title_fill']
    ws[f'A{row}'].alignment = styles['center_align']
    ws[f'A{row}'].border = styles['thin_border']
    row += 2
    
    # Sources section
    ws[f'A{row}'] = "INTELLIGENCE SOURCES"
    ws[f'A{row}'].font = styles['header_font']
    ws[f'A{row}'].fill = styles['header_fill']
    ws[f'A{row}'].border = styles['thin_border']
    ws.merge_cells(f'A{row}:E{row}')
    row += 1
    
    # Source headers
    headers = ['Source Name', 'TLP', 'Reliability', 'Date', 'URL', 'Type']
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = styles['header_font']
        cell.fill = styles['header_fill']
        cell.border = styles['thin_border']
    row += 1
    
    # Sources data
    sources = data.get('sources', [])
    if not sources:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        ws.cell(row=row, column=1, value='No sources provided').border = styles['thin_border']
        row += 1
    else:
        for source in sources:
            ws.cell(row=row, column=1, value=source.get('name', '')).border = styles['thin_border']
            ws.cell(row=row, column=2, value=source.get('tlp', '')).border = styles['thin_border']
            ws.cell(row=row, column=3, value=source.get('reliability', '')).border = styles['thin_border']
            ws.cell(row=row, column=4, value=source.get('date', '')).border = styles['thin_border']
            ws.cell(row=row, column=5, value=source.get('url', '')).border = styles['thin_border']
            ws.cell(row=row, column=6, value=source.get('type', '')).border = styles['thin_border']
            row += 1
    
    # Distribution section
    row += 1
    ws[f'A{row}'] = "DISTRIBUTION LIST"
    ws[f'A{row}'].font = styles['header_font']
    ws[f'A{row}'].fill = styles['header_fill']
    ws[f'A{row}'].border = styles['thin_border']
    ws.merge_cells(f'A{row}:E{row}')
    row += 1
    
    ws[f'A{row}'] = "Name"
    ws[f'B{row}'] = "Role"
    ws[f'C{row}'] = "Organization"
    ws[f'D{row}'] = "Email"
    for col in ['A', 'B', 'C', 'D']:
        ws[f'{col}{row}'].font = styles['header_font']
        ws[f'{col}{row}'].fill = styles['header_fill']
        ws[f'{col}{row}'].border = styles['thin_border']
    row += 1
    
    recipients = data.get('recipients', [])
    for recipient in recipients:
        ws.cell(row=row, column=1, value=recipient.get('name', '')).border = styles['thin_border']
        ws.cell(row=row, column=2, value=recipient.get('role', '')).border = styles['thin_border']
        ws.cell(row=row, column=3, value=recipient.get('organization', '')).border = styles['thin_border']
        ws.cell(row=row, column=4, value=recipient.get('email', '')).border = styles['thin_border']
        row += 1
    
    # Handling instructions
    row += 1
    ws[f'A{row}'] = "HANDLING INSTRUCTIONS"
    ws[f'A{row}'].font = styles['header_font']
    ws[f'A{row}'].fill = styles['header_fill']
    ws[f'A{row}'].border = styles['thin_border']
    ws.merge_cells(f'A{row}:E{row}')
    row += 1
    
    ws[f'A{row}'] = data.get('distribution_handling', 'Standard TLP handling applies')
    ws.merge_cells(f'A{row}:E{row}')
    ws[f'A{row}'].alignment = styles['wrap_align']
    ws[f'A{row}'].border = styles['thin_border']
    row += 2

    row = _write_section_header(ws, row, "FEEDBACK & DISCLAIMER", styles, cols=5)
    row = _write_field_row(ws, row, 'Feedback Contact', data.get('feedback_contact'), styles)
    row = _write_field_row(ws, row, 'Disclaimer', data.get('disclaimer'), styles)
    
    # Set column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 15


def _get_excel_styles():
    """Return common Excel styles"""
    violet_header = '6D28D9'
    return {
        'header_font': Font(name='Calibri', size=11, bold=True, color='FFFFFF'),
        'title_font': Font(name='Calibri', size=16, bold=True, color='4C1D95'),
        'normal_font': Font(name='Calibri', size=11, color='1E1033'),
        'bold_font': Font(name='Calibri', size=11, bold=True, color='1E1033'),
        'header_fill': PatternFill(start_color=violet_header, end_color=violet_header, fill_type='solid'),
        'title_fill': PatternFill(start_color='EDE9FE', end_color='EDE9FE', fill_type='solid'),
        'thin_border': Border(
            left=Side(style='thin', color='C4B5FD'),
            right=Side(style='thin', color='C4B5FD'),
            top=Side(style='thin', color='C4B5FD'),
            bottom=Side(style='thin', color='C4B5FD'),
        ),
        'medium_border': Border(
            left=Side(style='medium', color='7C3AED'),
            right=Side(style='medium', color='7C3AED'),
            top=Side(style='medium', color='7C3AED'),
            bottom=Side(style='medium', color='7C3AED'),
        ),
        'center_align': Alignment(horizontal='center', vertical='center'),
        'wrap_align': Alignment(horizontal='left', vertical='center', wrap_text=True),
    }
